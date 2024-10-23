from rest_framework import filters, status, viewsets
from rest_framework.decorators import api_view, permission_classes
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.parsers import JSONParser
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth.hashers import make_password
from django.db import transaction, IntegrityError
from django.db.models import Sum, Q, IntegerField, ExpressionWrapper, Count, Value, F, Prefetch
from django.db.models.functions import Coalesce, ExtractWeekDay, ExtractHour, ExtractMonth, TruncDate, TruncHour
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.validators import URLValidator
from django.core.files.storage import default_storage
from django.utils import timezone
from django.shortcuts import render
from django_filters.rest_framework import DjangoFilterBackend
from datetime import datetime, timedelta, time
from collections import defaultdict
from googletrans import Translator
from operator import itemgetter
from collections import OrderedDict
from order import order_helper
from core.models import (
    Vendor, Product, ProductCategory, ProductCategoryJoint, ProductImage, ProductAndModifierGroupJoint,
    ProductModifier, ProductModifierGroup, Platform, ProductModifierAndModifierGroupJoint, Tax
)
from order.models import (
    Order, OrderPayment, Customer, Address, LoyaltyProgramSettings, LoyaltyPointsCreditHistory, LoyaltyPointsRedeemHistory,
    Order_Discount, SplitOrderItem,
)
from pos.models import (
    StoreTiming, Banner, POSSetting, Department, CoreUserCategory, WorkingShift, CoreUser, POSPermission, CashRegister,
)
from koms.models import Order_tables, Order_content, Order as KOMSOrder, Order_modifer, Station, Staff
from kiosk.models import KioskOrderData
from woms.models import HotelTable, Waiter, Floor
from koms.serializers.content_history_serializer import Content_history_serializer
from kiosk.serializer import KiosK_create_order_serializer
from pos.serializers import (
    StoreTimingSerializer, WaiterSerializer, FloorSerializer, HotelTableSerializer, ProductCategorySerializer, ProductSerializer,
    ProductCategoryJointSerializer, ProductImagesSerializer, ProductModGroupJointSerializer, ModifierGroupSerializer,
    ModifierSerializer, DiscountCouponModelSerializer, StationModelSerializer, ChefModelSerializer, BannerModelSerializer,
    DepartmentModelSerializer, WorkingShiftModelSerializer, CoreUserModelSerializer, POSPermissionModelSerializer,
)
from woms.views import get_table_data, filter_tables
from koms.views import notify, allStationWiseCategory, allStationWiseRemove, allStationWiseSingle, waiteOrderUpdate, webSocketPush
from pos.filters import WaiterFilter, ChefFilter
from core.excel_file_upload import process_excel
from pos.utils import (
    order_count, get_product_by_category_data, get_product_data, get_modifier_data, process_product_excel,
    get_department_wise_categories, get_order_info_for_socket#, update_images, update_categories,
)
from inventory.utils import (
    single_category_sync_with_odoo, delete_category_in_odoo, single_product_sync_with_odoo,
    delete_product_in_odoo, single_modifier_group_sync_with_odoo, delete_modifier_group_in_odoo,
    single_modifier_sync_with_odoo, delete_modifier_in_odoo, sync_order_content_with_inventory,
)
from pos.language import (
    local_timezone, language_localization, payment_type_english, payment_status_english, order_type_english,
    master_order_status_number, koms_order_status_english, order_type_number, payment_type_number,
    check_key_exists, table_created_locale, table_deleted_locale,
)
import pandas
import openpyxl
import threading
import calendar
import pgeocode
import socket
import json
import re
import os



class CustomPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 50

    def get_paginated_response(self, data):
        return Response({
            'total_pages': self.page.paginator.num_pages if self.page.paginator.count > 0 else 1,
            'current_page': self.page.number,
            'page_size': self.page.paginator.per_page,
            'results': data
        })


class DepartmentModelViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all().order_by('-pk')
    serializer_class = DepartmentModelSerializer
    filter_backends = (DjangoFilterBackend, SearchFilter, OrderingFilter)
    filterset_fields = ('is_active',)
    search_fields = ('name', 'name_locale')
    ordering_fields = ('id', 'name',)
    # permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        vendor_id = self.request.GET.get('vendor')

        if vendor_id:
            return Department.objects.filter(vendor = vendor_id).order_by('-pk')
        
        return Department.objects.none()

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        serializer = self.get_serializer(queryset, many=True)
        
        return JsonResponse({"departments": serializer.data})


class WorkingShiftModelViewSet(viewsets.ModelViewSet):
    queryset = WorkingShift.objects.all().order_by('start_time')
    serializer_class = WorkingShiftModelSerializer
    filter_backends = (DjangoFilterBackend, SearchFilter, OrderingFilter)
    search_fields = ('name', 'name_locale')
    ordering_fields = ('id', 'start_time')
    # permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        vendor_id = self.request.GET.get('vendor')

        if vendor_id:
            return WorkingShift.objects.filter(vendor = vendor_id).order_by('start_time')
        
        return WorkingShift.objects.none()

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        serializer = self.get_serializer(queryset, many=True)
        
        return JsonResponse({"working_shifts": serializer.data})


class CoreUserModelViewSet(viewsets.ModelViewSet):
    queryset = CoreUser.objects.all().order_by('-pk')
    serializer_class = CoreUserModelSerializer
    filter_backends = (DjangoFilterBackend, SearchFilter, OrderingFilter)
    filterset_fields = ('is_head', 'core_user_category',)
    search_fields = ('first_name', 'last_name', 'email', 'phone_number',)
    ordering_fields = ('id', 'first_name', 'last_name',)
    # permission_classes = [IsAuthenticated]

    def get_queryset(self):
        vendor_id = self.request.GET.get('vendor')
        user_category_id = self.request.GET.get('user_category')

        queryset = CoreUser.objects.none()

        if vendor_id:
            queryset = CoreUser.objects.filter(vendor = vendor_id)

            if user_category_id:
                if user_category_id == '0':
                    user_category_id = None
                
                queryset = queryset.filter(core_user_category = user_category_id)
                
            queryset = queryset.order_by('-pk')

        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        serializer = self.get_serializer(queryset, many=True)
        
        return JsonResponse({"users": serializer.data})
    
    def perform_create(self, serializer):
        password = self.request.data.get('password')

        if password:
            serializer.save(password = make_password(password))
        else:
            serializer.save()

    def perform_update(self, serializer):
        password = self.request.data.get('password')

        if password:
            serializer.save(password = make_password(password))

        else:
            serializer.save()


class WaiterViewSet(viewsets.ModelViewSet):
    queryset = Waiter.objects.all()
    serializer_class = WaiterSerializer
    filter_class = WaiterFilter
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]

    def get_queryset(self):
        vendor_id = self.request.GET.get('vendorId', None)

        if vendor_id:
            queryset = Waiter.objects.filter(vendorId=vendor_id)

            return queryset
        
        else:
            return Waiter.objects.none()

    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # name_query = request.query_params.get('name', None)
        name_query = request.GET.get('name')
        language = request.GET.get('language', 'English')

        if name_query:
            if language == "English":
                queryset = queryset.filter(name__icontains=name_query)

            else:
                queryset = queryset.filter(name_locale__icontains=name_query)

        serializer = self.get_serializer(queryset, many=True)
        data = {"waiters": serializer.data}
        
        return Response(data, status=status.HTTP_200_OK)


class FloorViewSet(viewsets.ModelViewSet):
    queryset = Floor.objects.all()
    serializer_class = FloorSerializer
    # permission_classes = [IsAuthenticated]

    def get_queryset(self):
        vendor_id = self.request.GET.get('vendorId', None)

        if vendor_id:
            return Floor.objects.filter(vendorId=vendor_id).order_by('id')
        
        return Floor.objects.none()
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        serializer = self.get_serializer(queryset, many=True)
        
        return JsonResponse({"floors": serializer.data})
    

class HotelTableViewSet(viewsets.ModelViewSet):
    queryset = HotelTable.objects.all()
    serializer_class = HotelTableSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_fields = ('floor',)
    # permission_classes = [IsAuthenticated]

    def get_queryset(self):
        vendor_id = self.request.GET.get('vendorId', None)

        if vendor_id:
            return HotelTable.objects.filter(vendorId = vendor_id).order_by('tableNumber')

        return HotelTable.objects.none()
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        serializer = self.get_serializer(queryset, many=True)
        
        return JsonResponse({"tables": serializer.data})
    
    def perform_create(self, serializer):
        instance = serializer.save()

        serialized_data = self.get_serializer(instance).data

        table_number = serialized_data.get('tableNumber')
        floor_name = serialized_data.get('floor')
        vendor_id = serialized_data.get('vendorId')

        language = self.request.GET.get('language', 'English')

        if language == "English":
            notify(
                type = 3,
                msg = '0',
                desc = f"Table no.{table_number} created on {floor_name}",
                stn = ['POS'],
                vendorId = instance.vendorId.pk
            )

        else:
            notify(
                type = 3,
                msg = '0',
                desc = table_created_locale(table_number, floor_name),
                stn = ['POS'],
                vendorId = instance.vendorId.pk
            )

        table_data = get_table_data(hotelTable=instance, vendorId=vendor_id)
        
        webSocketPush(
            message = {"result": table_data, "UPDATE": "UPDATE"},
            room_name = f"WOMSPOS------{language}-{str(vendor_id)}",
            username = "CORE",
        )

        vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

        waiter_heads = Waiter.objects.filter(is_waiter_head=True, vendorId=vendor_id)

        if waiter_heads:
            for head in waiter_heads:
                if language == "English":
                    notify(
                        type = 3,
                        msg = '0',
                        desc = f"Table no.{table_number} created on {floor_name}",
                        stn = [f'WOMS{head.pk}'],
                        vendorId = vendor_id
                    )

                else:
                    notify(
                        type = 3,
                        msg = '0',
                        desc = table_created_locale(table_number, floor_name),
                        stn = [f'WOMS{head.pk}'],
                        vendorId = vendor_id
                    )

                webSocketPush(
                    message = {"result": table_data, "UPDATE": "UPDATE"},
                    room_name = f"WOMS{str(head.pk)}------English-{str(vendor_id)}",
                    username = "CORE",
                )
                
                if vendor_instance.secondary_language and (language != "English"):
                    webSocketPush(
                        message = {"result": table_data, "UPDATE": "UPDATE"},
                        room_name = f"WOMS{str(head.pk)}------{language}-{str(vendor_id)}",
                        username = "CORE",
                    )

    def perform_destroy(self, instance):
        instance.delete()

        vendor_id = instance.vendorId.pk

        language = self.request.GET.get('language', 'English')

        if language == "English":
            notify(
                type = 3,
                msg = '0',
                desc = f"Table no.{instance.tableNumber} deleted on {instance.floor.name}",
                stn = ['POS'],
                vendorId = vendor_id
            )

        else:
            notify(
                type = 3,
                msg = '0',
                desc = table_deleted_locale(instance.tableNumber, instance.floor.name),
                stn = ['POS'],
                vendorId = vendor_id
            )

        all_tables_data = filter_tables("POS", "All", "All", "All", "All", instance.floor.pk, vendor_id, language=language)
        
        webSocketPush(
            message = {"result": all_tables_data, "UPDATE": "UPDATE"},
            room_name = f"WOMSPOS------{language}-{str(vendor_id)}",
            username = "CORE",
        )
        
        vendor_instance = Vendor.objects.filter(pk=vendor_id).first()
        
        waiter_heads = Waiter.objects.filter(is_waiter_head=True, vendorId=vendor_id)

        if waiter_heads:
            for head in waiter_heads:
                if language == "English":
                    notify(
                        type = 3,
                        msg = '0',
                        desc = f"Table no.{instance.tableNumber} deleted on {instance.floor.name}",
                        stn = [f'WOMS{head.pk}'],
                        vendorId = vendor_id
                    )

                else:
                    notify(
                        type = 3,
                        msg = '0',
                        desc = table_deleted_locale(instance.tableNumber, instance.floor.name),
                        stn = [f'WOMS{head.pk}'],
                        vendorId = vendor_id
                    )

                webSocketPush(
                    message = {"result": all_tables_data, "UPDATE": "UPDATE"},
                    room_name = f"WOMS{str(head.pk)}------English-{str(vendor_id)}",
                    username = "CORE",
                )
                
                if vendor_instance.secondary_language and (language != "English"):
                    webSocketPush(
                        message = {"result": all_tables_data, "UPDATE": "UPDATE"},
                        room_name = f"WOMS{str(head.pk)}------{language}-{str(vendor_id)}",
                        username = "CORE",
                    )
  

class ProductCategoryViewSet(viewsets.ModelViewSet):
    queryset = ProductCategory.objects.all()
    serializer_class = ProductCategorySerializer
    filter_backends = (DjangoFilterBackend, SearchFilter, OrderingFilter)
    search_fields = ('categoryName', 'categoryName_locale',)
    pagination_class = CustomPagination
    # permission_classes = [IsAuthenticated]

    def get_queryset(self):
        vendor_id = self.request.GET.get('vendorId', None)

        if vendor_id:
            return ProductCategory.objects.filter(vendorId = vendor_id).order_by('-pk')
        
        return ProductCategory.objects.none()
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)

        serializer = self.get_serializer(page, many=True)

        return self.get_paginated_response(serializer.data)
    
    def create(self, request, *args, **kwargs):
        try:
            plu = request.data.get('categoryPLU')
            vendor_id = request.data.get('vendorId')

            existing_category = ProductCategory.objects.filter(categoryPLU=plu, vendorId=vendor_id).first()

            if existing_category:
                return Response({'error': 'Category with this PLU already exists.'}, status=status.HTTP_400_BAD_REQUEST)

            try:
                image_url = request.data.get('categoryImageUrl')

                if image_url:
                    validator = URLValidator()

                    validator(image_url)

            except Exception as e:
                return Response({'error': 'Invalid Image URL'}, status=status.HTTP_400_BAD_REQUEST)
            
            with transaction.atomic():
                serializer = self.get_serializer(data=request.data)
                serializer.is_valid(raise_exception=True)
                self.perform_create(serializer)

                inventory_platform = Platform.objects.filter(Name="Inventory", isActive=True, VendorId=vendor_id).first()
                
                if inventory_platform:
                    sync_status = single_category_sync_with_odoo(serializer.instance)
                        
                    if sync_status == 0:
                        notify(type=3, msg='0', desc='Category did not synced with Inventory', stn=['POS'], vendorId=vendor_id)
                    
                    else:
                        notify(type=3, msg='0', desc='Category synced with Inventory', stn=['POS'], vendorId=vendor_id)
                
                headers = self.get_success_headers(serializer.data)

                return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
            
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def update(self, request, *args, **kwargs):
        try:
            with transaction.atomic():
                partial = kwargs.pop('partial', False)
                instance = self.get_object()

                new_plu = request.data.get('categoryPLU')

                if new_plu != instance.categoryPLU:
                    existing_category = ProductCategory.objects.filter(categoryPLU=new_plu, vendorId=instance.vendorId).first()

                    if existing_category:
                        return Response({'error': 'Category with this PLU already exists.'}, status=status.HTTP_400_BAD_REQUEST)
                    
                try:
                    image_url = request.data.get('categoryImageUrl')

                    if image_url:
                        validator = URLValidator()

                        validator(image_url)

                except Exception as e:
                    return Response({'error': 'Invalid Image URL'}, status=status.HTTP_400_BAD_REQUEST)

                serializer = self.get_serializer(instance, data=request.data, partial=partial)
                serializer.is_valid(raise_exception=True)
                self.perform_update(serializer)

                vendor_id = serializer.instance.vendorId.pk
                
                inventory_platform = Platform.objects.filter(Name="Inventory", isActive=True, VendorId=vendor_id).first()
                
                if inventory_platform:
                    sync_status = single_category_sync_with_odoo(serializer.instance)
                        
                    if sync_status == 0:
                        notify(type=3, msg='0', desc='Category did not synced with Inventory', stn=['POS'], vendorId=vendor_id)
                    
                    else:
                        notify(type=3, msg='0', desc='Category synced with Inventory', stn=['POS'], vendorId=vendor_id)
                
                return Response(serializer.data, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def destroy(self, request, *args, **kwargs):
        try:
            with transaction.atomic():
                instance = self.get_object()

                vendor_id = instance.vendorId.pk
                
                inventory_platform = Platform.objects.filter(Name="Inventory", isActive=True, VendorId=vendor_id).first()

                if inventory_platform:
                    delete_status, error_message, request_data = delete_category_in_odoo(inventory_platform.baseUrl, instance.categoryPLU, vendor_id)
                
                    if delete_status == 0:
                        notify(type=3, msg='0', desc='Category did not synced with Inventory', stn=['POS'], vendorId=vendor_id)
                    
                    else:
                        notify(type=3, msg='0', desc='Category synced with Inventory', stn=['POS'], vendorId=vendor_id)
                
                self.perform_destroy(instance)
                return Response(status=status.HTTP_204_NO_CONTENT)
                    
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ModifierGroupViewSet(viewsets.ModelViewSet):
    queryset = ProductModifierGroup.objects.all()
    serializer_class = ModifierGroupSerializer
    filter_backends = (DjangoFilterBackend, SearchFilter, OrderingFilter)
    filterset_fields = ('active',)
    search_fields = ('name', 'name_locale')
    pagination_class = CustomPagination
    # permission_classes = [IsAuthenticated]

    def get_queryset(self):
        vendor_id = self.request.GET.get('vendorId', None)

        if vendor_id:
            return ProductModifierGroup.objects.filter(vendorId = vendor_id).order_by('-id')
        
        return ProductModifierGroup.objects.none()
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)

        serializer = self.get_serializer(page, many=True)

        return self.get_paginated_response(serializer.data)
        
    def create(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)

            plu = data.get('PLU')
            vendor_id = data.get('vendorId')

            existing_modifier_group = ProductModifierGroup.objects.filter(PLU=plu, vendorId=vendor_id).first()

            if existing_modifier_group:
                return Response({'error': 'Modifier group with this PLU already exists.'}, status=status.HTTP_400_BAD_REQUEST)

            with transaction.atomic():
                serializer = self.get_serializer(data=request.data)
                serializer.is_valid(raise_exception=True)
                self.perform_create(serializer)

                inventory_platform = Platform.objects.filter(Name="Inventory", isActive=True, VendorId=vendor_id).first()
                
                if inventory_platform:
                    sync_status = single_modifier_group_sync_with_odoo(serializer.instance)
                        
                    if sync_status == 0:
                        notify(type=3, msg='0', desc='Modifier group did not synced with Inventory', stn=['POS'], vendorId=vendor_id)
                    
                    else:
                        notify(type=3, msg='0', desc='Modifier group synced with Inventory', stn=['POS'], vendorId=vendor_id)
                
                headers = self.get_success_headers(serializer.data)
                return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
        
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    def update(self, request, *args, **kwargs):
        try:
            with transaction.atomic():
                partial = kwargs.pop('partial', False)
                instance = self.get_object()

                new_plu = request.data.get('PLU')

                if new_plu != instance.PLU:
                    existing_category = ProductModifierGroup.objects.filter(PLU=new_plu, vendorId=instance.vendorId).first()

                    if existing_category:
                        return Response({'error': 'Modifier group with this PLU already exists.'}, status=status.HTTP_400_BAD_REQUEST)
                    
                serializer = self.get_serializer(instance, data=request.data, partial=partial)
                serializer.is_valid(raise_exception=True)
                self.perform_update(serializer)

                vendor_id = instance.vendorId.pk
                
                inventory_platform = Platform.objects.filter(Name="Inventory", isActive=True, VendorId=vendor_id).first()
                
                if inventory_platform:
                    sync_status = single_modifier_group_sync_with_odoo(serializer.instance)
                        
                    if sync_status == 0:
                        notify(type=3, msg='0', desc='Modifier group did not synced with Inventory', stn=['POS'], vendorId=vendor_id)
                    
                    else:
                        notify(type=3, msg='0', desc='Modifier group synced with Inventory', stn=['POS'], vendorId=vendor_id)
                
                return Response(serializer.data)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def destroy(self, request, *args, **kwargs):
        try:
            with transaction.atomic():
                instance = self.get_object()
                
                vendor_id = instance.vendorId.pk
                
                inventory_platform = Platform.objects.filter(Name="Inventory", isActive=True, VendorId=vendor_id).first()
                
                if inventory_platform:
                    sync_status = delete_modifier_group_in_odoo(inventory_platform.baseUrl, instance.PLU, vendor_id)
                        
                    if sync_status == 0:
                        notify(type=3, msg='0', desc='Modifier group did not synced with Inventory', stn=['POS'], vendorId=vendor_id)
                    
                    else:
                        notify(type=3, msg='0', desc='Modifier group synced with Inventory', stn=['POS'], vendorId=vendor_id)
                
                self.perform_destroy(instance)
                return Response(status=status.HTTP_204_NO_CONTENT)
                
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class DiscountCouponModelViewSet(viewsets.ModelViewSet):
    queryset = Order_Discount.objects.all().order_by('-pk')
    serializer_class = DiscountCouponModelSerializer
    filter_backends = (DjangoFilterBackend, SearchFilter, OrderingFilter)
    filterset_fields = ('is_active',)
    search_fields = ("discountName", "discountCode", "value")
    pagination_class = CustomPagination
    # permission_classes = [IsAuthenticated]

    def get_queryset(self):
        vendor_id = self.request.GET.get('vendorId', None)

        if vendor_id:
            return Order_Discount.objects.filter(vendorId = vendor_id).order_by("-pk")
        
        return Order_Discount.objects.none()

    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)

        serializer = self.get_serializer(page, many=True)

        return self.get_paginated_response(serializer.data)


class StationModelViewSet(viewsets.ModelViewSet):
    queryset = Station.objects.all()
    serializer_class = StationModelSerializer
    filter_backends = (DjangoFilterBackend, SearchFilter, OrderingFilter)
    search_fields = ('station_name', 'station_name_locale',)
    # permission_classes = [IsAuthenticated]

    def get_queryset(self):
        vendor_id = self.request.GET.get('vendorId', None)

        if vendor_id:
            return Station.objects.filter(vendorId = vendor_id).order_by("-pk")
        
        return Station.objects.none()

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        serializer = self.get_serializer(queryset, many=True)

        return JsonResponse({"stations": serializer.data})


class ChefModelViewSet(viewsets.ModelViewSet):
    queryset = Staff.objects.all().order_by('-pk')
    serializer_class = ChefModelSerializer
    filter_class = ChefFilter
    filter_backends = (DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter)
    search_fields = ('first_name', 'last_name',)
    pagination_class = CustomPagination

    def get_queryset(self):
        vendor_id = self.request.GET.get('vendorId', None)

        if vendor_id:
            return Staff.objects.filter(vendorId=vendor_id).order_by("-pk")
        
        return Staff.objects.none()

    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)

        if page:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class BannerModelViewSet(viewsets.ModelViewSet):
    queryset = Banner.objects.all()
    serializer_class = BannerModelSerializer
    # permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        vendor_id = self.request.GET.get('vendorId')

        if vendor_id:
            platform_type = self.request.GET.get('platform_type')
            
            if platform_type:
                return Banner.objects.filter(platform_type = platform_type, vendor = vendor_id).order_by("-pk")
            
            return Banner.objects.filter(vendor = vendor_id).order_by("-pk")
        
        return Banner.objects.none()
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        serializer = self.get_serializer(queryset, many=True)

        return JsonResponse({"banners": serializer.data})


class POSPermissionModelViewSet(viewsets.ModelViewSet):
    queryset = POSPermission.objects.all()
    serializer_class = POSPermissionModelSerializer
    # permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        vendor_id = self.request.GET.get('vendor')

        if vendor_id:
            return POSPermission.objects.filter(vendor = vendor_id).order_by("-pk")
        
        return POSPermission.objects.none()
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        serializer = self.get_serializer(queryset, many=True)

        return JsonResponse({"permissions": serializer.data})



@api_view(["GET"])
def get_tax(request):
    platform = request.GET.get("platform")
    language = request.GET.get("language", "English")
    vendor_id = request.GET.get("vendorId")

    if not vendor_id:
        return Response("Vendor ID empty", status = status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id = int(vendor_id)

    except ValueError:
        return Response("Invalid vendor ID", status = status.HTTP_400_BAD_REQUEST)
    
    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor_instance:
        return Response("Vendor with given ID does not exist", status = status.HTTP_404_NOT_FOUND)

    tax_list = []

    taxes = Tax.objects.filter(vendor = vendor_id)

    if (platform == "Website") or (platform == "Mobile App"):
        for tax in taxes:
            if language == "English":
                tax_name = tax.name

            else:
                tax_name = tax.name_locale

            tax_list.append({
                "id": tax.pk,
                "type": tax_name,
                "rate": tax.percentage,
                "value": round((tax.percentage / 100), 3),
                "is_active": tax.is_active,
            })

    else:
        for tax in taxes:
            tax_list.append({
                "id": tax.pk,
                "type": tax.name,
                "type_locale": tax.name_locale,
                "rate": tax.percentage,
                "value": round((tax.percentage / 100), 3),
                "is_active": tax.is_active,
            })

    return JsonResponse({"taxes": tax_list})


@api_view(["POST"])
def create_tax(request):
    tax_type = request.data.get("type")
    rate = request.data.get("rate")
    is_active = request.data.get("is_active")
    vendor_id = request.data.get("vendor_id")
    
    try:
        if not all([tax_type, rate, vendor_id]):
            raise ValueError
    
        vendor_id = int(vendor_id)
        rate = float(rate)
        is_active = bool(is_active)
    
    except ValueError:
        return Response("Invalid request data", status = status.HTTP_400_BAD_REQUEST)
    
    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor_instance:
        return Response("Vendor with given ID does not exist", status = status.HTTP_404_NOT_FOUND)

    existing_tax = Tax.objects.filter(name = tax_type, vendor = vendor_id).first()

    if existing_tax:
        return Response("Entry already created", status = status.HTTP_409_CONFLICT)

    tax = Tax.objects.create(
        name = tax_type,
        percentage = rate,
        is_active = is_active,
        vendorId = vendor_instance
    )

    tax_info = {
        "type": tax.name,
        "rate": tax.percentage,
        "value": round((tax.percentage / 100), 3),
        "is_active": tax.is_active,
    }

    return JsonResponse(tax_info, status = status.HTTP_201_CREATED)


@api_view(["PUT"])
def update_tax(request):
    tax_id = request.data.get("id")
    tax_type = request.data.get("type")
    rate = request.data.get("rate")
    is_active = request.data.get("is_active")
    vendor_id = request.data.get("vendor_id")
    
    try:
        if not all((tax_id, tax_type, rate, vendor_id)):
            raise ValueError
    
        tax_id = int(tax_id)
        vendor_id = int(vendor_id)
        rate = float(rate)
        is_active = bool(is_active)
    
    except ValueError:
        return Response("Invalid request data", status=status.HTTP_400_BAD_REQUEST)
    
    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor_instance:
        return Response("Vendor with given ID does not exist", status=status.HTTP_404_NOT_FOUND)

    tax = Tax.objects.filter(pk=tax_id, vendorId=vendor_id).first()

    if not tax:
        return Response("No record found", status=status.HTTP_404_NOT_FOUND)
    
    tax.name = tax_type
    tax.percentage = rate
    tax.is_active = is_active

    tax.save()

    return JsonResponse({
        "type": tax.name,
        "rate": tax.percentage,
        "value": round((tax.percentage / 100), 3),
        "is_active": tax.is_active,
    })


@api_view(["DELETE"])
def delete_tax(request):
    tax_id = request.data.get("id")
    vendor_id = request.data.get("vendor_id")
    
    try:
        if not all((tax_id, vendor_id)):
            raise ValueError
    
        tax_id = int(tax_id)
        vendor_id = int(vendor_id)
    
    except ValueError:
        return Response("Invalid request data", status = status.HTTP_400_BAD_REQUEST)
    
    vendor_instance = Vendor.objects.filter(pk = vendor_id).first()

    if not vendor_instance:
        return Response("Vendor with given ID does not exist", status = status.HTTP_404_NOT_FOUND)

    tax = Tax.objects.filter(pk = tax_id, vendorId = vendor_id).first()

    if not tax:
        return Response("No record found", status = status.HTTP_404_NOT_FOUND)
    
    tax.delete()

    return Response(status = status.HTTP_204_NO_CONTENT)


@api_view(["POST"])
def pos_user_login(request):
    try:
        username = request.data.get("username")
        password = request.data.get("password")

        if not username or not password:
            return JsonResponse({"message": "Invalid request data"}, status = status.HTTP_400_BAD_REQUEST)

        user = authenticate(request, username = username, password = password)
        
        if not user:
            return JsonResponse({"message": "Invalid login credentials"}, status = status.HTTP_400_BAD_REQUEST)

        if user.is_active == False:
            return JsonResponse({"message": "User not active"}, status = status.HTTP_400_BAD_REQUEST)
        
        pos_user_info = CoreUser.objects.filter(username = username, password = user.password) \
            .select_related('vendor', 'core_user_category', 'core_user_category__department') \
            .values(
                "pk", "first_name", "last_name", "vendor__pk", "core_user_category__pk", "core_user_category__is_active",
                "core_user_category__department__pk", "core_user_category__department__is_active"
            ).first()
        
        vendor_id = pos_user_info["vendor__pk"]

        vendor_info = Vendor.objects.filter(pk = vendor_id, is_active = True).values(
            "is_franchise_owner", "primary_language", "secondary_language", "currency", "currency_symbol"
        ).first()

        if not vendor_info:
            return JsonResponse({"message": "Invalid Vendor for the user or Vendor not active"}, status = status.HTTP_400_BAD_REQUEST)

        platforms = Platform.objects.filter(isActive = True, VendorId = vendor_id) \
            .filter(Q(Name = "WOMS") | Q(Name = "POS", expiryDate__date__gte = timezone.now().date())) \
            .values_list("Name", "pk")

        platforms = dict(platforms)

        pos_id = platforms.get("POS")

        if not pos_id:
            return JsonResponse({"message": "Contact your administrator to activate POS"}, status = status.HTTP_400_BAD_REQUEST)

        if not pos_user_info["core_user_category__pk"] or pos_user_info["core_user_category__is_active"] == False:
            return JsonResponse({"message": "User role not configured or not active"}, status = status.HTTP_400_BAD_REQUEST)
            
        if not pos_user_info["core_user_category__department__pk"] or pos_user_info["core_user_category__department__is_active"] == False:
            return JsonResponse({"message": "User department not configured or not active"}, status = status.HTTP_400_BAD_REQUEST)

        pos_permissions = POSPermission.objects.filter(
            core_user_category = pos_user_info["core_user_category__pk"], 
            vendor = vendor_id
        ).values(
            "show_dashboard", "show_tables_page", "show_place_order_page", "show_order_history_page", "show_product_menu",
            "show_store_time_setting", "show_tax_setting", "show_delivery_charge_setting", "show_loyalty_points_setting",
            "show_cash_register_setting", "show_customer_setting", "show_printer_setting", "show_payment_machine_setting",
            "show_banner_setting", "show_excel_file_setting", "show_employee_setting", "show_reports", "show_sop", "show_language_setting"
        ).first()

        if not pos_permissions:
            return JsonResponse({"message": "POS permissions not specified for the user"}, status = status.HTTP_400_BAD_REQUEST)

        related_vendor_list = []
        
        if vendor_info["is_franchise_owner"] == True:
            related_vendor_list = list(Vendor.objects.filter(franchise = vendor_id).values("pk", "is_active", "franchise_location"))

        login(request, user)

        # JWT Authentication
        # refresh = RefreshToken.for_user(user)

        # access_token = str(refresh.access_token)
        
        return JsonResponse({
            "message": "",
            # "access_token": access_token,
            # "refresh_token": str(refresh),
            "user_id": pos_user_info["pk"],
            "first_name": pos_user_info["first_name"],
            "last_name": pos_user_info["last_name"],
            "pos_platform_id": pos_id,
            "woms_platform_id": platforms.get("WOMS", 0),
            "primary_language": vendor_info["primary_language"],
            "secondary_language": vendor_info["secondary_language"] if vendor_info["secondary_language"] else "",
            "currency": vendor_info["currency"],
            "currency_symbol": vendor_info["currency_symbol"],
            "vendor_id": vendor_id,
            "is_franchise_owner": vendor_info["is_franchise_owner"],
            "permissions": pos_permissions,
            "related_vendors": related_vendor_list,
        }, status = status.HTTP_200_OK)
    
    except Exception as e:
        return JsonResponse({"message": str(e)}, status = status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def get_pos_permissions(request):
    vendor_id = request.data.get("vendor")
    user_id = request.data.get("user")

    try:
        vendor_id = int(vendor_id)
        user_id = int(user_id)

        if vendor_id <= 0 or user_id <= 0:
            return JsonResponse({"message": "Invalid Vendor ID or User ID"}, status = status.HTTP_400_BAD_REQUEST)
    
    except:
        return JsonResponse({"message": "Invalid Vendor ID or User ID"}, status = status.HTTP_400_BAD_REQUEST)

    vendor_instance_id = Vendor.objects.filter(pk = vendor_id, is_active = True).values_list("pk", flat = True).first()

    if not vendor_instance_id:
        return JsonResponse({"message": "Vendor not found"}, status = status.HTTP_400_BAD_REQUEST)
    
    pos_id = Platform.objects.filter(
        Name = "POS", isActive = True, expiryDate__date__gte = timezone.now().date(), VendorId = vendor_id
    ).values_list("pk", flat = True).first()

    if not pos_id:
        return JsonResponse({"message": "Contact your administrator to activate POS"}, status = status.HTTP_400_BAD_REQUEST)

    pos_user_info = CoreUser.objects.filter(pk = user_id, is_active = True) \
        .select_related('core_user_category', 'core_user_category__department') \
        .values(
            "pk", "core_user_category__pk", "core_user_category__is_active",
            "core_user_category__department__pk", "core_user_category__department__is_active"
        ).first()

    if not pos_user_info:
        return JsonResponse({"message": "User not found or not active"}, status = status.HTTP_400_BAD_REQUEST)
    
    if not pos_user_info["core_user_category__pk"] or pos_user_info["core_user_category__is_active"] == False:
        return JsonResponse({"message": "User role not configured or not active"}, status = status.HTTP_400_BAD_REQUEST)
        
    if not pos_user_info["core_user_category__department__pk"] or pos_user_info["core_user_category__department__is_active"] == False:
        return JsonResponse({"message": "User department not configured or not active"}, status = status.HTTP_400_BAD_REQUEST)

    pos_permissions = POSPermission.objects.filter(
        core_user_category = pos_user_info["core_user_category__pk"], vendor = vendor_id
    ).values(
        "show_dashboard", "show_tables_page", "show_place_order_page", "show_order_history_page", "show_product_menu",
        "show_store_time_setting", "show_tax_setting", "show_delivery_charge_setting", "show_loyalty_points_setting",
        "show_cash_register_setting", "show_customer_setting", "show_printer_setting", "show_payment_machine_setting",
        "show_banner_setting", "show_excel_file_setting", "show_employee_setting", "show_reports", "show_sop", "show_language_setting"
    ).first()

    if not pos_permissions:
        return JsonResponse({"message": "POS permissions not specified for the user"}, status = status.HTTP_400_BAD_REQUEST)
    
    return JsonResponse({"message": "", "permissions": pos_permissions}, status = status.HTTP_200_OK)


@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def dashboard(request):
    try:
        vendor_id = request.GET.get("vendor")
        language = request.GET.get("language", "English")
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        get_all_vendor_data = request.GET.get("get_all_vendor_data")

        try:
            vendor_id = int(vendor_id)

            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

            if vendor_id <= 0 or not language or start_date > end_date:
                return Response("Invalid request data", status = status.HTTP_400_BAD_REQUEST)

            get_all_vendor_data = {'True': True, 'False': False}.get(get_all_vendor_data.capitalize())

        except:
            return Response("Invalid request data", status = status.HTTP_400_BAD_REQUEST)
        
        vendor_info = Vendor.objects.filter(pk = vendor_id).values("pk", "is_franchise_owner").first()

        if not vendor_info:
            return Response("Vendor not found", status = status.HTTP_400_BAD_REQUEST)

        vendor_ids = (vendor_id,)
        
        if get_all_vendor_data == True and vendor_info["is_franchise_owner"] == True:
            vendor_ids = vendor_ids + tuple(
                Vendor.objects.filter(franchise = vendor_id).exclude(pk = vendor_id).values_list("id", flat = True)
            )

        orders = Order.objects.filter(
            OrderDate__date__range = (start_date, end_date),
            vendorId__in = vendor_ids,
            masterOrder = None
        )

        koms_orders = KOMSOrder.objects.filter(
            arrival_time__date__range = (start_date, end_date), vendorId__in = vendor_ids
        )
        
        completed_status_code = master_order_status_number["Completed"]
        canceled_status_code = master_order_status_number["Canceled"]
        inprogress_status_code = master_order_status_number["Inprogress"]
        open_status_code = master_order_status_number["Open"]
        prepared_status_code = master_order_status_number["Prepared"]
        
        order_stats = orders.aggregate(
            total_orders = Count('id'),
            total_orders_canceled = Count('id', filter = Q(Status = canceled_status_code)),
            total_orders_completed = Count('id', filter = Q(Status = completed_status_code, orderpayment__status = True)),
            total_orders_inprogress = Count('id', filter = Q(Status__in = (inprogress_status_code, open_status_code, prepared_status_code))),
            total_orders_pickedup = Count('id', filter = Q(orderType = order_type_number["Pickup"])),
            total_orders_delivered = Count('id', filter = Q(orderType = order_type_number["Delivery"])),
            total_orders_dined = Count('id', filter = Q(orderType = order_type_number["Dinein"])),
            online_orders_count = Count('id', filter = Q(platform__Name__in = ('Mobile App', 'Website'))),
            subtotal_sum = Coalesce(Sum('subtotal', filter = Q(orderpayment__status = True) & ~Q(Status = canceled_status_code)), Value(0.0)),
            discount_sum = Coalesce(Sum('discount', filter = Q(orderpayment__status = True) & ~Q(Status = canceled_status_code)), Value(0.0))
        )

        new_orders_count = koms_orders.filter(order_status = 1).count()

        onhold_orders_count = koms_orders.filter(order_status = 4).count()

        order_stats['total_orders_inprogress'] = order_stats['total_orders_inprogress'] - new_orders_count

        total_revenue = "{:.2f}".format(order_stats["subtotal_sum"] - order_stats["discount_sum"])

        active_product_count = Product.objects.filter(isDeleted = False, vendorId = vendor_id).count()

        online_order_platform_id = Platform.objects.filter(
            Name__in = ('Mobile App', 'Website'), isActive = True, VendorId = vendor_id
        ).values_list('pk', flat=True).first()
        
        sales_order_list = []
        
        if orders.exists():
            current_start_date = current_end_date = datetime.now().date()

            if start_date == end_date:
                if start_date == current_start_date and end_date == current_end_date:
                    end_datetime = datetime.now().replace(minute = 59, second = 59, microsecond = 0)
            
                elif start_date != current_start_date and end_date != current_end_date:
                    end_datetime = datetime.combine(start_date, time(23, 59, 59))

                current_datetime = datetime.combine(start_date, time(0, 0, 0))
            
                while current_datetime <= end_datetime:
                    current_hour_start = current_datetime
                    current_hour_end = current_datetime + timedelta(hours = 1)

                    filtered_orders = orders.filter(OrderDate__range = (current_hour_start, current_hour_end))

                    hour_stats = filtered_orders.aggregate(
                        total_sale = Coalesce(Sum(F('subtotal') - F('discount'), filter = Q(orderpayment__status = True) & ~Q(Status = canceled_status_code)), Value(0.0)),
                        total_orders_count = Count('id'),
                        cancelled_orders_count = Count('id', filter=Q(Status = canceled_status_code))
                    )

                    if hour_stats['total_orders_count'] > 0:
                        sales_order_list.append({
                            "date": current_hour_start.astimezone(local_timezone).strftime('%Y-%m-%d %H:%M'),
                            "total_sale": str(hour_stats["total_sale"]),
                            "total_orders_count": filtered_orders.count(),
                            "cancelled_orders_count": hour_stats["cancelled_orders_count"],
                        })

                    current_datetime = current_datetime + timedelta(hours = 1)

            else:
                for unique_date in orders.dates('OrderDate', 'day'):
                    filtered_orders = orders.filter(OrderDate__date = unique_date)

                    day_stats = filtered_orders.aggregate(
                        total_sale = Coalesce(Sum(F('subtotal') - F('discount'), filter = Q(orderpayment__status=True) & ~Q(Status = canceled_status_code)), Value(0.0)),
                        total_orders_count = Count('id'),
                        cancelled_orders_count = Count('id', filter = Q(Status = canceled_status_code))
                    )

                    if filtered_orders.count() != 0:
                        sales_order_list.append({
                            "date": unique_date.strftime("%Y-%m-%d 00:00"),
                            "total_sale": str(day_stats["total_sale"]),
                            "total_orders_count": filtered_orders.count(),
                            "cancelled_orders_count": day_stats["cancelled_orders_count"],
                        })
        
            if len(sales_order_list) == 1:
                date_value = sales_order_list[0]["date"]

                date_obj = datetime.strptime(date_value, '%Y-%m-%d %H:%M')

                if start_date == end_date:
                    date_obj = date_obj - timedelta(hours = 1)

                    date_value = date_obj.replace(minute = 0).strftime('%Y-%m-%d %H:%M')

                else:
                    date_obj = date_obj - timedelta(days = 1)

                    date_value = date_obj.replace(hour = 1, minute = 0).strftime('%Y-%m-%d %H:%M')

                sales_order_list.append({
                    "date": date_value,
                    "total_sale": "0.0",
                    "total_orders_count": 0,
                    "cancelled_orders_count": 0,
                })
        
        order_items = Order_content.objects.filter(
            orderId__order_status = 10,
            orderId__master_order__Status = completed_status_code,
            orderId__master_order__orderpayment__status = True,
            orderId__master_order__OrderDate__date__range = (start_date, end_date),
            orderId__master_order__vendorId = vendor_id,
            orderId__vendorId = vendor_id
        )
        
        total_items_sold = order_items.values('SKU').distinct().count()
        
        top_selling_items = order_items.values('SKU').annotate(quantity_sold = Sum('quantity')).order_by('-quantity_sold')[:6]

        list_of_items = []

        for item in top_selling_items:
            product_info = Product.objects.filter(PLU = item['SKU'], vendorId = vendor_id).values(
                "pk", "productName", "productName_locale", "productPrice"
            ).first()

            image_url = ProductImage.objects.filter(product = product_info["pk"], vendorId = vendor_id) \
                .values_list('url', flat = True).first()
            
            item['id'] = product_info["pk"]
            item['product_name'] = product_info["productName_locale"] if language != "English" else product_info["productName"]
            item['image'] = image_url or 'https://www.stockvault.net/data/2018/08/31/254135/preview16.jpg'
            item['price'] = product_info["productPrice"]
            item['sale'] = item['quantity_sold'] * product_info["productPrice"]

            list_of_items.append(item)

        order_details = {
            "online_order_platform_id": str(online_order_platform_id) or "", # Required for Flutter model
            "active_products": active_product_count,
            "total_sale": total_revenue,
            "total_orders": order_stats["total_orders"],
            "items_sold": total_items_sold,
            "new_orders": new_orders_count,
            "orders_inprogress": order_stats["total_orders_inprogress"],
            "orders_completed": order_stats["total_orders_completed"],
            "orders_canceled": order_stats["total_orders_canceled"],
            "orders_onhold": onhold_orders_count,
            "orders_pickedup": order_stats["total_orders_pickedup"],
            "orders_delivered": order_stats["total_orders_delivered"],
            "orders_dined": order_stats["total_orders_dined"],
            "online_orders": order_stats["online_orders_count"],
            "sales_order": sales_order_list,
            "top_selling": list_of_items
        }

        return Response(order_details)
    
    except Exception as e:
        return Response(str(e), status = status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST"])
def order_status_type_summary(request):
    try:
        try:
            vendor_id = int(request.data.get('vendor_id'))
            order_type_code = int(request.data.get('order_type_code'))
            page_number = int(request.data.get('page_number', 1))
            page_limit = int(request.data.get('page_limit', 10))

            start_date = datetime.strptime(request.data.get('start_date'), '%Y-%m-%d').date()
            end_date = datetime.strptime(request.data.get('end_date'), '%Y-%m-%d').date()

            if vendor_id <= 0 or order_type_code < 0 or page_number <= 0 or page_limit <= 0 or start_date > end_date:
                return Response("Invalid request data", status = status.HTTP_400_BAD_REQUEST)
            
        except:
            return Response("Invalid request data", status = status.HTTP_400_BAD_REQUEST)

        if not Vendor.objects.filter(pk = vendor_id).exists():
            return Response("Vendor does not exist", status = status.HTTP_400_BAD_REQUEST)

        order_filters = {
            "master_order__OrderDate__date__range": (start_date, end_date),
            "master_order__vendorId": vendor_id,
            "vendorId": vendor_id
        }
        
        if order_type_code != 0:
            order_filters["order_type"] = order_type_code

        order_status_aggregation = {
            'total_orders': Count('id'),
            'onhold_orders': Count('id', filter = Q(order_status = 4)),
            'closed_orders': Count('id', filter = Q(
                order_status = 10,
                master_order__Status = master_order_status_number["Completed"],
                master_order__orderpayment__status = True
            )),
            'cancelled_orders': Count('id', filter = Q(
                order_status = 5,
                master_order__Status = master_order_status_number["Canceled"]
            ))
        }
        
        order_list = []
        
        trunc_function = TruncDate

        if start_date == end_date:
            trunc_function = TruncHour
            
            order_filters["arrival_time__date__range"] = (start_date, end_date)

        order_list = KOMSOrder.objects.filter(**order_filters).annotate(order_date = trunc_function('arrival_time')) \
            .values('order_date').annotate(**order_status_aggregation).order_by('order_date')

        paginated_data = []
        
        paginator = Paginator(order_list, page_limit)
        page = paginator.get_page(page_number) 

        for order in page:
            if start_date == end_date:
                order_date = order['order_date'].astimezone(local_timezone).strftime("%Y-%m-%d %H:%M")
            
            else:
                order_date = datetime.combine(order['order_date'], datetime.min.time()).astimezone(local_timezone).strftime("%Y-%m-%d 00:00")

            paginated_data.append({
                "order_date": order_date,
                "total_orders": order['total_orders'],
                "closed_orders": order['closed_orders'],
                "onhold_orders": order['onhold_orders'],
                "cancelled_orders": order['cancelled_orders'],
            })

        return JsonResponse({
            "page_number": page.number,
            "total_pages": paginator.num_pages,
            "orders": paginated_data,
        })
    
    except Exception as e:
        return Response(str(e), status = status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST"])
def top_selling_product_details(request):
    try:
        try:
            vendor_id = int(request.data.get('vendor_id'))
            product_id = int(request.data.get('product_id'))
            start_date = datetime.strptime(request.data.get('start_date'), '%Y-%m-%d').date()
            end_date = datetime.strptime(request.data.get('end_date'), '%Y-%m-%d').date()
            page_number = int(request.data.get('page_number', 1))
            page_limit = int(request.data.get('page_limit', 10))

            if vendor_id <= 0 or product_id <=0 or page_limit <= 0 or page_number <= 0 or start_date > end_date:
                return Response("Invalid request data", status = status.HTTP_400_BAD_REQUEST)

        except:
            return Response("Invalid request data", status = status.HTTP_400_BAD_REQUEST)

        if not Vendor.objects.filter(pk = vendor_id).exists():
            return Response("Vendor not found", status = status.HTTP_400_BAD_REQUEST)

        product_info = Product.objects.filter(pk = product_id, vendorId = vendor_id).values("PLU", "productPrice").first()

        if not product_info:
            return Response("Product not found", status = status.HTTP_400_BAD_REQUEST)
        
        orders = Order_content.objects.filter(
            SKU = product_info["PLU"],
            orderId__order_status = 10,
            orderId__master_order__Status = master_order_status_number["Completed"],
            orderId__master_order__orderpayment__status = True,
            orderId__master_order__OrderDate__date__range = (start_date, end_date),
            orderId__master_order__vendorId = vendor_id,
            orderId__vendorId = vendor_id
        ).order_by("-orderId__master_order__OrderDate__date")
        
        if not orders.exists():
            return JsonResponse({"page_number": 1, "total_pages": 1, "orders": []})
        
        trunc_func = TruncDate

        datetime_format = "%Y-%m-%d"

        if start_date == end_date:
            trunc_func = TruncHour

            datetime_format = "%Y-%m-%d %H:00"
        
        orders = orders.annotate(order_date = trunc_func('orderId__master_order__OrderDate')).values('order_date') \
            .annotate(quantity_sold = Sum('quantity'), total_sale = Sum(F('quantity') * product_info["productPrice"])) \
            .order_by('order_date')
        
        order_summary = []

        for summary in orders:
            order_summary.append({
                "order_date": summary['order_date'].strftime(datetime_format),
                "quantity_sold": summary['quantity_sold'],
                "total_sale": summary['total_sale']
            })
            
        paginated_data = []

        paginator = Paginator(order_summary, page_limit)
        page = paginator.get_page(page_number) 
        
        for order in page:
            paginated_data.append({
                "order_date": order['order_date'],
                "quantity_sold": order['quantity_sold'],
                "total_sale": order['total_sale']
            })

        return JsonResponse({"page_number": page.number, "total_pages": paginator.num_pages, "orders": paginated_data})
    
    except Exception as e:
        return Response(str(e), status = status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST"])
def excel_download_for_dashboard(request):
    try:
        try:
            vendor_id = int(request.data.get("vendor_id"))
            start_date = datetime.strptime(request.data.get("start_date"), '%Y-%m-%d').date()
            end_date = datetime.strptime(request.data.get("end_date"), '%Y-%m-%d').date()
            platform_id = request.data.get("platform_id", "All")
            order_type = request.data.get("order_type", "All")
            order_status = request.data.get("order_status", "All")
            language = request.data.get("language", "English")

            if vendor_id <= 0 or start_date > end_date or not language:
                return JsonResponse({"error": "Invalid request data"}, status = status.HTTP_400_BAD_REQUEST)
            
            koms_order_filters = {"arrival_time__range": (start_date, end_date), "vendorId": vendor_id}
            
            platform_parameter = "All" if language == "English" else language_localization["All"]

            if platform_id != "All":
                platform_id = int(platform_id)

                if platform_id <= 0:
                    return JsonResponse({"error": "Invalid request data"}, status = status.HTTP_400_BAD_REQUEST)

                platform_info = Platform.objects.filter(pk = platform_id, VendorId = vendor_id).values("Name", "Name_locale").first()
            
                if not platform_info:
                    return JsonResponse({"error": "Platform does not exist"}, status = status.HTTP_400_BAD_REQUEST)

                koms_order_filters["master_order__platform"] = platform_id

                platform_parameter = platform_info["Name"] if language == "English" else platform_info["Name_locale"]

            if order_type != "All":
                order_type = int(order_type)

                if order_type <= 0:
                    return JsonResponse({"error": "Invalid request data"}, status = status.HTTP_400_BAD_REQUEST)

                koms_order_filters["order_type"] = order_type

            if order_status != "All":
                order_status = int(order_status)

                if order_status <= 0:
                    return JsonResponse({"error": "Invalid request data"}, status = status.HTTP_400_BAD_REQUEST)

                koms_order_filters["order_status"] = order_status
            
        except:
            return JsonResponse({"error": "Invalid request data"}, status = status.HTTP_400_BAD_REQUEST)
        
        if check_key_exists("order_type", order_type) == False:
            return JsonResponse({"error": "Order Type does not exist"}, status = status.HTTP_400_BAD_REQUEST)
        
        if check_key_exists("koms_order_status", order_status) == False:
            return JsonResponse({"error": "Order Status does not exist"}, status = status.HTTP_400_BAD_REQUEST)
        
        order_type_parameter = order_type_english[order_type]
        order_status_parameter = koms_order_status_english[order_status]
                    
        if language != "English":
            order_type_parameter = language_localization[order_type_english[order_type]]
            order_status_parameter = language_localization[koms_order_status_english[order_status]]

        start_date_parameter = start_date.strftime('%d-%m-%Y')
        end_date_parameter = end_date.strftime('%d-%m-%Y')

        start_date = datetime.strptime(str(start_date) + " 00:00:00.000000", '%Y-%m-%d %H:%M:%S.%f')
        end_date = datetime.strptime(str(end_date) + " 23:59:59.000000", '%Y-%m-%d %H:%M:%S.%f')
        
        order_details = {}

        order_ids = tuple(KOMSOrder.objects.filter(**koms_order_filters).values_list('pk', flat=True))

        order_content = Order_content.objects.filter(orderId__in = order_ids).select_related("orderId", "orderId__master_order") \
            .values("orderId__master_order__pk", "orderId__order_status")

        for content in order_content:
            order_id = content["orderId__master_order__pk"]
            koms_order_status = content["orderId__order_status"]
            
            order_detail = Order.objects.filter(pk = order_id).select_related("platform").values(
                "pk", "arrivalTime", "orderType", "platform__Name", "platform__Name_locale"
            ).first()

            if not order_detail:
                return JsonResponse(
                    {"error": f"Order details not found for order ID {order_id}"},
                    status = status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            
            order_time = order_detail["arrivalTime"].astimezone(local_timezone).replace(tzinfo = None)

            order_type = order_type_english[order_detail["orderType"]]
            order_status = koms_order_status_english[koms_order_status]
    
            if language != "English":
                order_type = language_localization[order_type_english[order_detail["orderType"]]]
                order_status = language_localization[koms_order_status_english[koms_order_status]]
            
            payment_detail = OrderPayment.objects.filter(orderId = order_detail["pk"]).last()

            if not payment_detail:
                transaction_id = ""
                payment_type = ""

                payment_status = payment_status_english["Unknown"] if language == "English" else language_localization[payment_status_english["Unknown"]]

            if payment_detail.status == True:
                payment_status = payment_status_english["True"] if language == "English" else language_localization[payment_status_english["True"]]
            
            else:
                payment_status = payment_status_english["False"] if language == "English" else language_localization[payment_status_english["False"]]
            
            amount_paid = payment_detail.paid
            transaction_id = payment_detail.paymentKey

            payment_type = payment_type_english[payment_detail.type] if language == "English" else language_localization[payment_type_english[payment_detail.type]]
                
            platform_name = order_detail["platform__Name"] if language == "English" else order_detail["platform__Name_locale"]

            order_details[order_id] = {
                "order_id": order_id,
                "platform_name": platform_name,
                "amount_paid": amount_paid,
                "order_type": order_type,
                "order_status": order_status,
                "order_time": order_time,
                "payment_type": payment_type,
                "payment_status": payment_status,
                "transaction_id": transaction_id,
            }

        new_order_details = OrderedDict()
        sorted_items = []

        for key, value in order_details.items():
            order_time = value.get("order_time", "")
            sorted_items.append((order_time, int(key)))

        sorted_items.sort()

        for order_time, order_id in sorted_items:
            new_order_details[order_id] = order_details[order_id]

        order_details = new_order_details
        
        order_count = len(order_details)
        
        total_amount_paid = 0.0

        subtotal_sum, discount_sum = Order.objects.filter(
            OrderDate__date__range = (start_date, end_date),
            orderpayment__status = True,
            vendorId = vendor_id
        ).exclude(Status = master_order_status_number["Canceled"]) \
        .aggregate(subtotal_sum = Sum('subtotal'), discount_sum = Sum('discount')).values()

        subtotal_sum = subtotal_sum or 0.0
        discount_sum = discount_sum or 0.0
        
        total_amount_paid = "{:.2f}".format(subtotal_sum - discount_sum)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        if language == "English":
            sheet.append(['Start Date', f'{start_date_parameter}', '', '', '', '', '', '', ''])
            sheet.append(['End Date', f'{end_date_parameter}', '', '', '', '', '', '', ''])
            sheet.append(['Platform', f'{platform_parameter}', '', '', '', '', '', '', ''])
            sheet.append(['Order Type', f'{order_type_parameter}', '', '', '', '', '', '', ''])
            sheet.append(['Order Status', f'{order_status_parameter}', '', '', '', '', '', '', ''])
            sheet.append(['', '', '', '', '', '', '', '', ''])

            sheet.append(['Order ID', 'Platform', 'Amount', 'Order Type', 'Order Status', 'Order Time', 'Payment Type', 'Payment Status', 'Transaction ID'])

        else:
            sheet.append([language_localization['Start Date'], f'{start_date_parameter}', '', '', '', '', '', '', ''])
            sheet.append([language_localization['End Date'], f'{end_date_parameter}', '', '', '', '', '', '', ''])
            sheet.append([language_localization['Platform'], f'{platform_parameter}', '', '', '', '', '', '', ''])
            sheet.append([language_localization['Order Type'], f'{order_type_parameter}', '', '', '', '', '', '', ''])
            sheet.append([language_localization['Order Status'], f'{order_status_parameter}', '', '', '', '', '', '', ''])
            sheet.append(['', '', '', '', '', '', '', '', ''])

            sheet.append([
                language_localization['Order ID'],
                language_localization['Platform'],
                language_localization['Amount'],
                language_localization['Order Type'],
                language_localization['Order Status'],
                language_localization['Order Time'],
                language_localization['Payment Type'],
                language_localization['Payment Status'],
                language_localization['Transaction ID']
            ])

        for order_id, details in order_details.items():
            sheet.append([
                details["order_id"],
                details["platform_name"],
                details["amount_paid"],
                details["order_type"],
                details["order_status"],
                details["order_time"].strftime('%d-%m-%Y %I:%M:%S %p') if details["order_time"] else "",
                details["payment_type"],
                details["payment_status"],
                details["transaction_id"],
            ])
        
        if language == "English":
            sheet.append([f'Total orders = {order_count}', '', f'Total revenue = {total_amount_paid}', '', '', '', '', '', ''])

        else:
            sheet.append([f'{language_localization["Total orders"]} = {order_count}', '', f'{language_localization["Total revenue"]} = {total_amount_paid}', '', '', '', '', '', ''])
        
        directory = os.path.join(settings.MEDIA_ROOT, 'Excel Downloads')
        os.makedirs(directory, exist_ok = True) # Create the directory if it doesn't exist inside MEDIA_ROOT

        file_name = f"Order_data_Vendor{vendor_id}.xlsx"

        relative_file_path = os.path.join('Excel Downloads', file_name)

        file_path = os.path.join(settings.MEDIA_ROOT, relative_file_path)

        workbook.save(file_path)

        print(f"Excel file '{file_name}' has been created.")
        
        response = "/media/" + relative_file_path
        
        return HttpResponse(response, status = status.HTTP_200_OK)
    
    except Exception as e:
        print(e)
        return JsonResponse({"error": str(e)}, status = status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_categories(request):
    try:
        vendor_id = request.GET.get("vendorId")
        language = request.GET.get("language", "English")

        if not vendor_id:
            return JsonResponse({"message": "Invalid Vendor ID", "categories": []}, status = status.HTTP_400_BAD_REQUEST)

        category_list = []

        categories = ProductCategoryJoint.objects.filter(vendorId = vendor_id).select_related("category").values(
            'category__id',
            'category__categoryName',
            'category__categoryName_locale',
            'category__categoryDescription',
            'category__categoryDescription_locale',
            'category__categoryPLU',
            'category__categoryImageUrl',
            'category__is_active',
        ).distinct('category')

        for category in categories:
            category_name = category["category__categoryName"]
            category_description = category["category__categoryDescription"]
            
            if language != "English":
                category_name = category["category__categoryName_locale"]
                category_description = category["category__categoryDescription_locale"]
            
            category_list.append({
                "categoryId": category["category__id"],
                "categoryPlu": category["category__categoryPLU"],
                "name": category_name,
                "description": category_description or "",
                "image": category["category__categoryImageUrl"] or "https://www.stockvault.net/data/2018/08/31/254135/preview16.jpg",
                "is_active": category["category__is_active"]
            })

        return JsonResponse({"message": "", "categories": category_list}, status = status.HTTP_200_OK)
    
    except Exception as e:
        return JsonResponse({"message": str(e), "categories": []}, status = status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def productByCategory(request, id = 0):
    try:
        vendor_id = request.GET.get("vendorId")
        language = request.GET.get("language", "English")
        search_text = request.GET.get("search")
        platform = request.GET.get("platform")
        product_tag = request.GET.get("tag")

        if not vendor_id:
            return JsonResponse({"message": "Invalid Vendor ID", "products": {}}, status = status.HTTP_400_BAD_REQUEST)

        if id != 0:
            category_details = ProductCategory.objects.filter(pk = id)

        else:
            category_details = ProductCategory.objects.filter(vendorId = vendor_id)

        if search_text:
            product_filter = {"vendorId": vendor_id}

            if (platform == "Website" or platform == "Mobile App") and product_tag == "veg":
                product_filter["tag"] = "veg"

            products = Product.objects.filter(**product_filter).filter(
                Q(productName__icontains = search_text) | Q(productName_locale__icontains = search_text)
            )
            
            product_list = get_product_by_category_data(products, language, vendor_id)

            return JsonResponse({"message": "", "products": {"1": product_list}}, status = status.HTTP_200_OK)
        
        products_by_category = {}

        for category in category_details:
            product_ids = ProductCategoryJoint.objects.filter(category = category.pk, vendorId = vendor_id).values_list("product", flat = True)

            filtered_products = Product.objects.filter(pk__in = product_ids, vendorId = vendor_id)
            
            product_list = get_product_by_category_data(filtered_products, language, vendor_id)
            
            products_by_category[category.pk] = product_list

        return JsonResponse({"message": "", "products": products_by_category}, status = status.HTTP_200_OK)
    
    except Exception as e:
        return JsonResponse({"message": str(e), "products": {}}, status = status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST"])
def product_on_off(request):
    try:
        product = Product.objects.filter(pk = request.data.get("productId"), vendorId = request.data.get("vendorId"))

        if not product.exists():
            return JsonResponse({'error': 'Product not found'}, status = status.HTTP_400_BAD_REQUEST)
        
        product.update(active = request.data.get("status"))

        product = product.values("pk", "active").first()

        return JsonResponse({'productId': product["pk"], 'status': product["active"]})
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status = status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST"])
def modifier_on_off(request):
    try:
        modifier = ProductModifier.objects.filter(pk = request.data.get('id'), vendorId = request.GET.get("vendorId"))

        if not modifier.exists():
            return JsonResponse({"message": "Modifier not found"}, status = status.HTTP_400_BAD_REQUEST)
        
        modifier.update(active = request.data.get('active'))

        return JsonResponse({"message": "Modifier status updated successfully"})
    
    except Exception as e:
        return JsonResponse({"message": str(e)}, status = status.HTTP_500_INTERNAL_SERVER_ERROR)


# remove this API
@api_view(['GET'])
def show_tableCapacity(request):
    vendorId=request.GET.get("vendorId")
    try:
        data=HotelTable.objects.filter(vendorId=vendorId)
        tableCapacity =list(set([ i.tableCapacity for i in data]))
        table = [str(i) for i in tableCapacity]
        return JsonResponse({ "tableCapacity": table}, safe=False)
    except Exception as e:
        return JsonResponse({"error":str(e)})


def order_data(vendor_id, page_number, search, order_status, order_type, platform, is_dashboard = 0, s_date = None, e_date = None, language = "English"):
    try:
        if not vendor_id:
            return "Vendor ID cannot be empty"
            
        order_details = {}

        start_date_parameter = s_date or ""
        end_date_parameter = e_date or ""
        search_text_parameter = search if search != "All" else ""
        platform_parameter = platform if platform != "All" else ""
        page_number_parameter = page_number if page_number != "All" else 1
        order_status_parameter = order_status if order_status != "All" else ""
        order_type_parameter = order_type if order_type != "All" else ""

        current_date = datetime.today().strftime("%Y-%m-%d")

        order_data = None
        
        koms_order_filters = {"vendorId": vendor_id}
        
        if s_date and e_date:
            s_date = str(s_date).replace("T", "-")
            e_date = str(e_date).replace("T", "-")
            
            koms_order_filters["arrival_time__date__range"] = (s_date, e_date)
        
        else:
            koms_order_filters["arrival_time__date"] = current_date
        
        if order_status != "All":
            if is_dashboard == 0:
                koms_order_filters["order_status"] = order_status

            elif is_dashboard == 1:
                if order_status == 10 :
                    if s_date and e_date:
                        koms_order_filters["order_status__in"] = (4, 5, 10)

                    else:
                        koms_order_filters["order_status__in"] = (5, 10)

                else:
                    koms_order_filters["order_status"] = order_status

        else:
            if is_dashboard == 1:
                koms_order_filters["order_status__in"] = (2, 3, 4, 6, 7, 8, 9)

        if order_type != "All":
            koms_order_filters["order_type"] = order_type

        if platform != "All":
            if is_dashboard == 1:
                koms_order_filters["master_order__platform"] = platform

            else:
                online_platform_ids = Platform.objects.filter(Name__in = ("Website", "Mobile App")).values_list("pk", flat = True)

                if platform in online_platform_ids:
                    koms_order_filters["master_order__platform__Name__in"] = ("Website", "Mobile App")

                else:
                    koms_order_filters["master_order__platform"] = platform

        if search != 'All':
            output = re.search(r'\d+', search)

            if output:
                master_order_id = output.group()

                koms_order_filters["master_order__pk__icontains"] = master_order_id

                order_data = KOMSOrder.objects.filter(**koms_order_filters)

            else:
                output = re.search(r'[A-Za-z ]+', search)

                if output:
                    customer_name = output.group()

                    order_data = KOMSOrder.objects.filter(**koms_order_filters).filter(
                        Q(master_order__customerId__FirstName__icontains = customer_name) | \
                        Q(master_order__customerId__LastName__icontains = customer_name)
                    )

        else:
            order_data = KOMSOrder.objects.filter(**koms_order_filters)

        if not order_data.exists():
            response_data = {
                'page_number': 1,
                'total_pages': 1,
                'data_count': 0,
                'order_details': {},
            }

        order_data = order_data.order_by("-master_order__OrderDate") \
            .select_related("master_order", "master_order__platform", "master_order__customerId", "vendorId") \
            .values(
                "pk", "master_order__pk", "master_order__TotalAmount", "master_order__subtotal", "master_order__tax",
                "master_order__delivery_charge", "master_order__discount", "master_order__platform__pk", "master_order__platform__Name",
                "master_order__platform__Name_locale", "master_order__customerId__pk", "master_order__customerId__FirstName",
                "master_order__customerId__LastName", "master_order__customerId__Phone_Number", "master_order__customerId__Email",
                "vendorId", "vendorId__franchise_location"
            )
        
        paginator = Paginator(order_data, 10)

        page_obj = paginator.get_page(page_number)

        orders_for_page = page_obj.object_list

        for order in orders_for_page:
            order_details[order["pk"]] = get_order_info_for_socket(order_info = order, language = language, vendor_id = vendor_id)

        response_data = {
            'page_number': page_obj.number,
            'total_pages': paginator.num_pages,
            'data_count': paginator.count,
            'order_details': order_details,
        }
        
        webSocketPush(
            message = response_data,
            room_name = f"POS{order_status_parameter}-{search_text_parameter}-{platform_parameter}-{order_type_parameter}-{page_number_parameter}-{start_date_parameter}-{end_date_parameter}-{is_dashboard}-{language}-{str(vendor_id)}",
            username = "CORE",
        )

        webSocketPush(
            message = response_data,
            room_name = f"POS{order_status_parameter}-{search_text_parameter}-{platform_parameter}-{order_type_parameter}--{start_date_parameter}-{end_date_parameter}-{is_dashboard}-{language}-{str(vendor_id)}",
            username = "CORE",
        )
        
        return response_data
    
    except Exception as e:
        print(e)
        return str(e)


def order_data_start_thread(vendor_id, page_number, search, order_status, order_type, platform, is_dashboard = 0, s_date = None, e_date = None, language = "English"):
        print("Starting koms thread...")
        
        thr = threading.Thread(target = order_data, args = (), kwargs = {
            "vendor_id":vendor_id,
            "page_number":page_number,
            "search":search,
            "order_status":order_status,
            'order_type':order_type,
            "platform":platform,
            "search":search,
            "platform":platform,
            "s_date":s_date,
            "e_date":e_date,
            "is_dashboard":is_dashboard,
            "language": language
        })
        
        thr.setDaemon(True)
        thr.start()

        return {"connecting":False}


@api_view(['POST'])
def createOrder(request):
    try:
        vendor_id = request.GET.get('vendorId', None)
        language = request.GET.get("language", "English")

        if vendor_id is None:
            return JsonResponse({"message": "Vendor Id cannot be empty"}, status=status.HTTP_400_BAD_REQUEST, safe=False)
    
        platform = Platform.objects.filter(Name=request.data.get('platform'), VendorId=vendor_id).first()

        if (not platform) or (platform.isActive == False):
            return JsonResponse({"message": "Contact your administrator to activate the platform"}, status=status.HTTP_400_BAD_REQUEST, safe=False)
        
        orderid = vendor_id + str(platform.pk) + datetime.now().strftime("%H%M%S%f")[:15]

        payment_type_string = request.data.get("payment_details").get("paymentType")

        if payment_type_string:
            payment_type_string = payment_type_string.capitalize()

        payment_type = payment_type_number[payment_type_string] if payment_type_string else payment_type_number["Cash"]
        
        result = {
            "language": language,
            "internalOrderId": orderid,
            "vendorId": vendor_id,
            "externalOrderId": orderid,
            "orderType": request.data.get("type"),
            "pickupTime": '',
            "arrivalTime": '',
            "deliveryIsAsap": 'true',
            "note": request.data.get('customerNote'),
            "tables": request.data.get('tables'),
            "items": [],
            "remake": False,
            "customerName": request.data.get('name') if request.data.get('name') else "",
            "status": "pending",
            "platform": request.data.get('platform'),
            "customer": {
                "fname": request.data.get('FirstName') if request.data.get('FirstName') else "Guest",
                "lname": request.data.get('LastName') if request.data.get('LastName') else "",
                "email": request.data.get('Email') if request.data.get('Email') else "",
                "phno": request.data.get('Phone_Number') if request.data.get('Phone_Number') else "0",
                "address1": request.data.get('address_line_1') if request.data.get('address_line_1') else "",
                "address2": request.data.get('address_line_2') if request.data.get('address_line_2') else "",
                "city": request.data.get('city') if request.data.get('city') else "",
                "state": request.data.get('state') if request.data.get('state') else "",
                "country": request.data.get('country') if request.data.get('country') else "",
                "zip": request.data.get('zipcode') if request.data.get('zipcode') else "",
                "vendorId": vendor_id
            },
            "discount":{
                "value": request.data.get('discount'),
                "calType": 2
            },
            "payment": {
                "tipAmount": request.data.get('tip',0.0),
                "payConfirmation": request.data.get("payment_details").get("paymentKey") if request.data.get("payment_details").get("paymentKey") else "",
                "payAmount": request.data.get("finalTotal",0.0),
                "payType": payment_type,
                "mode": payment_type,
                "default": request.data.get("payment_details").get("paymentStatus") if request.data.get("payment_details").get("paymentStatus") else  False,
                "platform": request.data.get("payment_details").get("platform") if request.data.get("payment_details").get("platform") else "N/A",
                "custProfileId": "",
                "custPayProfileId": "",
                "payData": "",
                "CardId": "NA",
                "expDate": "",
                "transcationId": request.data.get("payment_details").get("paymentKey"),
                "lastDigits": "",
                "billingZip": ""
            },
            "tax": request.data.get("tax"),
            "subtotal": request.data.get("subtotal"),
            "finalTotal": request.data.get("finalTotal"),
            "is_wordpress": request.data.get("is_wordpress"),
            "points_redeemed": request.data.get("points_redeemed"),
            "points_redeemed_by": request.data.get("points_redeemed_by"),
        }
        
        # if request.data.get('promocodes'):
        #     discount = Order_Discount.objects.get(pk=request.data.get('promocodes')[0]['id'])

        #     result['discount'] = {
        #         "discountCode": discount.discountCode,
        #         "discountId": discount.plu,
        #         "status": True,
        #         "discountName": discount.discountName,
        #         "discountCost": discount.value
        #     }
        
        items = []

        for item in request.data["products"]:
            product_instance = Product.objects.get(pk=item['productId'], vendorId=vendor_id)
            
            modifier_list_1 = []
            modifier_list_2 = []
            
            for modifier_group in item['modifiersGroup']:
                for modifier in modifier_group['modifiers']:
                    modifier_data = {
                        "plu": ProductModifier.objects.get(pk=modifier["modifierId"]).modifierPLU,
                        "name": modifier['name'],
                        "status": modifier["status"],
                        "group": modifier_group['id']
                    }

                    modifier_list_1.append(modifier_data)

            for modifier_group in item['modifiersGroup']:
                for modifier in modifier_group['modifiers']:
                    if modifier["status"]:
                        modifier = {
                            "plu": ProductModifier.objects.get(pk=modifier["modifierId"]).modifierPLU,
                            "name": modifier['name'],
                            "status": modifier["status"],
                            "quantity": modifier["quantity"],
                            "group": modifier_group['id']
                        }

                        modifier_list_2.append(modifier)
            
            itemData = {
                "plu": product_instance.productParentId.PLU if product_instance.productParentId != None else product_instance.PLU,
                "sku": item.get("sku"),
                "productName": product_instance.productName,
                # Variation Id instead of name
                "variantName": str(item["variation_id"]) if item.get("variation_id") else "txt",
                "quantity": item["quantity"],
                "subItems": modifier_list_1,
                "itemRemark": item["note"],  # Note Unavailable
                "unit": "qty",  # Default
                "modifiers": modifier_list_2
            }
            
            if product_instance.productParentId != None:
                itemData["variant"] = {"plu": product_instance.PLU}

            items.append(itemData)

        result["items"] = items

        result["tip"] = request.data['tip'] 

        tokenlist = KioskOrderData.objects.filter(date=datetime.today().date()).values_list('token')

        token = 1 if len(tokenlist)==0 else max(tokenlist)[0] + 1

        response = order_helper.OrderHelper.openOrder(result, vendor_id)
        
        saveData = KiosK_create_order_serializer(data={
            'orderdata': str(result),
            'date': datetime.today().date(),
            'token': token
        })
        
        if saveData.is_valid():
            saveData.save()

            if Order.objects.filter(externalOrderId=orderid).exists():
                return JsonResponse({'token': token, "external_order_id": orderid})
            
            else:
                return JsonResponse({'token': token, "external_order_id": ""})
        
        return JsonResponse({"message": "Something went wrong"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    except Exception as e:
        print(e)
        return JsonResponse({"message": "Something went wrong"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# Replica of order_data function
@api_view(["POST"])
def get_orders(request):
    try:
        vendor_id = request.data.get("vendor_id")
        page_number = request.data.get("page_number", 1)
        search = request.data.get("search")
        order_status = request.data.get("order_status")
        order_type = request.data.get("order_type")
        platform = request.data.get("platform")
        is_dashboard = request.data.get("is_dashboard")
        start_date = request.data.get("start_date")
        end_date = request.data.get("end_date")
        language = request.data.get("language", "English")
        get_all_vendor_data = request.data.get("get_all_vendor_data")
        franchise_vendor_id = request.data.get("franchise_vendor_id")
        is_range = request.data.get("is_range")

        try:
            vendor_id = int(vendor_id)

        except:
            return Response("Invalid vendor ID", status = status.HTTP_400_BAD_REQUEST)
        
        vendor = Vendor.objects.filter(pk = vendor_id, is_active = True)

        if not vendor.exists():
            return Response("Vendor does not exist or not active", status = status.HTTP_400_BAD_REQUEST)
            
        if get_all_vendor_data == True:
            vendor_ids = tuple(Vendor.objects.filter(franchise = vendor_id).values_list("pk", flat = True))

        elif get_all_vendor_data == False and franchise_vendor_id:
            vendor_ids = (franchise_vendor_id,)
            vendor_id = franchise_vendor_id

        else:
            vendor_ids = (vendor_id,)

        current_date = datetime.today().strftime("%Y-%m-%d")
        
        koms_order_filters = {"vendorId__in": vendor_ids}
        
        if start_date and end_date:
            koms_order_filters["arrival_time__date__range"] = (start_date, end_date)
        
        else:
            koms_order_filters["arrival_time__date"] = current_date
        
        if order_status != "All":
            if is_dashboard == 0:
                koms_order_filters["order_status"] = order_status

            elif is_dashboard == 1:
                if order_status == 10 :
                    if start_date and end_date:
                        koms_order_filters["order_status__in"] = (4, 5, 10)

                    else:
                        koms_order_filters["order_status__in"] = (5, 10)

                else:
                    koms_order_filters["order_status"] = order_status

        else:
            if is_dashboard == 1 and start_date == current_date and end_date == current_date:
                if is_range == False:
                    koms_order_filters["order_status__in"] = (2, 3, 4, 6, 7, 8, 9)
            

        if order_type != "All":
            koms_order_filters["order_type"] = order_type

        if platform != "All":
            if is_dashboard == 1:
                koms_order_filters["master_order__platform"] = platform

            else:
                online_platform_ids = Platform.objects.filter(Name__in = ("Website", "Mobile App")).values_list("pk", flat = True)

                if platform in online_platform_ids:
                    koms_order_filters["master_order__platform__Name__in"] = ("Website", "Mobile App")

                else:
                    koms_order_filters["master_order__platform"] = platform

        if search != 'All':
            output = re.search(r'\d+', search)

            if output:
                master_order_id = output.group()

                koms_order_filters["master_order__pk__icontains"] = master_order_id

                order_data = KOMSOrder.objects.filter(**koms_order_filters)

            else:
                output = re.search(r'[A-Za-z ]+', search)

                if output:
                    customer_name = output.group()

                    order_data = KOMSOrder.objects.filter(**koms_order_filters).filter(
                        Q(master_order__customerId__FirstName__icontains = customer_name) | \
                        Q(master_order__customerId__LastName__icontains = customer_name)
                    )

        else:
            order_data = KOMSOrder.objects.filter(**koms_order_filters)
        
        if not order_data.exists():
            response_data = {'page_number': 1, 'total_pages': 1, 'data_count': 0, 'order_details': {}}

        order_details = {}

        if get_all_vendor_data == True:
            order_data = order_data.order_by("vendorId", "-master_order__OrderDate")

        else:
            order_data = order_data.order_by("-master_order__OrderDate")

        order_data = order_data.select_related("master_order", "master_order__platform", "master_order__customerId", "vendorId") \
        .values(
            "pk", "master_order__pk", "master_order__TotalAmount", "master_order__subtotal", "master_order__tax",
            "master_order__delivery_charge", "master_order__discount", "master_order__platform__pk", "master_order__platform__Name",
            "master_order__platform__Name_locale", "master_order__customerId__pk", "master_order__customerId__FirstName",
            "master_order__customerId__LastName", "master_order__customerId__Phone_Number", "master_order__customerId__Email",
            "vendorId", "vendorId__franchise_location"
        )
            
        paginator = Paginator(order_data, 10)

        page_obj = paginator.get_page(page_number)

        orders_for_page = page_obj.object_list

        for order in orders_for_page:
            order_details[order["pk"]] = get_order_info_for_socket(order_info = order, language = language, vendor_id = order["vendorId"])

        response_data = {
            'page_number': page_obj.number,
            'total_pages': paginator.num_pages,
            'data_count': paginator.count,
            'order_details': order_details,
        }
        
        return Response(response_data, status = status.HTTP_200_OK)
    
    except Exception as e:
        return Response(str(e), status = status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def platform_list(request):
    vendor_id = request.GET.get('vendorId')
    language = request.GET.get('language', 'English')

    if not vendor_id:
        return JsonResponse({"error": "Vendor Id cannot be empty"}, status = status.HTTP_400_BAD_REQUEST)

    platform_details = []

    if language == "English":
        platform_details.append({"id": "", "name": "All"})

    else:
        platform_details.append({"id": "", "name": language_localization["All"]})
    
    platforms = Platform.objects.filter(isActive = True, VendorId = vendor_id).exclude(Name__in = ("KOMS", "Inventory")) \
        .values("pk", "Name", "Name_locale")
    
    for platform in platforms:
        platform_details.append({
            "id": platform["pk"],
            "name": platform["Name"] if language == "English" else platform["Name_locale"]
        })

    return JsonResponse({'platforms': platform_details}, status = status.HTTP_200_OK)


@api_view(['POST'])
def make_payment(request):
    language = request.GET.get('language', 'English')

    try:
        vendorId = int(request.GET.get('vendorId'))

        if vendorId <= 0:
            return Response("Invalid Vendor ID", status = status.HTTP_400_BAD_REQUEST)

    except:
        return Response("Invalid Vendor ID", status = status.HTTP_400_BAD_REQUEST)

    vendor_info = Vendor.objects.filter(pk = vendorId).values("pk", "secondary_language").first()

    if not vendor_info:
        return Response("Vendor does not exist", status = status.HTTP_400_BAD_REQUEST)
    
    request_data = request.data
    
    order = KOMSOrder.objects.get(externalOrderId = request_data['orderid'])

    old_order_status = order.order_status
    
    master_order = Order.objects.filter(externalOrderId = order.externalOrderId, vendorId = vendorId).last()

    if request_data.get("payment_id"):
        payment = OrderPayment.objects.filter(pk = request_data.get("payment_id")).last()

    else:
        payment = OrderPayment.objects.filter(orderId = master_order.pk).last()

    if payment:
        payment.paymentBy = request_data['payment']['paymentBy']
        payment.paymentKey = request_data['payment']['paymentKey']
        payment.type = request_data['payment']['type']
        payment.status = True
        payment.platform = request_data['payment']['platform']

        payment.save()

    else:
        payment = OrderPayment.objects.create(
            orderId = master_order,
            paymentBy = request_data['payment']['paymentBy'],
            paymentKey = request_data['payment']['paymentKey'],
            paid = master_order.TotalAmount,
            due = 0,
            tip = 0,
            type = request_data['payment']['type'],
            status = True,
            platform = request_data['payment']['platform']
        )
    if request_data['payment']['type'] == 5:            # 5 = Mix payment type 
        OrderPayment.objects.filter(masterPaymentId = payment.pk).delete()
        for single_payment in request_data['payment']['payment_list']:
            OrderPayment.objects.create(
            orderId = master_order,
            masterPaymentId = payment,
            paymentBy = single_payment['paymentBy'],
            paymentKey = single_payment['paymentKey'],
            paid = single_payment['paid'],
            due = 0,
            tip = 0,
            type = single_payment['type'],
            status = True,
            platform = single_payment['platform']
            )
    all_status_true = True

    if request_data.get("payment_id"):
        master_order_id = OrderPayment.objects.filter(pk = request_data.get("payment_id")) \
            .values_list("orderId__masterOrder__pk", flat = True).last()

        split_orders = Order.objects.filter(masterOrder = master_order_id).values_list("pk", flat = True)

        for order_payment in OrderPayment.objects.filter(orderId__in = split_orders):
            if not order_payment.status:
                all_status_true = False
                break

        if all_status_true == True:
            OrderPayment.objects.filter(orderId = master_order_id).update(status = True)

    if master_order.orderType == 3:
        order.order_status = 10
        order.save()
        
        master_order.Status = 2
        master_order.save() # core order status this needs to be changed by updateCoreOrder function

        waiter_heads = Waiter.objects.filter(is_waiter_head = True, vendorId = vendorId).values_list("pk", flat = True)
        
        tables = Order_tables.objects.filter(orderId = order)
        
        for table in tables:
            table.tableId.status = 1 # EMPTY TABLE
            table.tableId.guestCount = 0
            table.tableId.save()

            waiter_id = 0

            if table.tableId.waiterId:
                waiter_id = table.tableId.waiterId.pk

            filtered_waiter_heads = set(waiter_heads) - set((waiter_id,))
            
            table_data = get_table_data(hotelTable = table.tableId, language = "English", vendorId = vendorId)
            
            webSocketPush(
                message = {"result": table_data, "UPDATE": "UPDATE"},
                room_name = f"WOMS{str(waiter_id)}------English-{str(vendorId)}",
                username = "CORE",
            )
            
            webSocketPush(
                message = {"result": table_data, "UPDATE": "UPDATE"},
                room_name = f"WOMSPOS------English-{str(vendorId)}",
                username = "CORE",
            )
            
            for waiter_head_id in filtered_waiter_heads:
                webSocketPush(
                    message = {"result": table_data, "UPDATE": "UPDATE"},
                    room_name = f"WOMS{str(waiter_head_id)}------English-{str(vendorId)}",
                    username = "CORE",
                )

            secondary_language = vendor_info["secondary_language"]
            
            if secondary_language:
                table_data_locale = get_table_data(hotelTable = table.tableId, language = secondary_language, vendorId = vendorId)
                
                webSocketPush(
                    message = {"result": table_data_locale, "UPDATE": "UPDATE"},
                    room_name = f"WOMS{str(waiter_id)}------{secondary_language}-{str(vendorId)}",
                    username = "CORE",
                )

                webSocketPush(
                    message = {"result": table_data_locale, "UPDATE": "UPDATE"},
                    room_name = f"WOMSPOS------{secondary_language}-{str(vendorId)}",
                    username = "CORE",
                )

                for waiter_head_id in filtered_waiter_heads:
                    webSocketPush(
                        message = {"result": table_data_locale, "UPDATE": "UPDATE"},
                        room_name = f"WOMS{str(waiter_head_id)}------{secondary_language}-{str(vendorId)}",
                        username = "CORE",
                    )

    elif ((master_order.orderType == 1) or (master_order.orderType == 2)) and (order.order_status == 3):
        order.order_status = 10
        order.save()
            
        master_order.Status = 2
        master_order.save()
    
    webSocketPush(
        message = {"id": order.pk, "orderId": order.externalOrderId, "UPDATE": "REMOVE",},
        room_name = str(vendorId)+'-'+str(old_order_status),
        username = "CORE",
    )  # WheelMan order remove order from old status
    
    allStationWiseRemove(id = order.pk, old = str(old_order_status), current = str(old_order_status), vendorId = vendorId)
    allStationWiseSingle(id = order.pk, vendorId = vendorId)
    allStationWiseCategory(vendorId = vendorId) 
    
    waiteOrderUpdate(orderid = order.pk, language = language, vendorId = vendorId)

    return JsonResponse({})


@api_view(["GET"])
def get_order_info(request):
    try:
        vendor_id = request.GET.get('vendorId')
        external_order_id = request.GET.get('id')
        platform_id = request.GET.get('platform')
        payment_id = request.GET.get("paymentId")
        language = request.GET.get('language', 'English')

        if not all((vendor_id, external_order_id, platform_id)):
            return Response({"error": "Invalid vendor ID, External Order ID or Platform ID"}, status = status.HTTP_400_BAD_REQUEST)
        
        try:
            vendor_id = int(vendor_id)
            platform_id = int(platform_id)

        except:
            return Response({"error": "Invalid vendor ID or Platform ID"}, status = status.HTTP_400_BAD_REQUEST)
        
        if not Vendor.objects.filter(pk = vendor_id, is_active = True).exists():
            return Response({"error": "Vendor does not exist or not active"}, status = status.HTTP_400_BAD_REQUEST)
        
        koms_order = KOMSOrder.objects.filter(externalOrderId = external_order_id, master_order__vendorId = vendor_id)\
            .select_related('master_order', 'master_order__customerId', 'master_order__platform') \
            .prefetch_related(
                Prefetch('master_order__customerId__address_set', queryset = Address.objects.filter(type = 'shipping_address', is_selected = True)),
                Prefetch('master_order__orderpayment_set', queryset = OrderPayment.objects.filter(masterPaymentId__isnull = True)),
                Prefetch('order_tables_set', queryset = Order_tables.objects.select_related('tableId', 'tableId__waiterId')),
                Prefetch('order_content_set', queryset = Order_content.objects.select_related('orderId')),
            ).first()

        if not koms_order:
            return JsonResponse({"error": "Order not found"}, status = status.HTTP_400_BAD_REQUEST)

        if not Platform.objects.filter(pk = platform_id, isActive = True, VendorId = vendor_id).exists():
            return JsonResponse({"error": "Platform invalid or not active"}, status = status.HTTP_400_BAD_REQUEST)

        order_id = koms_order.master_order.pk

        customer = koms_order.master_order.customerId

        address = customer.address_set.first()

        customer_details = {
            "FirstName": customer.FirstName or "",
            "LastName": customer.LastName or "",
            "Phone_Number": customer.Phone_Number or "",
            "Email": customer.Email or "",
            "loyalty_points_balance": customer.loyalty_points_balance,
            "Shipping_Address": {
                "address_line_1": address.address_line1 if address else "",
                "address_line_2": address.address_line2 if address else "",
                "city": address.city if address else "",
                "state": address.state if address else "",
                "country": address.country if address else "",
                "zipcode": address.zipcode if address else "",
                "type": address.type if address else "",
                "is_selected": address.is_selected if address else ""
            }
        }

        payment = koms_order.master_order.orderpayment_set.first()

        payment_details = {
            "paymentId": payment.pk if payment else 0,
            "paymentBy": payment.paymentBy if payment else "",
            "paymentKey": payment.paymentKey if payment else "",
            "amount_paid": payment.paid if payment else 0.0,
            "paymentType": payment_type_english.get(payment.type, "Cash") if payment else "Cash",
            "paymentStatus": payment.status if payment else False,
            "splitType": payment.splitType if payment else None,
            "split_payments": []
        }

        split_orders = Order.objects.filter(masterOrder = koms_order.master_order) \
            .prefetch_related('orderpayment_set', 'splitorderitem_set__order_content_id')
        
        for split_order in split_orders:
            split_payment = split_order.orderpayment_set.first()

            if split_payment:
                split_items = []

                for item in split_order.splitorderitem_set.all():
                    product_info = Product.objects.filter(
                        PLU = item.order_content_id.SKU, vendorId = vendor_id
                    ).values("pk", "productPrice").first()
                    
                    images = ProductImage.objects.filter(product_id = product_info["pk"]).values_list('url', flat=True)

                    image_list = list(images) if images else ['https://www.stockvault.net/data/2018/08/31/254135/preview16.jpg']
                    
                    modifiers = []

                    order_modifiers = Order_modifer.objects.filter(contentID = item.order_content_id.pk).values("SKU", "quantity")

                    for modifier in order_modifiers:
                        modifier_info = ProductModifier.objects.filter(modifierPLU__in = modifier["SKU"], vendorId = vendor_id) \
                            .values("pk", "modifierName", "modifierPrice").first()
                        
                        modifiers.append({
                            "modifer_id": modifier_info["pk"],
                            "modifer_name": modifier_info["modifierName"],
                            "modifer_quantity": modifier["quantity"],
                            "modifer_price": modifier_info["modifierPrice"],
                            "order_content_id": item.order_content_id.pk,
                        })
                    
                    split_items.append({
                        "order_content_id": item.order_content_id.pk,
                        "order_content_name": item.order_content_id.name,
                        "order_content_quantity": item.order_content_qty,
                        "order_content_price": product_info["productPrice"],
                        "order_content_images": image_list,
                        "order_content_modifer": modifiers
                    })
                
                payment_mode = payment_type_english[split_payment.type]
                
                if language != "English":
                    payment_mode = language_localization[payment_type_english[split_payment.type]]
                
                payment_details["split_payments"].append({
                    "paymentId": split_payment.pk,
                    "paymentBy": f"{split_order.customerId.FirstName or ''} {split_order.customerId.LastName or ''}",
                    "customer_name": f"{split_order.customerId.FirstName or ''} {split_order.customerId.LastName or ''}",
                    "customer_mobile": split_order.customerId.Phone_Number,
                    "customer_email": split_order.customerId.Email or "",
                    "paymentKey": split_payment.paymentKey,
                    "amount_paid": split_payment.paid,
                    "paymentType": split_payment.type,
                    "paymentSplitPer": (split_payment.paid / koms_order.master_order.TotalAmount) * 100,
                    "paymentStatus": split_payment.status,
                    "amount_subtotal": split_order.subtotal,
                    "amount_tax": split_order.tax,
                    "status": split_payment.status,
                    "platform": split_payment.platform,
                    "mode": payment_mode,
                    "splitType": split_payment.splitType,
                    "splitItems": split_items
                })

        table_details = []
        
        tables = koms_order.order_tables_set.filter(orderId__externalOrderId = external_order_id) \
            .select_related("tableId", "tableId__waiterId").values(
                "tableId__pk", "tableId__tableNumber", "tableId__status", "tableId__tableCapacity", "tableId__guestCount",
                "tableId__waiterId__pk", "tableId__waiterId__name", "tableId__waiterId__name_locale"
            )

        for table in tables:
            table_details.append({
                "tableId": table["tableId__pk"],
                "tableNumber": table["tableId__tableNumber"],
                "status": table["tableId__status"],
                "tableCapacity": table["tableId__tableCapacity"],
                "guestCount": table["tableId__guestCount"],
                "waiterId": table["tableId__waiterId__pk"],
                "waiterName": table["tableId__waiterId__name"] if language == "English" else table["tableId__waiterId__name_locale"],
            })
            
        order_items = []

        koms_order_details = koms_order.order_content_set.filter(orderId = koms_order.pk) \
            .values("pk", "SKU", "quantity", "note", "status")

        for content in koms_order_details:
            product = ProductCategoryJoint.objects.filter(product__PLU = content["SKU"], product__vendorId = vendor_id) \
                .select_related("category", "product").values(
                    "product__pk", "product__productName", "product__productName_locale", "product__taxable",
                    "product__productPrice", "category__categoryName", "category__categoryName_locale"
                ).first()

            images = ProductImage.objects.filter(product_id = product["product__pk"]).values_list('url', flat = True)
            
            image_list = list(images) if images else ['https://www.stockvault.net/data/2018/08/31/254135/preview16.jpg']
            
            modifier_groups = defaultdict(list)

            modifiers = Order_modifer.objects.filter(contentID = content["pk"], status = "1").values("SKU", "quantity")
            
            for modifier in modifiers:
                modifier_info = ProductModifierAndModifierGroupJoint.objects.filter(
                    modifier__modifierPLU = modifier["SKU"]
                ).select_related("modifier", "modifierGroup").values(
                    "modifier__pk", "modifier__modifierName", "modifier__modifierName_locale", "modifier__modifierPLU",
                    "modifier__modifierSKU", "modifier__modifierPrice", "modifier__modifierImg", "modifierGroup__pk",
                ).first()

                modifier_groups[modifier_info["modifierGroup__pk"]].append({
                    'modifierId': modifier_info["modifier__pk"],
                    'name': modifier_info["modifier__modifierName_locale"] if language != "English" else modifier_info["modifier__modifierName"],
                    'plu': modifier_info["modifier__modifierPLU"],
                    'sku': modifier_info["modifier__modifierSKU"],
                    'quantity': modifier["quantity"],
                    'cost': modifier_info["modifier__modifierPrice"],
                    'status': True,
                    'image': modifier_info["modifier__modifierImg"]
                })
            
            modifiers_group = []

            for group_id, modifiers in modifier_groups.items():
                group = ProductModifierGroup.objects.get(pk = group_id)

                modifiers_group.append({
                    "id": group.pk,
                    "name": group.name_locale if language != "English" else group.name,
                    "plu": group.PLU,
                    "min": group.min,
                    "max": group.max,
                    "active": group.active,
                    "modifiers": modifiers
                })
            
            order_items.append({
                "order_content_id": content["pk"],
                "productId": product["product__pk"],
                "quantity": content["quantity"],
                "plu": content["SKU"],
                "name": product["product__productName_locale"] if language != "English" else product["product__productName"],
                "isTaxable": product["product__taxable"],
                "imagePath": image_list[0],
                "images": image_list,
                "categoryName": product["category__categoryName_locale"] if language != "English" else product["category__categoryName"],
                "note": content["note"] or "",
                "cost": product["product__productPrice"],
                "status": int(content["status"]),
                "modifiersGroup": modifiers_group
            })
            
        total_points_redeemed = 0
        
        loyalty_program = LoyaltyProgramSettings.objects.filter(is_active = True, vendor = vendor_id).first()

        if loyalty_program:
            loyalty_points_redeem_history = LoyaltyPointsRedeemHistory.objects.filter(
                customer = koms_order.master_order.customerId.pk, order = order_id
            )

            if loyalty_points_redeem_history.exists():
                total_points_redeemed = loyalty_points_redeem_history.aggregate(Sum('points_redeemed'))['points_redeemed__sum']

                if not total_points_redeemed:
                    total_points_redeemed = 0

        order_info = {}
        
        order_info["staging_orderId"] = int(koms_order.pk)
        order_info["core_orderId"] = order_id
        order_info["orderId"] = external_order_id
        order_info["customerNote"] = koms_order.order_note or ""
        order_info["tax"] = koms_order.master_order.tax or 0.0
        order_info["discount"] = koms_order.master_order.discount or 0.0
        order_info["delivery_charge"] = koms_order.master_order.delivery_charge
        order_info["subtotal"] = koms_order.master_order.subtotal
        order_info["total_points_redeemed"] = total_points_redeemed
        order_info["finalTotal"] = koms_order.master_order.TotalAmount
        order_info["order_datetime"] = koms_order.master_order.OrderDate
        order_info["arrival_time"] = koms_order.master_order.arrivalTime
        order_info["type"] = koms_order.order_type
        order_info["products"] = order_items
        
        if payment_id:
            payment = OrderPayment.objects.filter(pk = payment_id).select_related("orderId").values(
                "pk", "paymentBy", "paymentKey", "paid", "type", "status", "orderId__pk", "orderId__externalOrderId",
                "orderId__tax", "orderId__discount", "orderId__delivery_charge", "orderId__subtotal", "orderId__TotalAmount"
            ).last()

            payment_details["paymentId"] = payment["pk"] or 0
            payment_details["paymentBy"] = payment["paymentBy"] or ""
            payment_details["paymentKey"] = payment["paymentKey"] or ""
            payment_details["amount_paid"] = payment["paid"] or 0.0
            payment_details["paymentType"] = payment_type_english[payment["type"]] if payment else "Cash"
            payment_details["paymentStatus"] = payment["status"] or False
            
            order_info["core_orderId"] = payment["orderId__pk"]
            order_info["orderId"] = payment["orderId__externalOrderId"]
            order_info["tax"] = payment["orderId__tax"] or 0.0
            order_info["discount"] = payment["orderId__discount"] or 0.0
            order_info["delivery_charge"] = payment["orderId__delivery_charge"]
            order_info["subtotal"] = payment["orderId__subtotal"]
            order_info["finalTotal"] = payment["orderId__TotalAmount"]
            
        platform_details = {
            "id": koms_order.master_order.platform.pk,
            "name": koms_order.master_order.platform.Name if language == "English" else koms_order.master_order.platform.Name_locale
        }

        return JsonResponse({
            "payment_details": payment_details,
            "customer_details": customer_details,
            "table_details": table_details,
            "platform_details": platform_details,
            "order_info": order_info
        })

    except Exception as e:
        return JsonResponse({"error": str(e)}, status = status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST"])
def update_order_content(request):
    vendor_id = request.GET['vendorId']
    external_order_id = request.GET['id']
    language = request.GET.get("language", "English")

    if not vendor_id:
        return JsonResponse({"message": "Vendor ID cannot be empty"}, status = status.HTTP_400_BAD_REQUEST)
    
    if not external_order_id:
        return JsonResponse({"message": "External Order ID cannot be empty"}, status = status.HTTP_400_BAD_REQUEST)

    request_data = request.data

    if not request_data:
        return JsonResponse({"message": "JSON data cannot be empty"}, status = status.HTTP_400_BAD_REQUEST)
    
    with transaction.atomic():
        try:
            core_id = request_data['order_info'].get('core_orderId')

            if core_id:
                core_order = Order.objects.get(pk = core_id)

                if core_order:
                    core_order.Notes = request_data['order_info'].get('customerNote', "")
                    core_order.subtotal = request_data['order_info'].get('subtotal')
                    core_order.tax = request_data['order_info'].get('tax')
                    core_order.discount = request_data['order_info'].get('discount')
                    core_order.delivery_charge = request_data['order_info'].get('delivery_charge')
                    core_order.TotalAmount = request_data['order_info'].get('finalTotal')
                    core_order.save()
            
            new_data = []

            for product in request_data["order_info"]["products"]:
                if product.get("order_content_id"):
                    order_detail = {
                        "order_content_id": product["order_content_id"],
                        "name": product["name"],
                        "quantity": product["quantity"],
                        "note": product["note"],
                        "SKU": product["plu"],
                        "categoryName": product["categoryName"],
                        "modifiers": []
                    }

                    if product.get("modifiersGroup"):
                        for modifier_group in product["modifiersGroup"]:
                            for modifier in modifier_group["modifiers"]:
                                if modifier["status"] == True:
                                    mod = {
                                        "order_modifier_id":modifier["order_modifier_id"],
                                        "name": modifier["name"],
                                        "SKU": modifier["plu"],
                                        "quantity": modifier["quantity"]
                                    }
                                
                                    order_detail["modifiers"].append(mod)

                    new_data.append(order_detail)

            modifier_details = defaultdict(list)

            old_data = []

            koms_order = KOMSOrder.objects.get(externalOrderId = external_order_id, vendorId = vendor_id)

            old_status_of_order = koms_order.order_status
            
            koms_order_id = koms_order.pk

            koms_order.order_note = request_data.get('order_info').get('customerNote')

            koms_order.is_edited = True

            koms_order.save()

            koms_order_content = Order_content.objects.filter(orderId = koms_order_id)

            for order in koms_order_content:
                koms_order_modifiers = Order_modifer.objects.filter(contentID = order.pk)

                for modifier in koms_order_modifiers:
                    modifier_details[order.pk].append({
                        "order_modifier_id": modifier.pk,
                        "name": modifier.name,
                        "quantityStatus": modifier.quantityStatus,
                        "SKU": modifier.SKU,
                        "status": modifier.status,
                        "quantity": modifier.quantity,
                        "isEdited": modifier.isEdited
                    })

                old_data.append({
                    "order_content_id": order.pk,
                    "name": order.name,
                    "quantity": order.quantity,
                    "quantityStatus": order.quantityStatus,
                    "note": order.note,
                    "SKU": order.SKU,
                    "categoryName": order.categoryName,
                    "stationId": order.stationId.pk,
                    "status": order.status,
                    "isEdited": order.isEdited,
                    "modifiers": modifier_details[order.pk]
                })

            old_products = {}
            new_products = {}
            deleted_products = set()

            for product in old_data:
                order_content_id = product.get('order_content_id')
                sku = product.get('SKU')

                old_products[(order_content_id, sku)] = []

            for product in new_data:
                order_content_id = product.get('order_content_id')
                sku = product.get('SKU')

                new_products[(order_content_id, sku)] = []

            for key in old_products:
                if key not in new_products:
                    deleted_products.add(key)
            
            for key in deleted_products:
                product = Order_content.objects.get(pk = key[0], SKU = key[1])
                product.quantityStatus = 1
                product.status = 5

                modifiers = Order_modifer.objects.filter(contentID = product.pk).update(quantityStatus = 0, status = 5) # quantityStatus=0 is deleted

                product.save()

            for item1 in old_data:
                for item2 in new_data:
                    if item1["SKU"] == item2["SKU"]:
                        if item1["quantity"] != item2["quantity"]:
                            product = Order_content.objects.get(pk = item2["order_content_id"], SKU = item2["SKU"])

                            product.quantity = item2["quantity"]

                            contdata = {
                                "contentID": product.pk,
                                "update_time": datetime.today().strftime("%Y-%m-%d"),
                                "quantity": item1["quantity"],
                                "unit": "qty"
                            }

                            cont = Content_history_serializer(data = contdata, partial = True)

                            if cont.is_valid():
                                cont.save()

                            product.isEdited = True
                            product.save()

                        if item1["note"] != item2["note"]:
                            product = Order_content.objects.get(pk = item2["order_content_id"], SKU = item2["SKU"])

                            product.note = item2["note"]

                            product.isEdited = True
                            product.save()

            for product in request_data["order_info"]["products"]:
                if product.get("order_content_id") == None:
                    koms_order_details = KOMSOrder.objects.get(pk = request_data["order_info"].get("staging_orderId"))

                    if koms_order_details.order_status == 1:
                        order_status = 1

                    else:
                        order_status = 8
                    
                    product_category_joint = ProductCategoryJoint.objects.filter(
                        product__PLU = product["plu"], product__vendorId = vendor_id, vendorId = vendor_id
                    ).select_related("category").first()

                    if product_category_joint:
                        category_station = product_category_joint.category.categoryStation

                    else:
                        category_station = Station.objects.filter(vendorId = vendor_id).first()
                    
                    item = Order_content.objects.create(
                        orderId = koms_order_details,
                        name = product["name"],
                        quantity = product["quantity"],
                        quantityStatus = 0,
                        unit = "qty",
                        note = product["note"],
                        SKU = product["plu"],
                        tag = 1,
                        categoryName = product["categoryName"],
                        stationId = category_station,
                        status = order_status,
                        isrecall = False,
                        isEdited = 0,
                    )

                    if koms_order_details.order_status != 1:
                        koms_order_details.order_status = 8
                        koms_order_details.save()

                    modifiers_group = product.get('modifiersGroup', [])

                    if modifiers_group:
                        for mod_group in modifiers_group:
                            modifiers = mod_group.get('modifiers', [])

                            for modifier in modifiers:
                                if modifier["status"] == True:
                                    item_modifier = Order_modifer.objects.create(
                                        contentID = Order_content.objects.get(pk = item.pk),
                                        name = modifier.get('name', ""),
                                        quantityStatus = 1, # not deleted
                                        unit = "qty",
                                        note = modifier.get('note', ""),
                                        SKU = modifier.get('plu'),
                                        status = 1,
                                        quantity = modifier.get('quantity', 0),
                                        isEdited = 0,
                                        group = mod_group.get('id')
                                    )

            if Platform.objects.filter(Name = "Inventory", isActive = True, VendorId = vendor_id).exists():
                sync_order_content_with_inventory(core_id, vendor_id)
            
            koms_order = KOMSOrder.objects.get(externalOrderId = external_order_id, vendorId = vendor_id)

            new_status_of_order = koms_order.order_status
            
            webSocketPush(
                message = {"id": koms_order_id, "orderId": koms_order.externalOrderId, "UPDATE": "REMOVE",},
                room_name = str(vendor_id) + '-' + str(old_status_of_order),
                username = "CORE",
            ) 

            allStationWiseRemove(
                id = koms_order_id, old = str(old_status_of_order), current = str(new_status_of_order), vendorId = vendor_id
            )

            allStationWiseSingle(id = koms_order_id, vendorId = vendor_id)
            waiteOrderUpdate(orderid = koms_order_id, language = language, vendorId = vendor_id)
            allStationWiseCategory(vendorId = vendor_id)
            
            return JsonResponse({"message": "Success"}, status = status.HTTP_200_OK)

        except Exception as e:
            transaction.set_rollback(True)
            return JsonResponse({"message": str(e)}, status = status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST", "PATCH"])
def split_order_and_payment(request):
    request_data = request.data

    vendorId = request.GET.get('vendorId')
    language = request.GET.get('language', 'English')

    if not vendorId:
        return Response("Vendor ID empty", status = status.HTTP_400_BAD_REQUEST)
    
    try:
        vendorId = int(vendorId)

    except ValueError:
        return Response("Invalid Vendor ID", status = status.HTTP_400_BAD_REQUEST)
    
    vendor_instance = Vendor.objects.filter(pk = vendorId).first()

    if not vendor_instance:
        return Response("Vendor does not exist", status = status.HTTP_400_BAD_REQUEST)
    
    order = KOMSOrder.objects.get(externalOrderId = request_data['orderid'])

    master_order = Order.objects.filter(pk = order.master_order.pk).last()

    payment = OrderPayment.objects.filter(orderId = master_order.pk).last()

    if not payment:
        return Response("Payment record does not exist", status = status.HTTP_400_BAD_REQUEST)
    
    payment.type = payment_type_number["Split"]

    payment.splitType = request_data.get('splitBy', None)

    payment.save()

    Order.objects.filter(masterOrder = master_order.pk).delete()

    count = 1

    for splitPayment in request_data["payments"]:
        split_customer = Customer.objects.filter(pk = splitPayment.get('customerId')) if splitPayment.get('customerId') else None

        split_order = Order(
            Status = master_order_status_number["Open"],
            masterOrder = master_order,
            TotalAmount = splitPayment.get("amount_paid", 0.0),
            OrderDate = timezone.now(),
            Notes = request_data.get("note"),
            externalOrderId = master_order.externalOrderId + f"_{count}",
            orderType = master_order.orderType,
            arrivalTime = timezone.now(),
            tax = splitPayment.get("amount_tax") or 0.0 ,
            discount = 0.0,
            tip = 0.0,
            delivery_charge = 0.0,
            subtotal = (splitPayment.get("amount_paid", 0.0)) - (splitPayment.get("amount_tax") or 0.0) ,
            customerId = split_customer.first() if split_customer and split_customer.exists() else master_order.customerId,
            vendorId = vendor_instance,
            platform = master_order.platform
        ).save()
        
        count = count + 1

        payment_type = splitPayment.get("paymentType", "Cash")

        payment_type = payment_type.capitalize()

        if request_data.get('splitBy', None) == "by_product":
            for item in splitPayment.get("splitItems",[]):
                order_content = Order_content.objects.get(pk = item['order_content_id'])

                SplitOrderItem(
                    order_id = split_order,
                    order_content_id = order_content,
                    order_content_qty = item.get('order_content_qty') or order_content.quantity
                ).save()

        OrderPayment(
            orderId = split_order,
            paymentBy = master_order.customerId.Email or "",
            paymentKey = splitPayment.get("paymentKey", ""),
            paid = splitPayment.get("amount_paid", 0.0),
            due = 0.0,
            tip = splitPayment.get('tip', 0.0),
            status = splitPayment.get("paymentStatus") or False,
            type = payment_type_number[payment_type],
            platform = splitPayment.get("platform") or "",
            splitType = request_data.get('splitBy', None),
        ).save()

    waiteOrderUpdate(orderid = order.pk, language = language, vendorId = vendorId)

    return Response({"success": True})


@api_view(["POST"])
def update_store_status(request):
    vendor_id = request.GET.get("vendorId")
    store_status = request.data.get("store")

    if not vendor_id or store_status is None:
        return JsonResponse({"message": "Invalid request data"}, status = status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id = int(vendor_id)

    except:
        return JsonResponse({"message": "Invalid vendor ID"}, status = status.HTTP_400_BAD_REQUEST)
    
    if not Vendor.objects.filter(pk = vendor_id).exists():
        return JsonResponse({"message": "Vendor does not exist"}, status = status.HTTP_400_BAD_REQUEST)

    pos_store_setting = POSSetting.objects.filter(vendor = vendor_id).first()

    if not pos_store_setting:
        return JsonResponse({"message": "POS setting not created for the Vendor"}, status = status.HTTP_400_BAD_REQUEST)

    pos_store_setting.store_status = store_status
    pos_store_setting.save()

    return JsonResponse({"message": "", "store_status": pos_store_setting.store_status}, status = status.HTTP_200_OK) 


@api_view(['GET'])
def get_store_timings(request):
    vendor_id = request.GET.get("vendorId")
    language = request.GET.get("language", "English")

    if not vendor_id:
        return JsonResponse({"message": "Invalid vendor ID"}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id = int(vendor_id)

    except ValueError:
        return JsonResponse({"message": "Invalid vendor ID"}, status=status.HTTP_400_BAD_REQUEST)
    
    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor_instance:
        return JsonResponse({"message": "Vendor does not exist"}, status=status.HTTP_400_BAD_REQUEST)

    store_timings = StoreTiming.objects.filter(vendor=vendor_id)

    slot = store_timings.filter(is_active=True , day=datetime.now().strftime("%A")).first()

    pos_store_setting = POSSetting.objects.filter(vendor=vendor_id).first()

    if not pos_store_setting:
        return JsonResponse({"message": "POS setting not created for the Vendor"}, status=status.HTTP_400_BAD_REQUEST)

    store_status = pos_store_setting.store_status

    store_status_value = store_status

    if store_status_value == False:
        store_status_final = False

    if not store_timings:
        return Response({"store_status": store_status_value, "store_timing": []})
    
    else:
        if store_status == False:
            store_status_final =  False

        elif slot == None:
            store_status_final = True 

        elif (slot.open_time < datetime.now().time() < slot.close_time) and not slot.is_holiday:
            store_status_final = True  

        else:
            store_status_final =  False

    store_timing_list = []
    
    for instance in store_timings:
        if language == "English":
            day = instance.day

        else:
            day = language_localization[instance.day]

        store_timing_list.append({
            "id": instance.pk,
            "slot_identity": instance.slot_identity,
            "day": day,
            "is_holiday": instance.is_holiday,
            "is_active": instance.is_active,
            "open_time": instance.open_time.strftime("%H:%M:%S"),
            "close_time": instance.close_time.strftime("%H:%M:%S"),
            "vendor": instance.vendor.pk,
        })

    return JsonResponse({"store_status": store_status_final, "store_timing": store_timing_list})


@api_view(['POST'])
def set_store_timings(request):
    vendor = request.GET.get("vendorId")

    body_data = request.data

    body_data['platform'] = None

    serialized = StoreTimingSerializer(data=body_data)

    # update store_timings
    if request.GET.get("id"):
        data = StoreTiming.objects.get(pk=request.GET.get("id"))
        serialized = StoreTimingSerializer(instance=data, data=body_data)

    if serialized.is_valid():
        slot = serialized.save()

        return Response(serialized.data, status=status.HTTP_200_OK)
    
    print("error :", serialized._errors)
    return Response(status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST','GET','DELETE'])
def delete_store_timings(request):
    try:
        data = StoreTiming.objects.get(pk=request.GET.get("id"))

        data.delete()

    except Exception as e:
        print(e)
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    return Response(status=status.HTTP_200_OK)


@api_view(["POST"])
def delete_excel(request):
    json_data = json.loads(request.body)

    path = json_data.get("path")

    if not path:
        return JsonResponse({"error": "File path empty"}, status=400)

    absolute_file_path = os.path.join(settings.MEDIA_ROOT, path.lstrip("/media/"))

    try:
        if os.path.exists(absolute_file_path):
            os.remove(absolute_file_path)
            return JsonResponse({"message": "File deleted"}, status=200)
        else:
            return JsonResponse({"error": "File not found"}, status=404)

    except Exception as e:
        return JsonResponse({"error": f"Failed to delete file: {str(e)}"}, status=500)


@api_view(["GET",])
def get_products(request):
    product_name = request.GET.get("productName")

    try:
        vendor_id = int(request.GET.get("vendorId"))
        page_number = int(request.GET.get("page", 1))
        page_size = int(request.GET.get("page_size", 10))

        if vendor_id <= 0 or page_number <= 0 or page_size <= 0:
            return Response("Invalid request data", status = status.HTTP_400_BAD_REQUEST)

    except:
        return Response("Invalid request data", status = status.HTTP_400_BAD_REQUEST)

    if not Vendor.objects.filter(pk = vendor_id).exists():
        return Response("Vendor does not exist", status = status.HTTP_400_BAD_REQUEST)

    products = Product.objects.filter(vendorId = vendor_id).order_by('-pk')
    
    if product_name:
        products = products.filter(Q(productName__icontains = product_name) | Q(productName_locale__icontains = product_name))  

    if not products:
        return JsonResponse({
            "total_pages": 0,
            "current_page": 0,
            "page_size": 0,
            "results": [],
        }, status = status.HTTP_200_OK)
    
    paginator = Paginator(products, page_size)

    paginated_products = paginator.page(page_number)

    current_page = paginated_products.number
    
    total_pages = paginator.num_pages

    response_data = []

    for single_product in paginated_products:
        product_info = get_product_data(single_product, vendor_id)

        response_data.append(product_info)

    response = {
        "total_pages": total_pages,
        "current_page": current_page,
        "page_size": page_size,
        "results": response_data,
    }

    return JsonResponse(response, status = status.HTTP_200_OK)


@api_view(["POST"])
def create_product(request):
    plu = request.data.get('PLU')

    try:
        vendor_id = int(request.data.get('vendorId'))

        if vendor_id <= 0:
            return Response("Invalid Vendor ID", status = status.HTTP_400_BAD_REQUEST)

    except:
        return Response("Invalid Vendor ID", status = status.HTTP_400_BAD_REQUEST)

    if not Vendor.objects.filter(pk=vendor_id).exists():
        return Response("Vendor does not exist", status = status.HTTP_400_BAD_REQUEST)

    existing_product = Product.objects.filter(PLU = plu, vendorId = vendor_id).first()

    if existing_product:
        return Response({'error': 'Product with this PLU already exists.'}, status = status.HTTP_400_BAD_REQUEST)

    product_serializer = ProductSerializer(data = request.data)

    if product_serializer.is_valid():
        categories_data = []
        images_data = []
        modifier_groups_data = []

        image_path_keys = []
        is_image_selected_keys = []

        with transaction.atomic():
            product = product_serializer.save()

            vendor_id = int(request.data.get('vendorId'))

            for key, value in request.data.items():
                if key.startswith('categories-'):
                    categories_data.append({'category': int(value), 'product': product.pk, 'vendorId': vendor_id})
                
                elif key.startswith('modifier_groups-'):
                    modifier_groups_data.append({'modifierGroup': int(value), 'product': product.pk, 'vendorId': vendor_id})

            for key in request.data.keys():
                if key.startswith('images-'):
                    if key.endswith('-image'):
                        image_path_keys.append(key)
                    
                    if key.endswith('-is_image_selected'):
                        is_image_selected_keys.append(key)

            if len(image_path_keys) == len(is_image_selected_keys):
                for iterator in range(len(image_path_keys)):
                    image_path = request.data.get(image_path_keys[iterator], None)
                    is_image_selected = request.data.get(is_image_selected_keys[iterator], None)

                    if (image_path == None):
                        pass

                    else:
                        try:
                            validator = URLValidator()

                            validator(image_path)

                        except Exception as e:
                            return Response({'error': 'Invalid Image URL'}, status=status.HTTP_400_BAD_REQUEST)
                        
                        images_data.append({'url': image_path, 'is_url_selected':is_image_selected, 'product': product.pk, 'vendorId': vendor_id})
            
            else:
                transaction.set_rollback(True)
                return Response("Image keys do not match", status=status.HTTP_400_BAD_REQUEST)
            
            for category_data in categories_data:
                product_category_joint_serializer = ProductCategoryJointSerializer(data=category_data)
                
                if product_category_joint_serializer.is_valid():
                    product_category_joint = product_category_joint_serializer.save()

                else:
                    print(product_category_joint_serializer.errors)

            for image_data in images_data:
                image_serializer = ProductImagesSerializer(data=image_data)

                if image_serializer.is_valid():
                    images = image_serializer.save()
                    
                else:
                    print("image save error :: ", image_serializer.errors)
                    transaction.set_rollback(True)
                    return Response("Not a valid image", status=status.HTTP_400_BAD_REQUEST)
            
            for modifier_group_data in modifier_groups_data:
                product_modGroup_joint_serializer = ProductModGroupJointSerializer(data=modifier_group_data)

                if product_modGroup_joint_serializer.is_valid():
                    product_modGroup_joint = product_modGroup_joint_serializer.save()

                else:
                    print(product_modGroup_joint_serializer.errors)
                    transaction.set_rollback(True)
                    return Response("Not a valid Modifier group", status=status.HTTP_400_BAD_REQUEST)

            inventory_platform = Platform.objects.filter(Name="Inventory", isActive=True, VendorId=vendor_id).first()
                
            if inventory_platform:
                sync_status = single_product_sync_with_odoo(product)
                    
                if sync_status == 0:
                    notify(type=3, msg='0', desc='Product did not synced with Inventory', stn=['POS'], vendorId=vendor_id)
                
                else:
                    notify(type=3, msg='0', desc='Product synced with Inventory', stn=['POS'], vendorId=vendor_id)
            
            product_info = get_product_data(product, vendor_id)
            
            return JsonResponse(product_info, status=status.HTTP_201_CREATED)
    
    return Response(product_serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["PUT", "PATCH"])
def update_product(request, product_id):
    try:
        vendor_id = int(request.POST.get('vendorId', None))

    except:
        return Response("Invalid Vendor ID", status = status.HTTP_400_BAD_REQUEST)

    vendor = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor:
        return Response("Vendor does not exist", status = status.HTTP_400_BAD_REQUEST)

    if product_id:
        product = Product.objects.filter(pk=product_id, vendorId=vendor_id).first()

        if product:
            product_serializer = ProductSerializer(product, data=request.data)

            # Check if a product with the new PLU already exists while editing
            new_plu = request.data.get('PLU')
            
            existing_product_with_new_plu = Product.objects.filter(PLU=new_plu, vendorId=vendor_id).exclude(pk=product_id).first()

            if existing_product_with_new_plu:
                return Response({'error': 'Product with this PLU already exists.'}, status=status.HTTP_400_BAD_REQUEST)
            
            if product_serializer.is_valid():
                category_ids = []
                modifier_group_ids = []
                images_data = []

                image_path_keys = []
                is_image_selected_keys = []

                with transaction.atomic():
                    updated_product = product_serializer.save()

                    for key, value in request.data.items():
                        if key.startswith('categories-'):
                            category_ids.append(int(value))
            
                        if key.startswith('modifier_groups-'):
                            modifier_group_ids.append(int(value))
                    
                        if key.startswith('images-'):
                            if key.endswith('-image'):
                                image_path_keys.append(key)
                            if key.endswith('-is_image_selected'):
                                is_image_selected_keys.append(key)

                    ProductImage.objects.filter(product=updated_product.pk, vendorId=vendor_id).delete()
                    
                    if len(image_path_keys) == len(is_image_selected_keys):
                        for iterator in range(len(image_path_keys)):
                            image_path = request.data.get(image_path_keys[iterator], None)
                            
                            is_image_selected = request.data.get(is_image_selected_keys[iterator], None)
                            
                            if ProductImage.objects.filter(product=updated_product.pk, image=image_path, is_image_selected=is_image_selected, vendorId=vendor_id).exists():
                                pass

                            else:
                                try:
                                    validator = URLValidator()

                                    validator(image_path)

                                except Exception as e:
                                    return Response({'error': 'Invalid Image URL'}, status=status.HTTP_400_BAD_REQUEST)
                                
                                images_data.append({"product":updated_product.pk, "url":image_path, "is_url_selected":is_image_selected, "vendorId":vendor_id})

                    else:
                        transaction.set_rollback(True)
                        return Response("Image keys do not match", status=status.HTTP_400_BAD_REQUEST)
                    
                    for image_data in images_data:
                        product_image_serializer = ProductImagesSerializer(data=image_data)
                        
                        if product_image_serializer.is_valid():
                            img=product_image_serializer.save()
                            
                        else:
                            print(product_image_serializer.errors)
                    
                    product_category_joint_ids = list(ProductCategoryJoint.objects.filter(product=updated_product.pk, vendorId=vendor_id).values_list('category_id', flat=True))
                    
                    category_ids_not_in_joint = []

                    for category_id in category_ids:
                        if category_id not in product_category_joint_ids:
                            category_ids_not_in_joint.append(category_id)

                    for category_id in category_ids_not_in_joint:
                        product_instance = Product.objects.filter(pk=updated_product.pk, vendorId=vendor_id).first()
                        category_instance = ProductCategory.objects.filter(pk=category_id).first()
                        vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

                        ProductCategoryJoint.objects.create(product=product_instance, category=category_instance, vendorId=vendor_instance)

                    missing_category_ids = list(set(product_category_joint_ids) - set(category_ids))

                    for category_id in missing_category_ids:
                        ProductCategoryJoint.objects.filter(product=updated_product.pk, category=category_id, vendorId=vendor_id).delete()
                    
                    product_modgroup_joint_ids = list(ProductAndModifierGroupJoint.objects.filter(product=updated_product.pk, vendorId=vendor_id).values_list('modifierGroup_id', flat=True))
                    
                    modifier_group_ids_not_in_joint = []

                    for group_id in modifier_group_ids:
                        if group_id not in product_modgroup_joint_ids:
                            modifier_group_ids_not_in_joint.append(group_id)

                    for group_id in modifier_group_ids_not_in_joint:
                        product_instance = Product.objects.filter(pk=updated_product.pk, vendorId=vendor_id).first()
                        modifier_group_instance = ProductModifierGroup.objects.filter(pk=group_id).first()
                        vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

                        ProductAndModifierGroupJoint.objects.create(product=product_instance, modifierGroup=modifier_group_instance, vendorId=vendor_instance)
                    
                    missing_modifier_group_ids = list(set(product_modgroup_joint_ids) - set(modifier_group_ids))

                    for group_id in missing_modifier_group_ids:
                        ProductAndModifierGroupJoint.objects.filter(product=updated_product.pk, modifierGroup=group_id, vendorId=vendor_id).delete()
                    
                    inventory_platform = Platform.objects.filter(Name="Inventory", isActive=True, VendorId=vendor_id).first()
                
                    if inventory_platform:
                        sync_status = single_product_sync_with_odoo(product)
                            
                        if sync_status == 0:
                            notify(type=3, msg='0', desc='Product did not synced with Inventory', stn=['POS'], vendorId=vendor_id)
                        
                        else:
                            notify(type=3, msg='0', desc='Product synced with Inventory', stn=['POS'], vendorId=vendor_id)
                    
                    product = Product.objects.filter(pk=product_id, vendorId=vendor_id).first()

                    product_info = get_product_data(product, vendor_id)
                    
                    return JsonResponse(product_info, status=status.HTTP_200_OK)
                
            else:
                return Response(product_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        else:
            return Response("Product does not exist", status=status.HTTP_404_NOT_FOUND)
    
    else:
        return Response("Product ID empty", status=status.HTTP_400_BAD_REQUEST)


@api_view(["DELETE"])
def delete_product(request, product_id):
    try:
        vendor_id = int(request.GET.get('vendorId'))
        product_id = int(product_id)

        if vendor_id <= 0 or product_id <=0:
            return Response("Invalid Vendor ID or Product ID", status = status.HTTP_400_BAD_REQUEST)

    except:
        return Response("Invalid Vendor ID or Product ID", status = status.HTTP_400_BAD_REQUEST)

    if not Vendor.objects.filter(pk = vendor_id).exists():
        return Response("Vendor does not exist", status = status.HTTP_400_BAD_REQUEST)
    
    product = Product.objects.filter(pk = product_id, vendorId = vendor_id).first()

    if not product:
        return Response("Product not found", status = status.HTTP_400_BAD_REQUEST)

    with transaction.atomic():
        try:
            inventory_platform_url = Platform.objects.filter(
                Name = "Inventory", isActive = True, VendorId = vendor_id
            ).values_list("baseUrl", flat = True).first()

            if inventory_platform_url:
                delete_status, error_message, request_data = delete_product_in_odoo(inventory_platform_url, product.PLU, vendor_id)
            
                if delete_status == 0:
                    notify(type = 3, msg = '0', desc = 'Category did not synced with Inventory', stn = ['POS'], vendorId = vendor_id)
                
                else:
                    notify(type = 3, msg = '0', desc = 'Category synced with Inventory', stn = ['POS'], vendorId = vendor_id)
            
            product.delete()

            return Response(status = status.HTTP_204_NO_CONTENT)
            
        except Exception as e:
            transaction.set_rollback(True)

            return Response({'error': str(e)}, status = status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET",])
def get_modifiers(request):
    modifier_name = request.GET.get("modifierName")

    try:
        vendor_id = int(request.GET.get("vendorId"))
        page_number = int(request.GET.get("page", 1))
        page_size = int(request.GET.get("page_size", 10))

        if vendor_id <= 0 or page_number <= 0 or page_size <= 0:
            return Response("Invalid request data", status = status.HTTP_400_BAD_REQUEST)

    except:
        return Response("Invalid Vendor ID", status = status.HTTP_400_BAD_REQUEST)

    if not Vendor.objects.filter(pk = vendor_id).exists():
        return Response("Vendor does not exist", status = status.HTTP_400_BAD_REQUEST)

    modifiers = ProductModifier.objects.filter(vendorId = vendor_id).order_by('-pk')
    
    if modifier_name:
        modifiers = modifiers.filter(
            Q(modifierName__icontains = modifier_name, vendorId = vendor_id) | Q(modifierName_locale__icontains = modifier_name)
        )

    if not modifiers:
        return JsonResponse({
            "total_pages": 0,
            "current_page": 0,
            "page_size": 0,
            "results": [],
        }, status = status.HTTP_200_OK)
    
    paginator = Paginator(modifiers, page_size)

    paginated_modifiers = paginator.page(page_number)

    current_page = paginated_modifiers.number

    total_pages = paginator.num_pages

    response_data = []

    for single_modifier in paginated_modifiers:
        modifier_info = get_modifier_data(single_modifier, vendor_id)

        response_data.append(modifier_info)

    return JsonResponse({
        "total_pages": total_pages,
        "current_page": current_page,
        "page_size": page_size,
        "results": response_data,
    }, status = status.HTTP_200_OK)

        
@api_view(["POST"])
def create_modifier(request):
    plu = request.data.get('modifierPLU')
    vendor_id = request.data.get('vendorId')

    if vendor_id is None:
        return Response("Vendor ID empty", status=status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id = int(vendor_id)
    
    except ValueError:
        return Response("Invalid Vendor ID", status=status.HTTP_400_BAD_REQUEST)

    vendor = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor:
        return Response("Vendor does not exist", status=status.HTTP_400_BAD_REQUEST)

    existing_modifier = ProductModifier.objects.filter(modifierPLU=plu, vendorId=vendor_id).first()

    if existing_modifier:
        return JsonResponse({'error':'Modifier with this PLU already exists'}, status=status.HTTP_400_BAD_REQUEST)
            
    try:
        image_url = request.data.get('modifierImg')

        if image_url:
            validator = URLValidator()

            validator(image_url)

    except Exception as e:
        return Response({'error': 'Invalid Image URL'}, status=status.HTTP_400_BAD_REQUEST)
    
    modifier_serializer = ModifierSerializer(data=request.data)

    if modifier_serializer.is_valid():
        modifier_groups_ids = []

        with transaction.atomic():
            modifier = modifier_serializer.save()

            vendor_id = int(request.data.get('vendorId'))

            for key, value in request.data.items():
                if key.startswith('modifier_groups-'):
                    modifier_groups_ids.append(int(value))
            
            for group_id in modifier_groups_ids:
                modifier_instance = modifier
                modifier_group_instance = ProductModifierGroup.objects.filter(pk=group_id, vendorId=vendor_id).first()
                vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

                ProductModifierAndModifierGroupJoint.objects.create(modifier=modifier_instance, modifierGroup=modifier_group_instance, vendor=vendor_instance)

            inventory_platform = Platform.objects.filter(Name="Inventory", isActive=True, VendorId=vendor_id).first()
                
            if inventory_platform:
                sync_status = single_modifier_sync_with_odoo(modifier)
                    
                if sync_status == 0:
                    notify(type=3, msg='0', desc='Modifier did not synced with Inventory', stn=['POS'], vendorId=vendor_id)
                
                else:
                    notify(type=3, msg='0', desc='Modifier synced with Inventory', stn=['POS'], vendorId=vendor_id)
            
            modifier_info = get_modifier_data(modifier, vendor_id)
            
            return JsonResponse(modifier_info, status=status.HTTP_201_CREATED)
    
    else:
        return Response(modifier_serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["PUT", "PATCH"])
def update_modifier(request, modifier_id):
    request_data = request.data
    
    vendor_id = request_data.get('vendorId')

    if vendor_id is None:
        return Response("Vendor ID empty", status=status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id = int(vendor_id)

    except ValueError:
        return Response("Invalid Vendor ID", status=status.HTTP_400_BAD_REQUEST)

    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor_instance:
        return Response("Vendor does not exist", status=status.HTTP_400_BAD_REQUEST)
    
    if not modifier_id:
        return Response("Modifier ID empty", status=status.HTTP_400_BAD_REQUEST)
    
    modifier_instance = ProductModifier.objects.filter(pk=modifier_id, vendorId=vendor_id).first()

    if not modifier_instance:
        return Response("Modifier does not exist", status=status.HTTP_404_NOT_FOUND)
    
    try:
        image_url = request.data.get('modifierImg')

        if image_url:
            validator = URLValidator()

            validator(image_url)

    except Exception as e:
        return Response({'error': 'Invalid Image URL'}, status=status.HTTP_400_BAD_REQUEST)
    
    modifier_serializer = ModifierSerializer(modifier_instance, data=request.data)

    new_plu = request.data.get('modifierPLU')

    existing_modifier_with_new_plu = ProductModifier.objects.filter(modifierPLU=new_plu, vendorId=vendor_id).exclude(pk=modifier_id).first()

    if existing_modifier_with_new_plu:
        return JsonResponse({'error': 'Modifier with this PLU already exists.'}, status=status.HTTP_400_BAD_REQUEST)
    
    if modifier_serializer.is_valid():
        received_modifier_group_ids = []
        
        with transaction.atomic():
            updated_modifier = modifier_serializer.save()

            for key, value in request.data.items():
                if key.startswith('modifier_groups-'):
                    received_modifier_group_ids.append(int(value))

            received_modifier_group_ids = set(received_modifier_group_ids)

            existing_modifier_group_ids = set(
                ProductModifierAndModifierGroupJoint.objects.filter(modifier=updated_modifier.pk, vendor=vendor_id).values_list("modifierGroup", flat=True)
            )

            new_modifier_group_ids = received_modifier_group_ids - existing_modifier_group_ids

            deleted_modifier_group_ids = existing_modifier_group_ids - received_modifier_group_ids

            if new_modifier_group_ids:
                for new_group_id in new_modifier_group_ids:
                    modifier_group_instance = ProductModifierGroup.objects.filter(pk=new_group_id, vendorId=vendor_id).first()

                    ProductModifierAndModifierGroupJoint.objects.create(modifier=updated_modifier, modifierGroup=modifier_group_instance, vendor=vendor_instance)
            
            if deleted_modifier_group_ids:
                for deleted_group_id in deleted_modifier_group_ids:
                    ProductModifierAndModifierGroupJoint.objects.filter(
                        modifier=updated_modifier.pk,
                        modifierGroup=deleted_group_id,
                        vendor=vendor_instance
                    ).delete()
            
            inventory_platform = Platform.objects.filter(Name="Inventory", isActive=True, VendorId=vendor_id).first()
                
            if inventory_platform:
                sync_status = single_modifier_sync_with_odoo(updated_modifier)
                    
                if sync_status == 0:
                    notify(type=3, msg='0', desc='Modifier did not synced with Inventory', stn=['POS'], vendorId=vendor_id)
                
                else:
                    notify(type=3, msg='0', desc='Modifier synced with Inventory', stn=['POS'], vendorId=vendor_id)
            
            modifier_info = get_modifier_data(updated_modifier, vendor_id)
            
            return JsonResponse(modifier_info, status=status.HTTP_200_OK)
        
    else:
        return Response(modifier_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        

@api_view(["DELETE"])
def delete_modifier(request, modifier_id):
    try:
        vendor_id = int(request.GET.get('vendorId'))
        modifier_id = int(modifier_id)

        if vendor_id <= 0 or modifier_id <= 0:
            return Response("Invalid Vendor ID or Modifier ID", status = status.HTTP_400_BAD_REQUEST)

    except:
        return Response("Invalid Vendor ID or Modifier ID", status = status.HTTP_400_BAD_REQUEST)

    if not Vendor.objects.filter(pk = vendor_id).exists():
        return Response("Vendor does not exist", status = status.HTTP_400_BAD_REQUEST)
    
    modifier = ProductModifier.objects.filter(pk = modifier_id, vendorId = vendor_id).first()

    if not modifier:
        return Response("Modifier not found", status = status.HTTP_400_BAD_REQUEST)
    
    with transaction.atomic():
        try:
            inventory_platform_url = Platform.objects.filter(
                Name ="Inventory", isActive =True, VendorId = vendor_id
            ).values_list("baseUrl", flat = True).first()
            
            if inventory_platform_url:
                sync_status = delete_modifier_in_odoo(inventory_platform_url, modifier.modifierPLU, vendor_id)
                    
                if sync_status == 0:
                    notify(type = 3, msg = '0', desc = 'Modifier did not synced with Inventory', stn = ['POS'], vendorId = vendor_id)
                
                else:
                    notify(type = 3, msg = '0', desc = 'Modifier synced with Inventory', stn = ['POS'], vendorId = vendor_id)
            
            modifier.delete()
        
            return Response(status = status.HTTP_204_NO_CONTENT)

        except Exception as e:
            transaction.set_rollback(True)
            
            return Response({'error': str(e)}, status = status.HTTP_500_INTERNAL_SERVER_ERROR)
    

@api_view(["GET",])
def get_customers(request):
    search_parameter = request.GET.get("search_parameter", None)

    try:
        vendor_id = int(request.GET.get("vendorId"))
        page_number = int(request.GET.get("page", 1))
        page_size = int(request.GET.get("page_size", 10))

        if vendor_id <= 0 or page_number <= 0 or page_size <= 0:
            return Response("Invalid request data", status = status.HTTP_400_BAD_REQUEST)

    except:
        return Response("Invalid request data", status = status.HTTP_400_BAD_REQUEST)

    if not Vendor.objects.filter(pk=vendor_id).exists():
        return Response("Vendor does not exist", status = status.HTTP_400_BAD_REQUEST)

    customers = Customer.objects.filter(VendorId = vendor_id).order_by('-pk')
    
    if search_parameter:
        customers = customers.filter(
                Q(Phone_Number__icontains = search_parameter) | \
                Q(FirstName__icontains = search_parameter) | \
                Q(LastName__icontains = search_parameter)
        )

    if not customers.exists():
        return JsonResponse({
            "total_pages": 0,
            "current_page": 0,
            "page_size": 0,
            "results": [],
        }, status = status.HTTP_200_OK)

    paginator = Paginator(customers, page_size)

    paginated_customers = paginator.page(page_number)

    current_page = paginated_customers.number
    
    total_pages = paginator.num_pages

    response_data = []

    for customer in paginated_customers:
        shipping_address = {}

        customer_address = Address.objects.filter(customer = customer.pk, type = "shipping_address", is_selected = True).first()
        
        if customer_address:
            shipping_address = {
                "id": customer_address.pk,
                "address_line_1": customer_address.address_line1 or "",
                "address_line_2": customer_address.address_line2 or "",
                "city": customer_address.city or "",
                "state": customer_address.state or "",
                "country": customer_address.country or "",
                "zipcode": customer_address.zipcode or "",
                "type": customer_address.type,
                "is_selected": customer_address.is_selected
            }

        else:
            shipping_address = {
                "id": 0,
                "address_line_1": "",
                "address_line_2": "",
                "city": "",
                "state": "",
                "country": "",
                "zipcode": "",
                "type": "",
                "is_selected": ""
            }
        
        total_revenue = 0.0

        loyalty_points_history = []

        orders = Order.objects.filter(customerId = customer.pk, vendorId = vendor_id)

        if orders:
            total_revenue = orders.aggregate(total_revenue = Sum('subtotal'))['total_revenue']

            for order in orders:
                loyalty_points_credit_history = LoyaltyPointsCreditHistory.objects.filter(
                    customer = customer.pk, order = order.pk, vendor = vendor_id
                ).order_by("-credit_datetime")

                if loyalty_points_credit_history:
                    for credit_point in loyalty_points_credit_history:
                        credit_transactions = {}

                        loyalty_points_redeem_history = LoyaltyPointsRedeemHistory.objects.filter(
                            customer = customer.pk, credit_history = credit_point.pk, vendor = vendor_id
                        )

                        if loyalty_points_redeem_history:
                            credit_transactions["id"] = credit_point.pk
                            credit_transactions["order_id"] = credit_point.order.externalOrderId
                            credit_transactions["points_credited"] = credit_point.points_credited
                            credit_transactions["credit_datetime"] = credit_point.credit_datetime
                            credit_transactions["expiry_date"] = credit_point.expiry_date
                            credit_transactions["is_expired"] = credit_point.is_expired
                            credit_transactions["total_points_redeemed"] = credit_point.total_points_redeemed
                            credit_transactions["balance_points"] = credit_point.balance_points

                            loyalty_points_history.append(credit_transactions)

        response_data.append({
            "id": customer.pk,
            "FirstName": customer.FirstName or "",
            "LastName": customer.LastName or "",
            "Email": customer.Email or "",
            "Phone_Number": customer.Phone_Number,
            "loyalty_points_balance": customer.loyalty_points_balance,
            "Shipping_Address": shipping_address,
            "total_revenue": total_revenue,
            "loyalty_points_history": loyalty_points_history
        })

    return JsonResponse({
        "total_pages": total_pages,
        "current_page": current_page,
        "page_size": page_size,
        "results": response_data,
    }, status = status.HTTP_200_OK)


@api_view(["POST"])
def create_customer(request):
    body_data = json.loads(request.body)

    vendor_id = body_data.get('vendorId', None)
    phone_number = body_data.get("Phone_Number", None)
    first_name = body_data.get("FirstName", None)
    last_name = body_data.get("LastName", None)
    email = body_data.get("Email", None)
    shipping_address = body_data.get("Shipping_Address", None)

    if not all((vendor_id, phone_number)):
        return Response("Vendor ID or Phone number is empty", status=status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id= int(vendor_id)
    except ValueError:
        return Response("Invalid Vendor ID", status=status.HTTP_400_BAD_REQUEST)

    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor_instance:
        return Response("Vendor does not exist", status=status.HTTP_400_BAD_REQUEST)
    
    existing_customer = Customer.objects.filter(Phone_Number=phone_number, VendorId=vendor_id).first()

    if existing_customer:
        return Response("Customer with this number already exist", status=status.HTTP_400_BAD_REQUEST)

    try:
        with transaction.atomic():
            customer_instance = Customer.objects.create(
                FirstName=first_name,
                LastName=last_name,
                Email=email,
                Phone_Number=phone_number,
                loyalty_points_balance=0,
                VendorId=vendor_instance
            )

            if shipping_address:
                shipping_address_address_line1 = shipping_address.get("address_line_1", None)
                shipping_address_address_line2 = shipping_address.get("address_line_2", None)
                shipping_address_city = shipping_address.get("city", None)
                shipping_address_state = shipping_address.get("state", None)
                shipping_address_country = shipping_address.get("country", None)
                shipping_address_zipcode = shipping_address.get("zipcode", None)

                shipping_address_instance = Address.objects.create(
                    address_line1=shipping_address_address_line1,
                    address_line2=shipping_address_address_line2,
                    city=shipping_address_city,
                    state=shipping_address_state,
                    country=shipping_address_country,
                    zipcode=shipping_address_zipcode,
                    type="shipping_address",
                    is_selected=True,
                    customer=customer_instance
                )
                
            else:
                shipping_address_instance = None

            customer_data = {}
            shipping_address_data = {}

            if shipping_address_instance:
                shipping_address_data["id"] = shipping_address_instance.pk
                shipping_address_data["address_line_1"] = shipping_address_instance.address_line1 if shipping_address_instance.address_line1 else ""
                shipping_address_data["address_line_2"] = shipping_address_instance.address_line2 if shipping_address_instance.address_line2 else ""
                shipping_address_data["city"] = shipping_address_instance.city if shipping_address_instance.city else ""
                shipping_address_data["state"] = shipping_address_instance.state if shipping_address_instance.state else ""
                shipping_address_data["country"] = shipping_address_instance.country if shipping_address_instance.country else ""
                shipping_address_data["zipcode"] = shipping_address_instance.zipcode if shipping_address_instance.zipcode else ""
                shipping_address_data["type"] = shipping_address_instance.type
                shipping_address_data["is_selected"] = shipping_address_instance.is_selected
            
            else:
                shipping_address_data["id"] = 0
                shipping_address_data["address_line_1"] = ""
                shipping_address_data["address_line_2"] = ""
                shipping_address_data["city"] = ""
                shipping_address_data["state"] = ""
                shipping_address_data["country"] = ""
                shipping_address_data["zipcode"] = ""
                shipping_address_data["type"] = ""
                shipping_address_data["is_selected"] = False
            
            customer_data["id"] = customer_instance.pk
            customer_data["FirstName"] = customer_instance.FirstName if customer_instance.FirstName else ""
            customer_data["LastName"] = customer_instance.LastName if customer_instance.LastName else ""
            customer_data["Email"] = customer_instance.Email if customer_instance.Email else ""
            customer_data["Phone_Number"] = customer_instance.Phone_Number if customer_instance.Phone_Number else ""
            customer_data["loyalty_points_balance"] = customer_instance.loyalty_points_balance
            customer_data["Shipping_Address"] = shipping_address_data
            customer_data["total_revenue"] = 0.0
            customer_data["loyalty_points_history"] = []
           
            return JsonResponse(customer_data, status=status.HTTP_200_OK)
        
    except Exception as e:
        print(e)
        return Response(e, status=status.HTTP_400_BAD_REQUEST)


@api_view(["PUT", "PATCH"])
def update_customer(request):
    body_data = json.loads(request.body)

    vendor_id = body_data.get('vendorId', None)
    phone_number = body_data.get("Phone_Number", None)
    customer_id = body_data.get("id")

    if not all((vendor_id, customer_id, phone_number)):
        return Response("Vendor ID, Customer ID or Phone number is empty", status=status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id, customer_id = map(int, (vendor_id, customer_id))
    except ValueError:
        return Response("Invalid Vendor ID or Customer ID", status=status.HTTP_400_BAD_REQUEST)

    vendor = Vendor.objects.filter(pk=vendor_id).first()
    customer_instance = Customer.objects.filter(pk=customer_id).first()

    if not all((vendor, customer_instance)):
        return Response("Vendor or Customer does not exist", status=status.HTTP_400_BAD_REQUEST)
    
    first_name = body_data.get("FirstName", None)
    last_name = body_data.get("LastName", None)
    email = body_data.get("Email", None)
    shipping_address = body_data.get("Shipping_Address", None)

    try:
        with transaction.atomic():
            if shipping_address:
                shipping_address_id = shipping_address.get("id", None)
                shipping_address_address_line1 = shipping_address.get("address_line_1", None)
                shipping_address_address_line2 = shipping_address.get("address_line_2", None)
                shipping_address_city = shipping_address.get("city", None)
                shipping_address_state = shipping_address.get("state", None)
                shipping_address_country = shipping_address.get("country", None)
                shipping_address_zipcode = shipping_address.get("zipcode", None)

                if not shipping_address_id:
                    shipping_address_instance = Address.objects.create(
                        address_line1=shipping_address_address_line1,
                        address_line2=shipping_address_address_line2,
                        city=shipping_address_city,
                        state=shipping_address_state,
                        country=shipping_address_country,
                        zipcode=shipping_address_zipcode,
                        type="shipping_address",
                        is_selected=True,
                        customer=customer_instance
                    )
                
                else:
                    shipping_address_instance = Address.objects.filter(customer=customer_instance.pk, type="shipping_address", is_selected=True).first()

                    if shipping_address_instance:
                        shipping_address_instance.address_line1 = shipping_address_address_line1
                        shipping_address_instance.address_line2 = shipping_address_address_line2
                        shipping_address_instance.city = shipping_address_city
                        shipping_address_instance.state = shipping_address_state
                        shipping_address_instance.country = shipping_address_country
                        shipping_address_instance.zipcode = shipping_address_zipcode

                        shipping_address_instance.save()

                    else:
                        shipping_address_instance = Address.objects.create(
                            address_line1=shipping_address_address_line1,
                            address_line2=shipping_address_address_line2,
                            city=shipping_address_city,
                            state=shipping_address_state,
                            country=shipping_address_country,
                            zipcode=shipping_address_zipcode,
                            type="shipping_address",
                            is_selected=True,
                            customer=customer_instance
                        )
            
            else:
                shipping_address_instance = None
                
            customer_instance.FirstName = first_name
            customer_instance.LastName = last_name
            customer_instance.Email = email
            customer_instance.Phone_Number = phone_number

            customer_instance.save()

            customer_data = {}
            shipping_address_data = {}

            if shipping_address_instance:
                shipping_address_data["id"] = shipping_address_instance.pk
                shipping_address_data["address_line_1"] = shipping_address_instance.address_line1 if shipping_address_instance.address_line1 else ""
                shipping_address_data["address_line_2"] = shipping_address_instance.address_line2 if shipping_address_instance.address_line2 else ""
                shipping_address_data["city"] = shipping_address_instance.city if shipping_address_instance.city else ""
                shipping_address_data["state"] = shipping_address_instance.state if shipping_address_instance.state else ""
                shipping_address_data["country"] = shipping_address_instance.country if shipping_address_instance.country else ""
                shipping_address_data["zipcode"] = shipping_address_instance.zipcode if shipping_address_instance.zipcode else ""
                shipping_address_data["type"] = shipping_address_instance.type
                shipping_address_data["is_selected"] = shipping_address_instance.is_selected
            
            else:
                shipping_address_data["id"] = 0
                shipping_address_data["address_line_1"] = ""
                shipping_address_data["address_line_2"] = ""
                shipping_address_data["city"] = ""
                shipping_address_data["state"] = ""
                shipping_address_data["country"] = ""
                shipping_address_data["zipcode"] = ""
                shipping_address_data["type"] = ""
                shipping_address_data["is_selected"] = False

            total_revenue = 0.0

            loyalty_points_history = []

            orders = Order.objects.filter(customerId=customer_instance.pk, vendorId=vendor_id)

            if orders:
                total_revenue = orders.aggregate(total_revenue=Sum('subtotal'))['total_revenue']
            
                for order in orders:
                        loyalty_points_credit_history = LoyaltyPointsCreditHistory.objects.filter(customer=customer_instance.pk, order=order.pk, vendor=vendor_id).order_by("-credit_datetime")

                        if loyalty_points_credit_history:
                            for credit_point in loyalty_points_credit_history:
                                credit_transactions = {}
                                # redeem_transactions = []

                                loyalty_points_redeem_history = LoyaltyPointsRedeemHistory.objects.filter(customer=customer_instance.pk, credit_history=credit_point.pk, vendor=vendor_id)

                                if loyalty_points_redeem_history:
                                    # for redeem_point in loyalty_points_redeem_history:
                                    #     redeem_history = {}

                                    #     redeem_history["id"] = redeem_point.pk
                                    #     redeem_history["order_id"] = redeem_point.order.externalOrderId
                                    #     redeem_history["points_redeemed"] = redeem_point.points_redeemed
                                    #     redeem_history["redeem_datetime"] = redeem_point.redeem_datetime
                                    #     redeem_history["redeemed_by"] = redeem_point.redeemed_by

                                    #     redeem_transactions.append(redeem_history)

                                    credit_transactions["id"] = credit_point.pk
                                    credit_transactions["order_id"] = credit_point.order.externalOrderId
                                    credit_transactions["points_credited"] = credit_point.points_credited
                                    credit_transactions["credit_datetime"] = credit_point.credit_datetime
                                    credit_transactions["expiry_date"] = credit_point.expiry_date
                                    credit_transactions["is_expired"] = credit_point.is_expired
                                    credit_transactions["total_points_redeemed"] = credit_point.total_points_redeemed
                                    credit_transactions["balance_points"] = credit_point.balance_points
                                    # credit_transactions["redeem_history"] = redeem_transactions

                                    loyalty_points_history.append(credit_transactions)
            
            customer_data["id"] = customer_instance.pk
            customer_data["FirstName"] = customer_instance.FirstName if customer_instance.FirstName else ""
            customer_data["LastName"] = customer_instance.LastName if customer_instance.LastName else ""
            customer_data["Email"] = customer_instance.Email if customer_instance.Email else ""
            customer_data["Phone_Number"] = customer_instance.Phone_Number if customer_instance.Phone_Number else ""
            customer_data["loyalty_points_balance"] = customer_instance.loyalty_points_balance
            customer_data["Shipping_Address"] = shipping_address_data
            customer_data["total_revenue"] = total_revenue
            customer_data["loyalty_points_history"] = loyalty_points_history
           
            return JsonResponse(customer_data, status=status.HTTP_200_OK)
        
    except IntegrityError as e:
        error_message = str(e)

        if 'duplicate key value violates unique constraint' in error_message and 'Phone_Number' in error_message and 'VendorId_id' in error_message:
            return Response("Customer with this phone number already exists", status=status.HTTP_400_BAD_REQUEST)
        else:
            raise e
    
    except Exception as e:
        print(e)
        return Response(e, status=status.HTTP_400_BAD_REQUEST)
    

@api_view(["DELETE"])
def delete_customer(request, customer_id):
    if not customer_id:
        return Response("Modifier ID empty", status=status.HTTP_400_BAD_REQUEST)
    
    customer = Customer.objects.filter(pk=customer_id).first()

    if customer:
        try:
            with transaction.atomic():
                customer.delete()
            
                return Response(status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    else:
        return Response("Customer not found", status=status.HTTP_404_NOT_FOUND)


@api_view(["GET"])
def get_orders_of_customer(request):
    vendor_id = request.GET.get("vendorId", None)
    customer_id = request.GET.get("customerId", None)
    page_number = request.GET.get("page", 1)
    page_size = request.GET.get("page_size", 10)
    language = request.GET.get("language", "English")

    if not all((vendor_id, customer_id)):
        return Response("Vendor ID or Customer ID is empty", status=status.HTTP_400_BAD_REQUEST)

    try:
        vendor_id, customer_id = map(int, (vendor_id, customer_id))

    except ValueError:
        return Response("Invalid Vendor ID or Customer ID", status=status.HTTP_400_BAD_REQUEST)

    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()
    customer_instance = Customer.objects.filter(pk=customer_id).first()

    if not all((vendor_instance, customer_instance)):
        return Response("Vendor or Customer does not exist", status=status.HTTP_404_NOT_FOUND)
    
    loyalty_settings = LoyaltyProgramSettings.objects.get(vendor=vendor_instance.pk)
    
    orders = Order.objects.filter(customerId=customer_id, vendorId=vendor_id).order_by("-arrivalTime")

    if orders:
        paginator = Paginator(orders, page_size)

        try:
            paginated_orders = paginator.page(page_number)

        except PageNotAnInteger:
            paginated_orders = paginator.page(1)
            
        except EmptyPage:
            paginated_orders = paginator.page(paginator.num_pages)

        order_list = []

        current_page = paginated_orders.number
        total_pages = paginator.num_pages
        
        for order in paginated_orders:
            koms_order = KOMSOrder.objects.filter(master_order=order.pk)                       # This is a temparory    
            if koms_order:                                                                     # Fix for Broken orders between 
                koms_order = koms_order.last()                                                  # core and KOMS
                koms_order = KOMSOrder.objects.filter(master_order=order.pk).last()

                order_contents = Order_content.objects.filter(orderId=koms_order.pk)

                order_items = []
                for content in order_contents:
                    product = ProductCategoryJoint.objects.get(product__vendorId=vendor_id, product__PLU=content.SKU)

                    modifier_list = []
                        
                    modifiers = Order_modifer.objects.filter(contentID=content.pk)
                        
                    if modifiers:
                        for modifier in modifiers:
                            product_modifier = ProductModifier.objects.filter(modifierPLU=modifier.SKU, vendorId=vendor_id).first()

                            if language == "English":
                                modifier_name = product_modifier.modifierName

                            else:
                                modifier_name = product_modifier.modifierName_locale
                            
                            modifier_list.append({
                                'modifier_name': modifier_name,
                                'modifier_plu': product_modifier.modifierPLU,
                                'modifier_quantity': modifier.quantity,
                                'modifier_price': product_modifier.modifierPrice if product_modifier else 0.0,
                                'modifier_img': product_modifier.modifierImg if product_modifier else ""
                            })

                    product_image = ProductImage.objects.filter(product=product.product.pk).first()

                    if product_image:
                        image_url = product_image.url

                    else:
                        image_url = 'https://www.stockvault.net/data/2018/08/31/254135/preview16.jpg'
                    
                    if language == "English":
                        product_name = product.product.productName
                        category_name = product.category.categoryName
                    
                    else:
                        product_name = product.product.productName_locale
                        category_name = product.category.categoryName_locale
                    
                    order_items.append({
                        'quantity': content.quantity,
                        'product_plu': content.SKU,
                        'product_name': product_name,
                        'tag': product.product.tag,
                        'is_unlimited': product.product.is_unlimited,
                        'preparation_time': product.product.preparationTime,
                        'is_taxable': product.product.taxable,
                        'product_image': image_url,
                        'category': category_name,
                        "product_note": content.note if content.note else "",
                        "unit_price": product.product.productPrice,
                        'modifiers': modifier_list
                    })
                
                payment_data = {}
                
                payment_details = OrderPayment.objects.filter(orderId=order.pk).last()
                
                if payment_details:
                    payment_type = payment_details.type

                    if payment_type == payment_type_number["Cash"]:
                        payment_data['paymentKey'] = ''
                        payment_data['platform'] = ''
                        payment_data["mode"] = "Cash"

                        if language != "English":
                            payment_data["mode"] = language_localization["Cash"]

                    else:
                        payment_data["paymentKey"] = payment_details.paymentKey if payment_details.paymentKey else ''
                        payment_data["platform"] = payment_details.platform if payment_details.platform else ''
                        payment_data["mode"] = payment_type_english[payment_type]

                        if language != "English":
                            payment_data["mode"] = language_localization[payment_type_english[payment_type]]
                    
                    payment_data["status"] = payment_details.status

                else:
                    payment_data = {
                        "paymentKey": "",
                        "platform": "",
                        "status": False,
                        "mode": "Cash" if language != "English" else language_localization["Cash"]
                    }

                table_numbers_list = ""

                table_details = Order_tables.objects.filter(orderId = koms_order.pk)

                if table_details:
                    for table in table_details:
                        table_numbers_list = table_numbers_list + str(table.tableId.tableNumber) + ","

                    table_numbers_list = table_numbers_list[:-1]

                loyalty_points_redeem_history = LoyaltyPointsRedeemHistory.objects.filter(
                    order = order.pk,
                    customer = customer_id,
                    vendor = vendor_id
                )

                if loyalty_points_redeem_history:
                    total_points_redeemed = loyalty_points_redeem_history.aggregate(Sum('points_redeemed'))['points_redeemed__sum']

                    if not total_points_redeemed:
                        total_points_redeemed = 0

                else:
                    total_points_redeemed = 0
                
                platform_name = order.platform.Name

                if language != "English":
                    platform_name = order.platform.Name_locale

                credit_points = 0
                
                loyalty_points_credit_history = LoyaltyPointsCreditHistory.objects.filter(
                    order = order.pk,
                    customer = customer_id,
                    vendor = vendor_id
                ).first()
                
                if loyalty_settings.is_active == True:
                    if (koms_order.order_status == 10) and (payment_details.status == True):
                        if loyalty_points_credit_history:
                            credit_points = loyalty_points_credit_history.points_credited

                    else:
                        if loyalty_settings.redeem_limit_applied_on == "subtotal":
                            credit_points = round(order.subtotal / loyalty_settings.amount_spent_in_rupees_to_earn_unit_point)

                        elif loyalty_settings.redeem_limit_applied_on == "final_total":
                            credit_points = round(order.TotalAmount / loyalty_settings.amount_spent_in_rupees_to_earn_unit_point)

                else:
                    if loyalty_points_credit_history:
                        credit_points = loyalty_points_credit_history.points_credited
                
                order_data = {
                    "orderId": order.pk,
                    "staging_order_id": koms_order.pk,
                    "external_order_id": order.externalOrderId,
                    "status": koms_order.order_status,
                    "order_note": order.Notes if order.Notes else "",
                    "total_tax": order.tax,
                    "total_discount": order.discount,
                    "delivery_charge": order.delivery_charge,
                    "subtotal": order.subtotal,
                    "total_amount": order.TotalAmount,
                    "pickup_time": koms_order.pickupTime.astimezone(local_timezone).strftime("%Y-%m-%dT%H:%M:%S"),
                    "order_datetime": order.OrderDate.astimezone(local_timezone).strftime("%Y-%m-%dT%H:%M:%S"),
                    "arrival_time": order.arrivalTime.astimezone(local_timezone).strftime("%Y-%m-%dT%H:%M:%S"),
                    "order_type": order.orderType,
                    "platform_name": platform_name,
                    "table_numbers": table_numbers_list,
                    "items": order_items,
                    "payment": payment_data,
                    "total_points_redeemed": total_points_redeemed,
                    "points_earned": credit_points # Key required for NextJS Webiste to show points immediately after placing order
                }

                order_list.append(order_data)
            
        response = {
            "total_pages": total_pages,
            "current_page": current_page,
            "page_size": int(page_size),
            "results": order_list
        }

        return JsonResponse(response, status=status.HTTP_200_OK)
    
    else:
        response = {
            "total_pages": 0,
            "current_page": 0,
            "page_size": 0,
            "results": [],
        }

        return JsonResponse(response, status=status.HTTP_200_OK)
    

@api_view(["GET"])
def get_loyalty_point_transactions_of_customer(request):
    vendor_id = request.GET.get("vendorId", None)
    customer_id = request.GET.get("customerId", None)
    page_number = request.GET.get("page", 1)
    page_size = request.GET.get("page_size", 10)

    if not all((vendor_id, customer_id)):
        return Response("Vendor ID or Customer ID is empty", status=status.HTTP_400_BAD_REQUEST)

    try:
        vendor_id, customer_id = map(int, (vendor_id, customer_id))
    except ValueError:
        return Response("Invalid Vendor ID or Customer ID", status=status.HTTP_400_BAD_REQUEST)

    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()
    customer_instance = Customer.objects.filter(pk=customer_id).first()

    if not all((vendor_instance, customer_instance)):
        return Response("Vendor or Customer does not exist", status=status.HTTP_404_NOT_FOUND)
    
    orders = Order.objects.filter(customerId=customer_id, vendorId=vendor_id).order_by("-OrderDate")

    if orders:
        paginator = Paginator(orders, page_size)

        try:
            paginated_orders = paginator.page(page_number)
        except PageNotAnInteger:
            paginated_orders = paginator.page(1)
        except EmptyPage:
            paginated_orders = paginator.page(paginator.num_pages)

        transaction_list = []

        current_page = paginated_orders.number
        total_pages = paginator.num_pages

        loyalty_points_credit_history = LoyaltyPointsCreditHistory.objects.filter(customer=customer_id)

        sum_of_points_credited = loyalty_points_credit_history.aggregate(Sum('points_credited'))['points_credited__sum']
        sum_of_balance_points = loyalty_points_credit_history.aggregate(Sum('balance_points'))['balance_points__sum']
        sum_of_points_redeemed = loyalty_points_credit_history.aggregate(Sum('total_points_redeemed'))['total_points_redeemed__sum']

        if not sum_of_points_credited:
            sum_of_points_credited = 0

        if not sum_of_balance_points:
            sum_of_balance_points = 0

        if not sum_of_points_redeemed:
            sum_of_points_redeemed = 0
        
        loyalty_points_redeem_history = LoyaltyPointsRedeemHistory.objects.filter(customer=customer_id)
        
        for order in paginated_orders:
            credit_transaction = {}
            redeem_transaction = []

            credit_history = loyalty_points_credit_history.filter(order=order.pk).first()

            if credit_history:
                credit_transaction["points_credited"] = credit_history.points_credited
                credit_transaction["credit_datetime"] = credit_history.credit_datetime
                credit_transaction["expiry_date"] = credit_history.expiry_date
                credit_transaction["is_expired"] = credit_history.is_expired
            
            else:
                credit_transaction["points_credited"] = 0
                credit_transaction["credit_datetime"] = ""
                credit_transaction["expiry_date"] = ""
                credit_transaction["is_expired"] = False

            redeem_history = loyalty_points_redeem_history.filter(order=order.pk)

            if redeem_history:
                redeemed_points = redeem_history.aggregate(Sum('points_redeemed'))['points_redeemed__sum']

                for history in redeem_history:
                    redeem_transaction.append({
                        "credit_history_id": history.credit_history.pk,
                        "points_redeemed": history.points_redeemed,
                        "redeem_datetime": history.redeem_datetime,
                        "redeemed_by": history.redeemed_by
                    })    
            
            else:
                redeemed_points = 0
            
            order_data = {
                "order_id": order.pk,
                "external_order_id": order.externalOrderId,
                "order_datetime": order.OrderDate.astimezone(local_timezone).strftime("%Y-%m-%dT%H:%M:%S"),
                "credit_history": credit_transaction,
                "redeemed_points": redeemed_points,
                "redeem_history": redeem_transaction
            }

            transaction_list.append(order_data)
        
        response = {
            "total_pages": total_pages,
            "current_page": current_page,
            "page_size": int(page_size),
            "current_points_balance": customer_instance.loyalty_points_balance,
            "sum_of_points_credited": sum_of_points_credited,
            "sum_of_balance_points": sum_of_balance_points,
            "sum_of_points_redeemed": sum_of_points_redeemed,
            "results": transaction_list
        }

        return JsonResponse(response, status=status.HTTP_200_OK)
    
    else:
        response = {
            "total_pages": 0,
            "current_page": 0,
            "page_size": 0,
            "current_points_balance": 0,
            "sum_of_points_credited": 0,
            "sum_of_balance_points": 0,
            "sum_of_points_redeemed": 0,
            "results": [],
        }

        return JsonResponse(response, status=status.HTTP_200_OK)


@api_view(['GET'])
def get_loyalty_points_settings(request):
    vendor_id = request.GET.get("vendorId", None)

    if vendor_id is None:
        return Response("Vendor ID empty", status=status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id = int(vendor_id)
    except ValueError:
        return Response("Invalid Vendor ID", status=status.HTTP_400_BAD_REQUEST)

    vendor = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor:
        return Response("Vendor does not exist", status=status.HTTP_404_NOT_FOUND)
    
    settings_data = {}
    
    settings = LoyaltyProgramSettings.objects.filter(vendor=vendor_id).first()

    if settings:
        settings_data["id"] = settings.pk
        settings_data["is_active"] = settings.is_active
        settings_data["amount_spent"] = settings.amount_spent_in_rupees_to_earn_unit_point
        settings_data["unit_point_value"] = settings.unit_point_value_in_rupees
        settings_data["expiry_days"] = settings.points_expiry_days
        settings_data["applied_on"] = settings.redeem_limit_applied_on
        settings_data["redeem_limit"] = settings.redeem_limit_in_percentage

        return JsonResponse(settings_data, status=status.HTTP_200_OK)
    
    else:
        settings_data["id"] = 0
        settings_data["is_active"] = False
        settings_data["amount_spent"] = 0
        settings_data["unit_point_value"] = 0
        settings_data["expiry_days"] = 0
        settings_data["applied_on"] = ""
        settings_data["redeem_limit"] = 0

        return JsonResponse(settings_data, status=status.HTTP_200_OK)


@api_view(["POST"])
def create_loyalty_points_settings(request):
    body_data = json.loads(request.body)

    if body_data:
        vendor_id = body_data.get('vendorId', None)

        if vendor_id is None:
            return Response("Vendor ID empty", status=status.HTTP_400_BAD_REQUEST)
        
        try:
            vendor_id = int(vendor_id)
        except ValueError:
            return Response("Invalid Vendor ID", status=status.HTTP_400_BAD_REQUEST)

        vendor = Vendor.objects.filter(pk=vendor_id).first()

        if not vendor:
            return Response("Vendor does not exist", status=status.HTTP_400_BAD_REQUEST)
        
        is_active = body_data.get('is_active')
        amount_spent = body_data.get('amount_spent')
        unit_point_value = body_data.get('unit_point_value')
        expiry_days = body_data.get('expiry_days')
        applied_on = body_data.get('applied_on')
        redeem_limit = body_data.get('redeem_limit')

        if (is_active == None) or (amount_spent == None) or (unit_point_value == None) or \
        (expiry_days == None) or (applied_on == None) or (redeem_limit == None):
            return Response("Invalid data sent", status=status.HTTP_400_BAD_REQUEST)

        settings_count = LoyaltyProgramSettings.objects.filter(vendor=vendor_id).count()

        if settings_count > 0:
            return Response("Loyalty program settings already created", status=status.HTTP_200_OK)
        
        with transaction.atomic():
            settings = LoyaltyProgramSettings.objects.create(
                is_active = is_active,
                amount_spent_in_rupees_to_earn_unit_point = amount_spent,
                unit_point_value_in_rupees = unit_point_value,
                points_expiry_days = expiry_days,
                redeem_limit_applied_on = applied_on,
                redeem_limit_in_percentage = redeem_limit,
                vendor = vendor
            )

            settings_data = {}

            settings_data["id"] = settings.pk
            settings_data["is_active"] = settings.is_active
            settings_data["amount_spent"] = settings.amount_spent_in_rupees_to_earn_unit_point
            settings_data["unit_point_value"] = settings.unit_point_value_in_rupees
            settings_data["expiry_days"] = settings.points_expiry_days
            settings_data["applied_on"] = settings.redeem_limit_applied_on
            settings_data["redeem_limit"] = settings.redeem_limit_in_percentage

            return JsonResponse(settings_data, status=status.HTTP_201_CREATED)


@api_view(["PUT", "PATCH"])
def update_loyalty_points_settings(request):
    body_data = json.loads(request.body)

    vendor_id = body_data.get('vendorId', None)

    if vendor_id is None:
        return Response("Vendor ID empty", status=status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id = int(vendor_id)
    except ValueError:
        return Response("Invalid Vendor ID", status=status.HTTP_400_BAD_REQUEST)

    vendor = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor:
        return Response("Vendor does not exist", status=status.HTTP_400_BAD_REQUEST)
    
    setting_id = body_data.get('id')
    is_active = body_data.get('is_active')
    amount_spent = body_data.get('amount_spent')
    unit_point_value = body_data.get('unit_point_value')
    expiry_days = body_data.get('expiry_days')
    applied_on = body_data.get('applied_on')
    redeem_limit = body_data.get('redeem_limit')

    settings = LoyaltyProgramSettings.objects.filter(pk=setting_id, vendor=vendor_id).first()

    if not settings:
        return Response("Settings not found", status=status.HTTP_400_BAD_REQUEST)
    
    with transaction.atomic():    
        settings.is_active = is_active
        settings.amount_spent_in_rupees_to_earn_unit_point = amount_spent
        settings.unit_point_value_in_rupees = unit_point_value
        settings.points_expiry_days = expiry_days
        settings.redeem_limit_applied_on = applied_on
        settings.redeem_limit_in_percentage = redeem_limit

        settings.save()

        settings_data = {}

        settings_data["id"] = settings.pk
        settings_data["is_active"] = settings.is_active
        settings_data["amount_spent"] = settings.amount_spent_in_rupees_to_earn_unit_point
        settings_data["unit_point_value"] = settings.unit_point_value_in_rupees
        settings_data["expiry_days"] = settings.points_expiry_days
        settings_data["applied_on"] = settings.redeem_limit_applied_on
        settings_data["redeem_limit"] = settings.redeem_limit_in_percentage

        return JsonResponse(settings_data, status=status.HTTP_200_OK)


@api_view(["POST"])
def redeem_loyalty_points(request):
    body_data = request.data

    vendor_id = body_data.get('vendor_id')
    customer_id = body_data.get('customer_id')
    external_order_id = body_data.get('external_order_id')
    is_wordpress = body_data.get('is_wordpress')
    language = body_data.get('language')

    if not all((vendor_id, customer_id, external_order_id)):
        return Response("Vendor ID, Customer ID, Order ID or is_wordpress is empty", status=status.HTTP_400_BAD_REQUEST)

    if is_wordpress is None:
        return Response("is_wordpress is empty", status=status.HTTP_400_BAD_REQUEST)

    try:
        vendor_id, customer_id = map(int, (vendor_id, customer_id))
    except ValueError:
        return Response("Invalid Vendor ID, Customer ID", status=status.HTTP_400_BAD_REQUEST)

    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()
    customer_instance = Customer.objects.filter(pk=customer_id, VendorId=vendor_id).first()
    master_order_instance = Order.objects.filter(externalOrderId=external_order_id, vendorId=vendor_id).first()

    if not all((vendor_instance, customer_instance, master_order_instance)):
        return Response("Vendor, Customer, or Order does not exist", status=status.HTTP_400_BAD_REQUEST)

    subtotal = body_data.get('subtotal')
    tax = body_data.get('tax')
    discount = body_data.get('discount')
    delivery_charge = body_data.get('delivery_charge')
    total_amount = body_data.get('total_amount')
    
    with transaction.atomic():    
        master_order_instance.subtotal = subtotal
        master_order_instance.tax = tax
        master_order_instance.discount = discount
        master_order_instance.delivery_charge = delivery_charge
        master_order_instance.TotalAmount = total_amount

        master_order_instance.save()

        print(master_order_instance.customerId.FirstName, master_order_instance.customerId.Phone_Number)

        if (master_order_instance.customerId.FirstName != "Guest") and (master_order_instance.customerId.Phone_Number != "0"):
            loyalty_program_settings = LoyaltyProgramSettings.objects.filter(vendor=vendor_id).first()

            if not loyalty_program_settings:
                return Response("Loyalty Program Settings not found for the vendor", status=status.HTTP_400_BAD_REQUEST)
            
            points_utilized = body_data.get('points_redeemed', 0)
            points_redeemed_by = body_data.get('redeemed_by')

            loyalty_points_credit_history = LoyaltyPointsCreditHistory.objects.filter(
                customer=master_order_instance.customerId.pk,
                is_expired=False,
                balance_points__gt=0,
                vendor=vendor_id
            ).order_by("expiry_date")

            points_total = 0
            points_to_redeem = 0
            transactions = []

            customer_instance = Customer.objects.get(pk=master_order_instance.customerId.pk)

            if points_utilized == 0:
                return Response(status=status.HTTP_200_OK)
            
            elif points_utilized < 0:
                return Response("Points redeemed cannot be negative", status=status.HTTP_400_BAD_REQUEST)

            elif customer_instance.loyalty_points_balance == 0:
                if points_utilized == 0:
                    return Response(status=status.HTTP_200_OK)
            
                elif points_utilized > 0:
                    return Response("0 points in account", status=status.HTTP_400_BAD_REQUEST)
            
            elif (customer_instance.loyalty_points_balance > 0):
                if points_utilized > customer_instance.loyalty_points_balance:
                    return Response("Points balance is low", status=status.HTTP_400_BAD_REQUEST)
                
                elif points_utilized > 0:
                    for credit_history_instance in loyalty_points_credit_history:
                        if (points_total != points_utilized) and (points_to_redeem >= 0):
                            points_to_redeem = points_utilized - points_total

                            if (credit_history_instance.balance_points < points_to_redeem) or \
                            (credit_history_instance.balance_points == points_to_redeem):
                                points_total = points_total + credit_history_instance.balance_points
                                
                                transactions.append({
                                    "credit_history_id": credit_history_instance.pk,
                                    "utilized_points": credit_history_instance.balance_points
                                })

                            elif (credit_history_instance.balance_points > points_to_redeem):
                                points_total = points_total + points_to_redeem
                                
                                transactions.append({
                                    "credit_history_id": credit_history_instance.pk,
                                    "utilized_points": points_to_redeem
                                })

                    for record in transactions:
                        loyalty_points_credit_history_instance = LoyaltyPointsCreditHistory.objects.filter(
                            pk=record["credit_history_id"]
                        ).first()
                        
                        loyalty_points_redeem_history_instance = LoyaltyPointsRedeemHistory(
                            customer = customer_instance,
                            order = master_order_instance,
                            credit_history = loyalty_points_credit_history_instance,
                            points_redeemed = record["utilized_points"],
                            redeem_datetime = timezone.now(), # timezone.localtime(timezone.now(), timezone='Asia/Kolkata')
                            redeemed_by = points_redeemed_by,
                            vendor = vendor_instance
                        )

                        loyalty_points_redeem_history_instance.save()

                        loyalty_points_credit_history_instance.total_points_redeemed = loyalty_points_credit_history_instance.total_points_redeemed + record["utilized_points"]
                        loyalty_points_credit_history_instance.balance_points = loyalty_points_credit_history_instance.balance_points - record["utilized_points"]
                        loyalty_points_credit_history_instance.save()

                        customer_instance.loyalty_points_balance = customer_instance.loyalty_points_balance - record["utilized_points"]
                        customer_instance.save()

                    koms_order = KOMSOrder.objects.filter(master_order=master_order_instance.pk).first()
                    if master_order_instance.masterOrder:
                        koms_order = KOMSOrder.objects.filter(master_order=master_order_instance.masterOrder.pk).first()
                    waiteOrderUpdate(orderid=koms_order.pk, language=language, vendorId=vendor_id) # call socket
                    
                    return Response("Points redeemed", status=status.HTTP_200_OK)

        else:
            return Response("Cannot redeemed point of Guest", status=status.HTTP_400_BAD_REQUEST)
        

def loyalty_points_redeem(vendor_id, customer_id, master_order_id, is_wordpress, points_utilized, points_redeemed_by):
    is_redeemed = False

    if not all((vendor_id, customer_id, master_order_id)):
        return is_redeemed

    if is_wordpress is None:
        return is_redeemed

    try:
        vendor_id, customer_id = map(int, (vendor_id, customer_id))
    except ValueError:
        return is_redeemed

    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()
    customer_instance = Customer.objects.filter(pk=customer_id, VendorId=vendor_id).first()
    master_order_instance = Order.objects.filter(pk=master_order_id, vendorId=vendor_id).first()

    if not all((vendor_instance, customer_instance, master_order_instance)):
        return is_redeemed
    
    with transaction.atomic():
        if (master_order_instance.customerId.FirstName != "Guest") and (master_order_instance.customerId.Phone_Number != "0"):
            loyalty_program_settings = LoyaltyProgramSettings.objects.filter(vendor=vendor_id).first()

            if not loyalty_program_settings:
                return is_redeemed

            loyalty_points_credit_history = LoyaltyPointsCreditHistory.objects.filter(
                customer=master_order_instance.customerId.pk,
                is_expired=False,
                balance_points__gt=0,
                vendor=vendor_id
            ).order_by("expiry_date")

            points_total = 0
            points_to_redeem = 0
            transactions = []

            if customer_instance.loyalty_points_balance > 0:
                for credit_history_instance in loyalty_points_credit_history:
                    if (credit_history_instance.balance_points == points_utilized):
                        if (credit_history_instance.points_credited == points_utilized):
                            LoyaltyPointsRedeemHistory.objects.create(
                                customer = customer_instance,
                                order = master_order_instance,
                                credit_history = credit_history_instance,
                                points_redeemed = points_utilized,
                                redeem_datetime = timezone.now(), # timezone.localtime(timezone.now(), timezone='Asia/Kolkata')
                                redeemed_by = points_redeemed_by,
                                vendor = vendor_instance
                            )

                            credit_history_instance.total_points_redeemed = points_utilized
                            credit_history_instance.balance_points = 0
                            credit_history_instance.save()

                            customer_instance.loyalty_points_balance = customer_instance.loyalty_points_balance - points_utilized
                            customer_instance.save()

                            break
                    
                        else:
                            LoyaltyPointsRedeemHistory.objects.create(
                                customer = customer_instance,
                                order = master_order_instance,
                                credit_history = credit_history_instance,
                                points_redeemed = points_utilized,
                                redeem_datetime = timezone.now(), # timezone.localtime(timezone.now(), timezone='Asia/Kolkata')
                                redeemed_by = points_redeemed_by,
                                vendor = vendor_instance
                            )

                            credit_history_instance.total_points_redeemed = credit_history_instance.total_points_redeemed + points_utilized
                            credit_history_instance.balance_points = 0
                            credit_history_instance.save()

                            customer_instance.loyalty_points_balance = customer_instance.loyalty_points_balance - points_utilized
                            customer_instance.save()
                            
                            break

                    elif (credit_history_instance.balance_points > points_utilized):
                        LoyaltyPointsRedeemHistory.objects.create(
                            customer = customer_instance,
                            order = master_order_instance,
                            credit_history = credit_history_instance,
                            points_redeemed = points_utilized,
                            redeem_datetime = timezone.now(), # timezone.localtime(timezone.now(), timezone='Asia/Kolkata')
                            redeemed_by = points_redeemed_by,
                            vendor = vendor_instance
                        )

                        credit_history_instance.total_points_redeemed = credit_history_instance.total_points_redeemed + points_utilized
                        credit_history_instance.balance_points = credit_history_instance.balance_points - points_utilized
                        credit_history_instance.save()

                        customer_instance.loyalty_points_balance = customer_instance.loyalty_points_balance - points_utilized
                        customer_instance.save()
                        
                        break

                    else:
                        if (points_total != points_utilized) and (points_to_redeem >= 0):
                            points_to_redeem = points_utilized - points_total

                            if (credit_history_instance.balance_points < points_to_redeem) or \
                            (credit_history_instance.balance_points == points_to_redeem):
                                points_total = points_total + credit_history_instance.balance_points
                                
                                transactions.append({
                                    "credit_history_id": credit_history_instance.pk,
                                    "utilized_points": credit_history_instance.balance_points
                                })

                            elif (credit_history_instance.balance_points > points_to_redeem):
                                points_total = points_total + points_to_redeem
                                
                                transactions.append({
                                    "credit_history_id": credit_history_instance.pk,
                                    "utilized_points": points_to_redeem
                                })
                        
                        else:
                            for record in transactions:
                                loyalty_points_credit_history_instance = LoyaltyPointsCreditHistory.objects.filter(
                                    pk=record["credit_history_id"]
                                ).first()

                                loyalty_points_redeem_history_instance = LoyaltyPointsRedeemHistory(
                                    customer = customer_instance,
                                    order = master_order_instance,
                                    credit_history = loyalty_points_credit_history_instance,
                                    points_redeemed = record["utilized_points"],
                                    redeem_datetime = timezone.now(), # timezone.localtime(timezone.now(), timezone='Asia/Kolkata')
                                    redeemed_by = points_redeemed_by,
                                    vendor = vendor_instance
                                )

                                loyalty_points_redeem_history_instance.save()

                                loyalty_points_credit_history_instance.total_points_redeemed = loyalty_points_credit_history_instance.total_points_redeemed + record["utilized_points"]
                                loyalty_points_credit_history_instance.balance_points = loyalty_points_credit_history_instance.balance_points - record["utilized_points"]
                                loyalty_points_credit_history_instance.save()

                                customer_instance.loyalty_points_balance = customer_instance.loyalty_points_balance - record["utilized_points"]
                                customer_instance.save()

                            break

                is_redeemed = True
                return is_redeemed
            
            return is_redeemed


@api_view(["GET"])
def top_selling_products_report(request):
    vendor_id = request.GET.get("vendorId")
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    order_type = request.GET.get('type')
    is_download = request.GET.get('download')
    top_number = request.GET.get('top')
    sort_by = request.GET.get('sort')
    language = request.GET.get('language', 'English')

    if not vendor_id:
        return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id = int(vendor_id)

    except ValueError:
        return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
    
    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor_instance:
        return Response("Vendor does not exist", status=status.HTTP_400_BAD_REQUEST)
    
    if (((not start_date) or (not end_date)) or (start_date > end_date)):
        return Response("Invalid start date or end date", status=status.HTTP_400_BAD_REQUEST)

    if is_download not in ("false", "true"):
        return Response("Invalid download parameter", status=status.HTTP_400_BAD_REQUEST)

    try:
        top_number = int(top_number)

    except ValueError:
        return Response("Invalid top parameter", status=status.HTTP_400_BAD_REQUEST)
    
    if sort_by not in ("ascending", "descending"):
        return Response("Invalid sort parameter", status=status.HTTP_400_BAD_REQUEST)
    
    platform = Platform.objects.filter(Name__in=('Mobile App', 'Website'), isActive=True, VendorId=vendor_id).first()
    
    if order_type == "all":
        order_items = Order_content.objects.filter(
            orderId__order_status = 10,
            orderId__master_order__Status = master_order_status_number["Completed"],
            orderId__master_order__OrderDate__date__range = (start_date, end_date),
            orderId__master_order__vendorId = vendor_id,
            orderId__vendorId = vendor_id,
        ).exclude(status = 5)
    
    elif order_type == "delivery":
        order_items = Order_content.objects.filter(
            orderId__order_status = 10,
            orderId__master_order__orderType = order_type_number["Delivery"],
            orderId__master_order__Status = master_order_status_number["Completed"],
            orderId__master_order__OrderDate__date__range = (start_date, end_date),
            orderId__master_order__vendorId = vendor_id,
            orderId__vendorId = vendor_id,
        ).exclude(status = 5)

    elif order_type == "pickup":
        order_items = Order_content.objects.filter(
            orderId__order_status = 10,
            orderId__master_order__orderType = order_type_number["Pickup"],
            orderId__master_order__Status = master_order_status_number["Completed"],
            orderId__master_order__OrderDate__date__range = (start_date, end_date),
            orderId__master_order__vendorId = vendor_id,
            orderId__vendorId=vendor_id,
        ).exclude(status = 5)

    elif order_type == "dinein":
        order_items = Order_content.objects.filter(
            orderId__order_status = 10,
            orderId__master_order__orderType = order_type_number["Dinein"],
            orderId__master_order__Status = master_order_status_number["Completed"],
            orderId__master_order__OrderDate__date__range = (start_date, end_date),
            orderId__master_order__vendorId = vendor_id,
            orderId__vendorId = vendor_id,
        ).exclude(status = 5)

    elif order_type == "online":
        if platform:
            order_items = Order_content.objects.filter(
                orderId__order_status = 10,
                orderId__master_order__platform = platform.pk,
                orderId__master_order__Status = master_order_status_number["Completed"],
                orderId__master_order__OrderDate__date__range = (start_date, end_date),
                orderId__master_order__vendorId = vendor_id,
                orderId__vendorId = vendor_id,
            ).exclude(status = 5)

        else:
            return Response("Contact you administrator to activate the platform", status=status.HTTP_400_BAD_REQUEST)
        
    elif order_type == "offline":
        if platform:
            order_items = Order_content.objects.filter(
                orderId__order_status = 10,
                orderId__master_order__Status = master_order_status_number["Completed"],
                orderId__master_order__OrderDate__date__range = (start_date, end_date),
                orderId__master_order__vendorId = vendor_id,
                orderId__vendorId = vendor_id,
            ).exclude(status = 5, orderId__master_order__platform = platform.pk)

        else:
            return Response("Contact you administrator to activate the platform", status=status.HTTP_400_BAD_REQUEST)

    else:
        return Response("Invalid type parameter", status=status.HTTP_400_BAD_REQUEST)

    top_selling_items = order_items.values('SKU').distinct() \
        .annotate(quantity_sold=ExpressionWrapper(Sum('quantity'),output_field=IntegerField())) \
        .filter(quantity_sold__gt=0).order_by('-quantity_sold')[:top_number]
    
    if sort_by == "descending":
        pass

    elif sort_by == "ascending":
        top_selling_items = sorted(top_selling_items, key=itemgetter('quantity_sold'))

    else:
        return Response("Invalid sort parameter", status=status.HTTP_400_BAD_REQUEST)
    
    list_of_items = []

    if is_download.lower() == "false":
        if order_items.exists():
            for item in top_selling_items:
                image_list = []

                products = Product.objects.filter(PLU=item['SKU'], vendorId=vendor_id).first()
                
                images = ProductImage.objects.filter(product=products.pk)

                for image in images:
                    if image.url:
                        image_list.append(image.url)

                product_name = ""

                if language == "English":
                    product_name = products.productName

                else:
                    product_name = products.productName_locale
                
                item['product_id'] = products.pk
                item['product_name'] = product_name
                item['image'] = image_list[0] if len(image_list)!=0 else 'https://www.stockvault.net/data/2018/08/31/254135/preview16.jpg'
                item['price'] = products.productPrice
                item['total_sale'] = item['quantity_sold'] * products.productPrice

                list_of_items.append(item)

        return JsonResponse({"top_selling_products": list_of_items})
    
    elif is_download.lower() == "true":
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')

        formatted_start_date = start_date.strftime('%d-%m-%Y')
        formatted_end_date = end_date.strftime('%d-%m-%Y')

        # Create a new Excel workbook and select the active sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        if language == "English":
            sheet.append(['Start Date', f'{formatted_start_date}'])
            sheet.append(['End Date', f'{formatted_end_date}'])
            sheet.append(['Order Type', f'{order_type}'])
            sheet.append(['Top', f'{top_number}'])
            sheet.append(['Sorted by', f'{sort_by}'])
            sheet.append(['Total records', f'{len(top_selling_items)}'])
            sheet.append([''])

            sheet.append(['Product Name', 'Quantity Sold', 'Unit Price', 'Total Sale'])

        else:
            sort_by = language_localization[sort_by]

            order_type = language_localization[order_type]

            sheet.append([language_localization["Start Date"], f'{formatted_start_date}'])
            sheet.append([language_localization["End Date"], f'{formatted_end_date}'])
            sheet.append([language_localization["Order Type"], f'{order_type}'])
            sheet.append([language_localization["Top"], f'{top_number}'])
            sheet.append([language_localization["Sorted by"], f'{sort_by}'])
            sheet.append([language_localization["Total records"], f'{len(top_selling_items)}'])
            sheet.append([''])

            sheet.append([
                language_localization['Product Name'],
                language_localization['Quantity Sold'],
                language_localization['Unit Price'],
                language_localization['Total Sale']
            ])

        if order_items.exists():    
            for item in top_selling_items:
                products = Product.objects.filter(PLU=item['SKU'], vendorId=vendor_id).first()

                product_name = ""

                if language == "English":
                    product_name = products.productName

                else:
                    product_name = products.productName_locale

                sheet.append([
                    product_name,
                    item['quantity_sold'],
                    products.productPrice,
                    item['quantity_sold'] * products.productPrice
                ])
        
        directory = os.path.join(settings.MEDIA_ROOT, 'Excel Downloads')
        os.makedirs(directory, exist_ok=True) # Create the directory if it doesn't exist inside MEDIA_ROOT

        file_name = f"Top_selling_products_{vendor_id}.xlsx"

        relative_file_path = os.path.join('Excel Downloads', file_name)

        file_path = os.path.join(settings.MEDIA_ROOT, relative_file_path)

        workbook.save(file_path)

        print(f"Excel file '{file_name}' has been created.")
        
        response = "/media/" + relative_file_path
        
        return HttpResponse(response, status=200)

    else:
        return Response("Invalid 'download' key", status=status.HTTP_400_BAD_REQUEST)


@api_view(["GET"])
def most_repeating_customers_report(request):
    vendor_id = request.GET.get("vendorId")
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    order_type = request.GET.get('type')
    is_download = request.GET.get('download')
    top_number = request.GET.get('top')
    sort_by = request.GET.get('sort')
    language = request.GET.get('language', 'English')

    if not vendor_id:
        return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id = int(vendor_id)

    except ValueError:
        return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
    
    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor_instance:
        return Response("Vendor does not exist", status=status.HTTP_400_BAD_REQUEST)
    
    if (((not start_date) or (not end_date)) or (start_date > end_date)):
        return Response("Invalid start date or end date", status=status.HTTP_400_BAD_REQUEST)
    
    if not is_download:
        is_download = "false"
    
    else:
        is_download = str(is_download)

    try:
        top_number = int(top_number)
        
    except ValueError:
        return Response("Invalid top parameter", status=status.HTTP_400_BAD_REQUEST)
    
    if not sort_by:
        sort_by = "descending "

    platform = Platform.objects.filter(Name__in=('Mobile App', 'Website'), isActive=True, VendorId=vendor_id).first()
    
    if order_type == "all":
        orders = Order.objects.filter(
            Status = master_order_status_number["Completed"],
            OrderDate__date__range = (start_date, end_date),
            vendorId = vendor_id
        )
    
    elif order_type == "delivery":
        orders = Order.objects.filter(
            orderType = order_type_number["Delivery"],
            Status = master_order_status_number["Completed"],
            OrderDate__date__range = (start_date, end_date),
            vendorId = vendor_id,
        )
    
    elif order_type == "pickup":
        orders = Order.objects.filter(
            orderType = order_type_number["Pickup"],
            Status = master_order_status_number["Completed"],
            OrderDate__date__range = (start_date, end_date),
            vendorId = vendor_id
        )

    elif order_type == "dinein":
        orders = Order.objects.filter(
            orderType = order_type_number["Dinein"],
            Status = master_order_status_number["Completed"],
            OrderDate__date__range = (start_date, end_date),
            vendorId = vendor_id
        )

    elif order_type == "online":
        if platform:
            orders = Order.objects.filter(
                platform = platform.pk,
                Status = master_order_status_number["Completed"],
                OrderDate__date__range = (start_date, end_date),
                vendorId = vendor_id
            )
        
        else:
            return Response("Contact you administrator to activate the platform", status=status.HTTP_400_BAD_REQUEST)
        
    elif order_type == "offline":
        if platform:
            orders = Order.objects.filter(
                Status = master_order_status_number["Completed"],
                OrderDate__date__range = (start_date, end_date),
                vendorId = vendor_id
            ).exclude(platform = platform.pk)
        
        else:
            return Response("Contact you administrator to activate the platform", status=status.HTTP_400_BAD_REQUEST)
    
    else:
        return Response("Invalid type parameter", status=status.HTTP_400_BAD_REQUEST)
    
    top_customers = orders.values('customerId').distinct() \
        .annotate(customer_count=Count('customerId')) \
        .filter(customer_count__gt=0) \
        .order_by('-customer_count')[:top_number]

    if sort_by == "descending":
        pass

    elif sort_by == "ascending":
        top_customers = sorted(top_customers, key=itemgetter('customer_count'))

    else:
        return Response("Invalid sort parameter", status=status.HTTP_400_BAD_REQUEST)

    customer_details = []
    
    if is_download.lower() == "false": 
        for customer in top_customers:
            if order_type == "all":
                order_items = Order_content.objects.filter(
                    orderId__order_status = 10,
                    orderId__master_order__customerId = customer["customerId"],
                    orderId__master_order__Status = master_order_status_number["Completed"],
                    orderId__master_order__OrderDate__date__range = (start_date, end_date),
                    orderId__master_order__vendorId = vendor_id,
                    orderId__vendorId = vendor_id,
                ).exclude(status = 5)
            
            elif order_type == "delivery":
                order_items = Order_content.objects.filter(
                    orderId__order_status = 10,
                    orderId__master_order__customerId = customer["customerId"],
                    orderId__master_order__orderType = order_type_number["Delivery"],
                    orderId__master_order__Status = master_order_status_number["Completed"],
                    orderId__master_order__OrderDate__date__range = (start_date, end_date),
                    orderId__master_order__vendorId = vendor_id,
                    orderId__vendorId = vendor_id,
                ).exclude(status = 5)

            elif order_type == "pickup":
                order_items = Order_content.objects.filter(
                    orderId__order_status = 10,
                    orderId__master_order__customerId = customer["customerId"],
                    orderId__master_order__orderType = order_type_number["Pickup"],
                    orderId__master_order__Status = master_order_status_number["Completed"],
                    orderId__master_order__OrderDate__date__range = (start_date, end_date),
                    orderId__master_order__vendorId = vendor_id,
                    orderId__vendorId = vendor_id,
                ).exclude(status = 5)

            elif order_type == "dinein":
                order_items = Order_content.objects.filter(
                    orderId__order_status = 10,
                    orderId__master_order__customerId = customer["customerId"],
                    orderId__master_order__orderType = order_type_number["Dinein"],
                    orderId__master_order__Status = master_order_status_number["Completed"],
                    orderId__master_order__OrderDate__date__range = (start_date, end_date),
                    orderId__master_order__vendorId = vendor_id,
                    orderId__vendorId = vendor_id,
                ).exclude(status = 5)

            elif order_type == "online":
                order_items = Order_content.objects.filter(
                    orderId__order_status = 10,
                    orderId__master_order__customerId = customer["customerId"],
                    orderId__master_order__platform = platform.pk,
                    orderId__master_order__Status = master_order_status_number["Completed"],
                    orderId__master_order__OrderDate__date__range = (start_date, end_date),
                    orderId__master_order__vendorId = vendor_id,
                    orderId__vendorId = vendor_id,
                ).exclude(status = 5)

            elif order_type == "offline":
                order_items = Order_content.objects.filter(
                    orderId__order_status = 10,
                    orderId__master_order__customerId = customer["customerId"],
                    orderId__master_order__Status = master_order_status_number["Completed"],
                    orderId__master_order__OrderDate__date__range = (start_date, end_date),
                    orderId__master_order__vendorId = vendor_id,
                    orderId__vendorId = vendor_id,
                ).exclude(status = 5, orderId__master_order__platform = platform.pk)

            list_of_items = []
            
            if order_items.exists():
                top_selling_items = order_items.values('SKU').distinct() \
                    .annotate(quantity_sold=ExpressionWrapper(Sum('quantity'), output_field=IntegerField())) \
                    .filter(quantity_sold__gt=0) \
                    .order_by('-quantity_sold')[:10]
                
                for item in top_selling_items:
                    image_list = []

                    products = Product.objects.filter(PLU=item['SKU'], vendorId=vendor_id).first()
                    
                    images = ProductImage.objects.filter(product=products.pk)

                    for image in images:
                        if image.url:
                            image_list.append(image.url)

                    product_name = ""

                    if language == "English":
                        product_name = products.productName

                    else:
                        product_name = products.productName_locale

                    item['product_id'] = products.pk
                    item['product_name'] = product_name
                    item['image'] = image_list[0] if len(image_list)!=0 else 'https://www.stockvault.net/data/2018/08/31/254135/preview16.jpg'
                    item['price'] = products.productPrice
                    item['total_sale'] = item['quantity_sold'] * products.productPrice

                    list_of_items.append(item)
        
            customer_orders = Order.objects.filter(
                customerId = customer["customerId"],
                Status = master_order_status_number["Completed"],
                OrderDate__date__range = (start_date, end_date),
                vendorId = vendor_id
            )

            total_orders_count = customer_orders.count()
            delivery_orders_count = customer_orders.filter(orderType = order_type_number["Delivery"]).count()
            pickup_orders_count = customer_orders.filter(orderType = order_type_number["Pickup"]).count()
            dinein_orders_count = customer_orders.filter(orderType = order_type_number["Dinein"]).count()

            if platform:
                online_orders_count = customer_orders.filter(platform=platform.pk).count()
                offline_orders_count = customer_orders.exclude(platform=platform.pk).count()
            
            customer_info = Customer.objects.filter(pk=customer["customerId"]).first()

            customer_address = Address.objects.filter(customer=customer["customerId"], type='shipping_address', is_selected=True).first()

            loyalty_points_credit_history = LoyaltyPointsCreditHistory.objects.filter(customer=customer["customerId"])

            sum_of_points_credited = loyalty_points_credit_history.aggregate(Sum('points_credited'))['points_credited__sum']
            sum_of_points_redeemed = loyalty_points_credit_history.aggregate(Sum('total_points_redeemed'))['total_points_redeemed__sum']

            if not sum_of_points_credited:
                sum_of_points_credited = 0

            if not sum_of_points_redeemed:
                sum_of_points_redeemed = 0
            
            if customer_address:    
                customer_details.append({
                    "customer_id": customer["customerId"],
                    "first_name": customer_info.FirstName,
                    "last_name": customer_info.LastName if customer_info.LastName else "",
                    "phone_number": customer_info.Phone_Number,
                    "email": customer_info.Email if customer_info.Email else "",
                    "address_line_1": customer_address.address_line1 if customer_address.address_line1 else "",
                    "address_line_2": customer_address.address_line2 if customer_address.address_line2 else "",
                    "city": customer_address.city if customer_address.city else "",
                    "state": customer_address.state if customer_address.state else "",
                    "pincode": customer_address.zipcode if customer_address.zipcode else "",
                    "loyalty_points_balance": customer_info.loyalty_points_balance,
                    "total_points_credited": sum_of_points_credited,
                    "total_points_redeemed": sum_of_points_redeemed,
                    "total_orders_count": total_orders_count,
                    "online_orders_count": online_orders_count if platform else 0,
                    "offline_orders_count": offline_orders_count if platform else 0,
                    "delivery_orders_count": delivery_orders_count,
                    "pickup_orders_count": pickup_orders_count,
                    "dinein_orders_count": dinein_orders_count,
                    "most_ordered_items": list_of_items,
                })

            else:
                customer_details.append({
                "customer_id": customer["customerId"],
                "first_name": customer_info.FirstName,
                "last_name": customer_info.LastName if customer_info.LastName else "",
                "phone_number": customer_info.Phone_Number,
                "email": customer_info.Email if customer_info.Email else "",
                "address_line_1": "",
                "address_line_2": "",
                "city": "",
                "state": "",
                "pincode": "",
                "loyalty_points_balance": customer_info.loyalty_points_balance,
                "total_points_credited": sum_of_points_credited,
                "total_points_redeemed": sum_of_points_redeemed,
                "total_orders_count": total_orders_count,
                "online_orders_count": online_orders_count if platform else 0,
                "offline_orders_count": offline_orders_count if platform else 0,
                "delivery_orders_count": delivery_orders_count,
                "pickup_orders_count": pickup_orders_count,
                "dinein_orders_count": dinein_orders_count,
                "most_ordered_items": list_of_items,
            })

        return JsonResponse({
            "most_repeating_customers": customer_details
        })
    
    elif is_download.lower() == "true":
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')

        formatted_start_date = start_date.strftime('%d-%m-%Y')
        formatted_end_date = end_date.strftime('%d-%m-%Y')

        # Create a new Excel workbook and select the active sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        
        if language == "English":
            sheet.append(['Start Date', f'{formatted_start_date}'])
            sheet.append(['End Date', f'{formatted_end_date}'])
            sheet.append(['Order Type', f'{order_type}'])
            sheet.append(['Top', f'{top_number}'])
            sheet.append(['Sorted by', f'{sort_by}'])
            sheet.append(['Total records', f'{len(top_customers)}'])
            sheet.append([''])

        else:
            order_type = language_localization[order_type]
            sort_by = language_localization[sort_by]

            sheet.append([language_localization['Start Date'], f'{formatted_start_date}'])
            sheet.append([language_localization['End Date'], f'{formatted_end_date}'])
            sheet.append([language_localization['Order Type'], f'{order_type}'])
            sheet.append([language_localization['Top'], f'{top_number}'])
            sheet.append([language_localization['Sorted by'], f'{sort_by}'])
            sheet.append([language_localization['Total records'], f'{len(top_customers)}'])
            sheet.append([''])
        
        if platform:
            if language == "English":
                sheet.append([
                    'Customer Name',
                    'Phone Number',
                    'Email ID',
                    'Address',
                    'Total Points Credited',
                    'Total Points Redeemed',
                    'Total Orders',
                    'Total Online Orders',
                    'Total Offline Orders',
                    'Total Delivery Orders',
                    'Total Pickup Orders',
                    'Total DineIn Orders',
                    'Most Ordered Items'
                ])

            else:
                sheet.append([
                    language_localization['Customer Name'],
                    language_localization['Phone Number'],
                    language_localization['Email ID'],
                    language_localization['Address'],
                    language_localization['Total Points Credited'],
                    language_localization['Total Points Redeemed'],
                    language_localization['Total Orders'],
                    language_localization['Total Online Orders'],
                    language_localization['Total Offline Orders'],
                    language_localization['Total Delivery Orders'],
                    language_localization['Total Pickup Orders'],
                    language_localization['Total DineIn Orders'],
                    language_localization['Most Ordered Items']
                ])

        else:
            if language == "English":
                sheet.append([
                    'Customer Name',
                    'Phone Number',
                    'Email ID',
                    'Address',
                    'Total Points Credited',
                    'Total Points Redeemed',
                    'Total Orders',
                    'Total Delivery Orders',
                    'Total Pickup Orders',
                    'Total DineIn Orders',
                    'Most Ordered Items'
                ])

            else:
                sheet.append([
                    language_localization['Customer Name'],
                    language_localization['Phone Number'],
                    language_localization['Email ID'],
                    language_localization['Address'],
                    language_localization['Total Points Credited'],
                    language_localization['Total Points Redeemed'],
                    language_localization['Total Orders'],
                    language_localization['Total Delivery Orders'],
                    language_localization['Total Pickup Orders'],
                    language_localization['Total DineIn Orders'],
                    language_localization['Most Ordered Items']
                ])

        for customer in top_customers:
            order_items = None

            if order_type == "all":
                order_items = Order_content.objects.filter(
                    orderId__order_status = 10,
                    orderId__master_order__customerId = customer["customerId"],
                    orderId__master_order__Status = master_order_status_number["Completed"],
                    orderId__master_order__OrderDate__date__range = (start_date, end_date),
                    orderId__master_order__vendorId = vendor_id,
                    orderId__vendorId = vendor_id,
                ).exclude(status = 5)
            
            elif order_type == "delivery":
                order_items = Order_content.objects.filter(
                    orderId__order_status = 10,
                    orderId__master_order__customerId = customer["customerId"],
                    orderId__master_order__orderType = order_type_number["Delivery"],
                    orderId__master_order__Status = master_order_status_number["Completed"],
                    orderId__master_order__OrderDate__date__range = (start_date, end_date),
                    orderId__master_order__vendorId = vendor_id,
                    orderId__vendorId = vendor_id,
                ).exclude(status = 5)

            elif order_type == "pickup":
                order_items = Order_content.objects.filter(
                    orderId__order_status = 10,
                    orderId__master_order__customerId = customer["customerId"],
                    orderId__master_order__orderType = order_type_number["Pickup"],
                    orderId__master_order__Status = master_order_status_number["Completed"],
                    orderId__master_order__OrderDate__date__range = (start_date, end_date),
                    orderId__master_order__vendorId = vendor_id,
                    orderId__vendorId = vendor_id,
                ).exclude(status = 5)

            elif order_type == "dinein":
                order_items = Order_content.objects.filter(
                    orderId__order_status = 10,
                    orderId__master_order__customerId = customer["customerId"],
                    orderId__master_order__orderType = order_type_number["Dinein"],
                    orderId__master_order__Status = master_order_status_number["Completed"],
                    orderId__master_order__OrderDate__date__range = (start_date, end_date),
                    orderId__master_order__vendorId = vendor_id,
                    orderId__vendorId = vendor_id,
                ).exclude(status = 5)

            elif order_type == "online":
                order_items = Order_content.objects.filter(
                    orderId__order_status = 10,
                    orderId__master_order__customerId = customer["customerId"],
                    orderId__master_order__platform = platform.pk,
                    orderId__master_order__Status = master_order_status_number["Completed"],
                    orderId__master_order__OrderDate__date__range = (start_date, end_date),
                    orderId__master_order__vendorId = vendor_id,
                    orderId__vendorId = vendor_id,
                ).exclude(status = 5)

            elif order_type == "offline":
                order_items = Order_content.objects.filter(
                    orderId__order_status = 10,
                    orderId__master_order__customerId = customer["customerId"],
                    orderId__master_order__Status = master_order_status_number["Completed"],
                    orderId__master_order__OrderDate__date__range = (start_date, end_date),
                    orderId__master_order__vendorId = vendor_id,
                    orderId__vendorId = vendor_id,
                ).exclude(status = 5, orderId__master_order__platform = platform.pk)
            
            list_of_items = ''
            
            if (order_items != None) and order_items.exists():
                top_selling_items = order_items.values('SKU').distinct() \
                    .annotate(quantity_sold=ExpressionWrapper(Sum('quantity'), output_field=IntegerField())) \
                    .filter(quantity_sold__gt=0) \
                    .order_by('-quantity_sold')[:10]

                for item in top_selling_items:
                    product = Product.objects.filter(PLU=item['SKU'], vendorId=vendor_id).first()

                    product_name = ""

                    if language == "English":
                        product_name = product.productName

                    else:
                        product_name = product.productName_locale
                    
                    if list_of_items:
                        list_of_items = list_of_items + ", " + product_name
                        
                    else:
                        list_of_items = product_name
        
            customer_orders = Order.objects.filter(
                customerId = customer["customerId"],
                Status = master_order_status_number["Completed"],
                OrderDate__date__range = (start_date, end_date),
                vendorId = vendor_id
            )

            total_orders_count = customer_orders.count()

            delivery_orders_count = customer_orders.filter(orderType = order_type_number["Delivery"]).count()
            pickup_orders_count = customer_orders.filter(orderType = order_type_number["Pickup"]).count()
            dinein_orders_count = customer_orders.filter(orderType = order_type_number["Dinein"]).count()

            if platform:
                online_orders_count = customer_orders.filter(platform=platform.pk).count()
                offline_orders_count = customer_orders.exclude(platform=platform.pk).count()
            
            customer_info = Customer.objects.filter(pk=customer["customerId"]).first()

            customer_address = Address.objects.filter(customer=customer["customerId"], type='shipping_address', is_selected=True).first()

            loyalty_points_credit_history = LoyaltyPointsCreditHistory.objects.filter(customer=customer["customerId"])

            sum_of_points_credited = loyalty_points_credit_history.aggregate(Sum('points_credited'))['points_credited__sum']
            sum_of_points_redeemed = loyalty_points_credit_history.aggregate(Sum('total_points_redeemed'))['total_points_redeemed__sum']

            if not sum_of_points_credited:
                sum_of_points_credited = 0

            if not sum_of_points_redeemed:
                sum_of_points_redeemed = 0

            if customer_info.FirstName and customer_info.LastName:
                customer_name = customer_info.FirstName + " " + customer_info.LastName
            elif customer_info.FirstName and not customer_info.LastName:
                customer_name = customer_info.FirstName

            if customer_address:
                address_line_1 = customer_address.address_line1 if customer_address.address_line1 else ""
                address_line_2 = customer_address.address_line2 if customer_address.address_line2 else ""
                city = customer_address.city if customer_address.city else ""
                state = customer_address.state if customer_address.state else ""
                country = customer_address.country if customer_address.country else ""
                zipcode = customer_address.zipcode if customer_address.zipcode else ""
                
                address = address_line_1 + ", " + address_line_2 + ", " + city + ", " + state + ", " + country + ", " + zipcode
            
            else:
                address = ""
            
            if platform:
                sheet.append([
                    customer_name,
                    customer_info.Phone_Number,
                    customer_info.Email if customer_info.Email else "",
                    address,
                    sum_of_points_credited,
                    sum_of_points_redeemed,
                    total_orders_count,
                    online_orders_count,
                    offline_orders_count,
                    delivery_orders_count,
                    pickup_orders_count,
                    dinein_orders_count,
                    list_of_items
                ])

            else:
                sheet.append([
                    customer_name,
                    customer_info.Phone_Number,
                    customer_info.Email if customer_info.Email else "",
                    address,
                    sum_of_points_credited,
                    sum_of_points_redeemed,
                    total_orders_count,
                    delivery_orders_count,
                    pickup_orders_count,
                    dinein_orders_count,
                    list_of_items
                ])
        
        directory = os.path.join(settings.MEDIA_ROOT, 'Excel Downloads')
        
        os.makedirs(directory, exist_ok=True) # Create the directory if it doesn't exist inside MEDIA_ROOT

        file_name = f"Most_repeating_customers_{vendor_id}.xlsx"

        relative_file_path = os.path.join('Excel Downloads', file_name)

        file_path = os.path.join(settings.MEDIA_ROOT, relative_file_path)

        workbook.save(file_path)

        print(f"Excel file '{file_name}' has been created.")
        
        response = "/media/" + relative_file_path
        
        return HttpResponse(response, status=200)
    
    else:
        return Response("Invalid 'download' key", status=status.HTTP_400_BAD_REQUEST)


@api_view(["GET"])
def customers_redeemed_most_points_report(request):
    vendor_id = request.GET.get("vendorId")
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    is_download = request.GET.get('download')
    top_number = request.GET.get('top')
    sort_by = request.GET.get('sort')
    language = request.GET.get('language', 'English')

    if not vendor_id:
        return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id = int(vendor_id)
    except ValueError:
        return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
    
    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor_instance:
        return Response("Vendor does not exist", status=status.HTTP_400_BAD_REQUEST)
    
    if (((not start_date) or (not end_date)) or (start_date > end_date)):
        return Response("Invalid start date or end date", status=status.HTTP_400_BAD_REQUEST)
    
    if not is_download:
        is_download = "false"
    
    else:
        is_download = str(is_download)

    try:
        top_number = int(top_number)
    except ValueError:
        return Response("Invalid top parameter", status=status.HTTP_400_BAD_REQUEST)
    
    if not sort_by:
        sort_by = "descending "
    
    platform = Platform.objects.filter(Name__in=('Mobile App', 'Website'), isActive=True, VendorId=vendor_id).first()
    
    orders = Order.objects.filter(
        Status = master_order_status_number["Completed"],
        OrderDate__date__range = (start_date, end_date),
        vendorId = vendor_id,
    )
    
    top_customers = LoyaltyPointsRedeemHistory.objects.filter(
        vendor=vendor_id,
        redeem_datetime__date__range=(start_date, end_date)
    ).values('customer').distinct() \
        .annotate(total_points_redeemed=Sum('points_redeemed')) \
        .filter(total_points_redeemed__gt=0) \
        .order_by('-total_points_redeemed')[:top_number]

    if sort_by == "descending":
        pass

    elif sort_by == "ascending":
        top_customers = sorted(top_customers, key=itemgetter('total_points_redeemed'))

    else:
        return Response("Invalid sort parameter", status=status.HTTP_400_BAD_REQUEST)

    customer_details = []
    
    if is_download.lower() == "false":
        for customer in top_customers:
            order_items = Order_content.objects.filter(
                orderId__order_status = 10,
                orderId__master_order__customerId = customer["customer"],
                orderId__master_order__Status = master_order_status_number["Completed"],
                orderId__master_order__OrderDate__date__range = (start_date, end_date),
                orderId__master_order__vendorId = vendor_id,
                orderId__vendorId = vendor_id,
            ).exclude(status = 5)
            
            list_of_items = []

            if order_items.exists():
                top_selling_items = order_items.values('SKU').distinct() \
                    .annotate(quantity_sold=ExpressionWrapper(Sum('quantity'), output_field=IntegerField())) \
                    .filter(quantity_sold__gt=0) \
                    .order_by('-quantity_sold')[:10]
                
                for item in top_selling_items:
                    image_list = []

                    products = Product.objects.filter(PLU=item['SKU'], vendorId=vendor_id).first()
                    
                    images = ProductImage.objects.filter(product=products.pk)

                    for image in images:
                        if image.url:
                            image_list.append(image.url)

                    product_name = ""

                    if language == "English":
                        product_name = products.productName

                    else:
                        product_name = products.productName_locale

                    item['product_id'] = products.pk
                    item['product_name'] = product_name
                    item['image'] = image_list[0] if len(image_list)!=0 else 'https://www.stockvault.net/data/2018/08/31/254135/preview16.jpg'
                    item['price'] = products.productPrice
                    item['total_sale'] = item['quantity_sold'] * products.productPrice

                    list_of_items.append(item)
        
            customer_orders = orders.filter(customerId=customer["customer"])

            total_orders_count = customer_orders.count()
            delivery_orders_count = customer_orders.filter(orderType = order_type_number["Delivery"]).count()
            pickup_orders_count = customer_orders.filter(orderType = order_type_number["Pickup"]).count()
            dinein_orders_count = customer_orders.filter(orderType = order_type_number["Dinein"]).count()
            
            online_orders_count = 0
            offline_orders_count = 0
            
            if platform:
                online_orders_count = customer_orders.filter(platform=platform.pk).count()
                offline_orders_count = customer_orders.exclude(platform=platform.pk).count()

            customer_info = Customer.objects.filter(pk=customer["customer"]).first()

            customer_address = Address.objects.filter(customer=customer["customer"], type='shipping_address', is_selected=True).first()
            
            sum_of_points_credited = LoyaltyPointsCreditHistory.objects.filter(
                customer=customer["customer"],
                credit_datetime__date__range=(start_date, end_date)
            ).aggregate(total_points_credited=Sum('points_credited'))['total_points_credited'] or 0

            if customer_address:    
                customer_details.append({
                    "customer_id": customer["customer"],
                    "first_name": customer_info.FirstName,
                    "last_name": customer_info.LastName if customer_info.LastName else "",
                    "phone_number": customer_info.Phone_Number,
                    "email": customer_info.Email if customer_info.Email else "",
                    "address_line_1": customer_address.address_line1 if customer_address.address_line1 else "",
                    "address_line_2": customer_address.address_line2 if customer_address.address_line2 else "",
                    "city": customer_address.city if customer_address.city else "",
                    "state": customer_address.state if customer_address.state else "",
                    "pincode": customer_address.zipcode if customer_address.zipcode else "",
                    "loyalty_points_balance": customer_info.loyalty_points_balance,
                    "total_points_redeemed": customer['total_points_redeemed'],
                    "total_points_credited": sum_of_points_credited,
                    "total_orders_count": total_orders_count,
                    "online_orders_count": online_orders_count,
                    "offline_orders_count": offline_orders_count,
                    "delivery_orders_count": delivery_orders_count,
                    "pickup_orders_count": pickup_orders_count,
                    "dinein_orders_count": dinein_orders_count,
                    "most_ordered_items": list_of_items,
                })

            else:
                customer_details.append({
                "customer_id": customer["customer"],
                "first_name": customer_info.FirstName,
                "last_name": customer_info.LastName if customer_info.LastName else "",
                "phone_number": customer_info.Phone_Number,
                "email": customer_info.Email if customer_info.Email else "",
                "address_line_1": "",
                "address_line_2": "",
                "city": "",
                "state": "",
                "pincode": "",
                "loyalty_points_balance": customer_info.loyalty_points_balance,
                "total_points_redeemed": customer['total_points_redeemed'],
                "total_orders_count": total_orders_count,
                "online_orders_count": online_orders_count,
                "offline_orders_count": offline_orders_count,
                "delivery_orders_count": delivery_orders_count,
                "pickup_orders_count": pickup_orders_count,
                "dinein_orders_count": dinein_orders_count,
                "most_ordered_items": list_of_items,
            })

        return JsonResponse({
            "customers_redeemed_most_points": customer_details
        })
    
    elif is_download.lower() == "true":
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')

        formatted_start_date = start_date.strftime('%d-%m-%Y')
        formatted_end_date = end_date.strftime('%d-%m-%Y')

        # Create a new Excel workbook and select the active sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        if platform:
            sheet.append(['Start Date', f'{formatted_start_date}'])
            sheet.append(['End Date', f'{formatted_end_date}'])
            sheet.append(['Top', f'{top_number}'])
            sheet.append(['Sorted by', f'{sort_by}'])
            sheet.append(['Total records', f'{len(top_customers)}'])
            sheet.append([''])

            sheet.append([
                'Customer Name',
                'Phone Number',
                'Email ID',
                'Address',
                'Total Points Redeemed',
                'Total Orders',
                'Total Online Orders',
                'Total Offline Orders',
                'Total Delivery Orders',
                'Total Pickup Orders',
                'Total DineIn Orders',
                'Most Ordered Items'
            ])

        else:    
            sheet.append(['Start Date', f'{formatted_start_date}'])
            sheet.append(['End Date', f'{formatted_end_date}'])
            sheet.append(['Top', f'{top_number}'])
            sheet.append(['Sorted by', f'{sort_by}'])
            sheet.append(['Total records', f'{len(top_customers)}'])
            sheet.append([''])

            sheet.append([
                'Customer Name',
                'Phone Number',
                'Email ID',
                'Address',
                'Total Points Redeemed',
                'Total Orders',
                'Total Delivery Orders',
                'Total Pickup Orders',
                'Total DineIn Orders',
                'Most Ordered Items'
            ])

        for customer in top_customers:
            order_items = Order_content.objects.filter(
                orderId__order_status = 10,
                orderId__master_order__customerId = customer["customer"],
                orderId__master_order__Status = master_order_status_number["Completed"],
                orderId__master_order__OrderDate__date__range = (start_date, end_date),
                orderId__master_order__vendorId = vendor_id,
                orderId__vendorId = vendor_id,
            ).exclude(status = 5)

            top_selling_items = order_items.values('SKU').distinct() \
                                .annotate(quantity_sold=ExpressionWrapper(Sum('quantity'), output_field=IntegerField())) \
                                .filter(quantity_sold__gt=0) \
                                .order_by('-quantity_sold')[:10]

            list_of_items = ''
              
            if order_items.exists():    
                for item in top_selling_items:
                    product = Product.objects.filter(PLU=item['SKU'], vendorId=vendor_id).first()

                    product_name = ""

                    if language == "English":
                        product_name = product.productName

                    else:
                        product_name = product.productName_locale
                    
                    if list_of_items:
                        list_of_items = list_of_items + ", " + product_name

                    else:
                        list_of_items = product_name
    
            customer_orders = orders.filter(customerId=customer["customer"])

            total_orders_count = customer_orders.count()
            delivery_orders_count = customer_orders.filter(orderType = order_type_number["Delivery"]).count()
            pickup_orders_count = customer_orders.filter(orderType = order_type_number["Pickup"]).count()
            dinein_orders_count = customer_orders.filter(orderType = order_type_number["Dinein"]).count()

            if platform:
                online_orders_count = customer_orders.filter(platform = platform.pk).count()
                offline_orders_count = customer_orders.exclude(platform = platform.pk).count()
            
            customer_info = Customer.objects.filter(pk = customer["customer"]).first()

            customer_address = Address.objects.filter(customer = customer["customer"], type = 'shipping_address', is_selected = True).first()

            if customer_info.FirstName and customer_info.LastName:
                customer_name = customer_info.FirstName + " " + customer_info.LastName

            elif customer_info.FirstName and not customer_info.LastName:
                customer_name = customer_info.FirstName

            if customer_address:
                address_line_1 = customer_address.address_line1 if customer_address.address_line1 else ""
                address_line_2 = customer_address.address_line2 if customer_address.address_line2 else ""
                city = customer_address.city if customer_address.city else ""
                state = customer_address.state if customer_address.state else ""
                country = customer_address.country if customer_address.country else ""
                zipcode = customer_address.zipcode if customer_address.zipcode else ""
                
                address = address_line_1 + ", " + address_line_2 + ", " + city + ", " + state + ", " + country + ", " + zipcode
            
            else:
                address = ""
            
            if platform:
                sheet.append([
                customer_name,
                customer_info.Phone_Number,
                customer_info.Email if customer_info.Email else "",
                address,
                customer["total_points_redeemed"],
                total_orders_count,
                online_orders_count,
                offline_orders_count,
                delivery_orders_count,
                pickup_orders_count,
                dinein_orders_count,
                list_of_items
            ])

            else:    
                sheet.append([
                    customer_name,
                    customer_info.Phone_Number,
                    customer_info.Email if customer_info.Email else "",
                    address,
                    customer["total_points_redeemed"],
                    total_orders_count,
                    delivery_orders_count,
                    pickup_orders_count,
                    dinein_orders_count,
                    list_of_items
                ])
        
        directory = os.path.join(settings.MEDIA_ROOT, 'Excel Downloads')
        os.makedirs(directory, exist_ok=True) # Create the directory if it doesn't exist inside MEDIA_ROOT

        file_name = f"Customers_redeemed_most_points_{vendor_id}.xlsx"

        relative_file_path = os.path.join('Excel Downloads', file_name)

        file_path = os.path.join(settings.MEDIA_ROOT, relative_file_path)

        workbook.save(file_path)

        print(f"Excel file '{file_name}' has been created.")
        
        response = "/media/" + relative_file_path
        
        return HttpResponse(response, status=200)
    
    else:
        return Response("Invalid 'download' key", status=status.HTTP_400_BAD_REQUEST)


@api_view(["GET"])
def finance_report(request):
    vendor_id = request.GET.get("vendorId")
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    is_download = request.GET.get('download')
    language = request.GET.get('language', 'English')

    if not vendor_id:
        return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id = int(vendor_id)
    except ValueError:
        return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
    
    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor_instance:
        return Response("Vendor does not exist", status=status.HTTP_400_BAD_REQUEST)
    
    if (((not start_date) or (not end_date)) or (start_date > end_date)):
        return Response("Invalid start date or end date", status=status.HTTP_400_BAD_REQUEST)
    
    if not is_download:
        is_download = "false"
    
    else:
        is_download = str(is_download)

    platform = Platform.objects.filter(Name__in=('Mobile App', 'Website'), isActive=True, VendorId=vendor_id).first()
    
    orders = OrderPayment.objects.filter(
        status = True,
        orderId__OrderDate__date__range = (start_date, end_date),
        orderId__vendorId = vendor_id,
        orderId__masterOrder = None
    ).exclude(orderId__Status = master_order_status_number["Canceled"])

    delivery_orders = orders.filter(orderId__orderType = order_type_number["Delivery"])
    pickup_orders = orders.filter(orderId__orderType = order_type_number["Pickup"])
    dinein_orders = orders.filter(orderId__orderType = order_type_number["Dinein"])
    cash_payment_orders = orders.filter(type = payment_type_number["Cash"])
    online_payment_orders = orders.filter(type = payment_type_number["Online"])
    card_payment_orders = orders.filter(type = payment_type_number["Card"])
    
    total_orders = orders.count()
    delivery_orders_count = delivery_orders.count()
    pickup_orders_count = pickup_orders.count()
    dinein_orders_count = dinein_orders.count()
    cash_payment_orders_count = cash_payment_orders.count()
    online_payment_orders_count = online_payment_orders.count()
    card_payment_orders_count = card_payment_orders.count()
    
    total_tax_collected = orders.aggregate(total_tax=Sum('orderId__tax'))['total_tax'] or 0.0
    tax_collected_from_delivery = delivery_orders.aggregate(total_tax=Sum('orderId__tax'))['total_tax'] or 0.0
    tax_collected_from_pickup = pickup_orders.aggregate(total_tax=Sum('orderId__tax'))['total_tax'] or 0.0
    tax_collected_from_dinein = dinein_orders.aggregate(total_tax=Sum('orderId__tax'))['total_tax'] or 0.0
    tax_collected_from_cash_payment = cash_payment_orders.aggregate(total_tax=Sum('orderId__tax'))['total_tax'] or 0.0
    tax_collected_from_online_payment = online_payment_orders.aggregate(total_tax=Sum('orderId__tax'))['total_tax'] or 0.0
    tax_collected_from_card_payment = card_payment_orders.aggregate(total_tax=Sum('orderId__tax'))['total_tax'] or 0.0

    total_orders_aggregation = orders.aggregate(
        subtotal_sum=Coalesce(Sum('orderId__subtotal'), 0.0),
        discount_sum=Coalesce(Sum('orderId__discount'), 0.0)
    )

    delivery_orders_aggregation = delivery_orders.aggregate(
        subtotal_sum=Coalesce(Sum('orderId__subtotal'), 0.0),
        discount_sum=Coalesce(Sum('orderId__discount'), 0.0)
    )

    pickup_orders_aggregation = pickup_orders.aggregate(
        subtotal_sum=Coalesce(Sum('orderId__subtotal'), 0.0),
        discount_sum=Coalesce(Sum('orderId__discount'), 0.0)
    )

    dinein_orders_aggregation = dinein_orders.aggregate(
        subtotal_sum=Coalesce(Sum('orderId__subtotal'), 0.0),
        discount_sum=Coalesce(Sum('orderId__discount'), 0.0)
    )

    cash_payment_orders_aggregation = cash_payment_orders.aggregate(
        subtotal_sum=Coalesce(Sum('orderId__subtotal'), 0.0),
        discount_sum=Coalesce(Sum('orderId__discount'), 0.0)
    )

    online_payment_orders_aggregation = online_payment_orders.aggregate(
        subtotal_sum=Coalesce(Sum('orderId__subtotal'), 0.0),
        discount_sum=Coalesce(Sum('orderId__discount'), 0.0)
    )

    card_payment_orders_aggregation = card_payment_orders.aggregate(
        subtotal_sum=Coalesce(Sum('orderId__subtotal'), 0.0),
        discount_sum=Coalesce(Sum('orderId__discount'), 0.0)
    )
    
    total_revenue = total_orders_aggregation['subtotal_sum'] - total_orders_aggregation['discount_sum']
    revenue_from_delivery = delivery_orders_aggregation['subtotal_sum'] - delivery_orders_aggregation['discount_sum']
    revenue_from_pickup = pickup_orders_aggregation['subtotal_sum'] - pickup_orders_aggregation['discount_sum']
    revenue_from_dinein = dinein_orders_aggregation['subtotal_sum'] - dinein_orders_aggregation['discount_sum']
    revenue_from_cash_payment = cash_payment_orders_aggregation['subtotal_sum'] - cash_payment_orders_aggregation['discount_sum']
    revenue_from_online_payment = online_payment_orders_aggregation['subtotal_sum'] - online_payment_orders_aggregation['discount_sum']
    revenue_from_card_payment = card_payment_orders_aggregation['subtotal_sum'] - card_payment_orders_aggregation['discount_sum']
    
    online_orders_count = 0
    offline_orders_count = 0
        
    tax_collected_from_online = 0.0
    tax_collected_from_offline = 0.0

    revenue_from_online = 0.0
    revenue_from_offline = 0.0
    
    if platform:
        online_orders = orders.filter(orderId__platform=platform.pk)
        offline_orders = orders.exclude(orderId__platform=platform.pk)

        online_orders_count = online_orders.count()
        offline_orders_count = offline_orders.count()

        tax_collected_from_online = online_orders.aggregate(total_tax=Sum('orderId__tax'))['total_tax'] or 0.0
        tax_collected_from_offline = offline_orders.aggregate(total_tax=Sum('orderId__tax'))['total_tax'] or 0.0

        online_orders_aggregation = online_orders.aggregate(
            subtotal_sum=Coalesce(Sum('orderId__subtotal'), 0.0),
            discount_sum=Coalesce(Sum('orderId__discount'), 0.0)
        )

        offline_orders_aggregation = offline_orders.aggregate(
            subtotal_sum=Coalesce(Sum('orderId__subtotal'), 0.0),
            discount_sum=Coalesce(Sum('orderId__discount'), 0.0)
        )
        
        revenue_from_online = online_orders_aggregation['subtotal_sum'] - online_orders_aggregation['discount_sum']
        revenue_from_offline = offline_orders_aggregation['subtotal_sum'] - offline_orders_aggregation['discount_sum']
    
    if is_download.lower() == "false":
        return JsonResponse({
            "total_orders": total_orders,
            "delivery_orders_count": delivery_orders_count,
            "pickup_orders_count": pickup_orders_count,
            "dinein_orders_count": dinein_orders_count,
            "online_orders_count": online_orders_count,
            "offline_orders_count": offline_orders_count,
            "cash_payment_orders_count": cash_payment_orders_count,
            "online_payment_orders_count": online_payment_orders_count,
            "card_payment_orders_count": card_payment_orders_count,
            "total_tax_collected": round(total_tax_collected, 2),
            "tax_collected_from_delivery": round(tax_collected_from_delivery, 2),
            "tax_collected_from_pickup": round(tax_collected_from_pickup, 2),
            "tax_collected_from_dinein": round(tax_collected_from_dinein, 2),
            "tax_collected_from_online": round(tax_collected_from_online, 2),
            "tax_collected_from_offline": round(tax_collected_from_offline, 2),
            "tax_collected_from_cash_payment": round(tax_collected_from_cash_payment, 2),
            "tax_collected_from_online_payment": round(tax_collected_from_online_payment, 2),
            "tax_collected_from_card_payment": round(tax_collected_from_card_payment, 2),
            "total_revenue": round(total_revenue, 2),
            "revenue_from_delivery": round(revenue_from_delivery, 2),
            "revenue_from_pickup": round(revenue_from_pickup, 2),
            "revenue_from_dinein": round(revenue_from_dinein, 2),
            "revenue_from_online": round(revenue_from_online, 2),
            "revenue_from_offline": round(revenue_from_offline, 2),
            "revenue_from_cash_payment": round(revenue_from_cash_payment, 2),
            "revenue_from_online_payment": round(revenue_from_online_payment, 2),
            "revenue_from_card_payment": round(revenue_from_card_payment, 2)
        })
    
    elif is_download.lower() == "true":
        report_data = {
            'Delivery Orders': {'orders': delivery_orders_count, 'tax': tax_collected_from_delivery, 'revenue': revenue_from_delivery},
            'Pickup Orders': {'orders': pickup_orders_count, 'tax': tax_collected_from_pickup, 'revenue': revenue_from_pickup},
            'DineIn Orders': {'orders': dinein_orders_count, 'tax': tax_collected_from_dinein, 'revenue': revenue_from_dinein},
            'Total': {'orders': total_orders, 'tax': total_tax_collected, 'revenue': total_revenue},
            'Cash Payment': {'orders': cash_payment_orders_count, 'tax': tax_collected_from_cash_payment, 'revenue': revenue_from_cash_payment},
            'Online Payment': {'orders': online_payment_orders_count, 'tax': tax_collected_from_online_payment, 'revenue': revenue_from_online_payment},
            'Card Payment': {'orders': card_payment_orders_count, 'tax': tax_collected_from_card_payment, 'revenue': revenue_from_card_payment},
        }

        if platform:
            report_data['Online Orders'] = {'orders': online_orders_count, 'tax': tax_collected_from_online, 'revenue': revenue_from_online}
            report_data['Offline Orders'] = {'orders': offline_orders_count, 'tax': tax_collected_from_offline, 'revenue': revenue_from_offline}

        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')

        formatted_start_date = start_date.strftime('%d-%m-%Y')
        formatted_end_date = end_date.strftime('%d-%m-%Y')

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        if language == "English":
            sheet.append(['Start Date', f'{formatted_start_date}'])
            sheet.append(['End Date', f'{formatted_end_date}'])
            sheet.append([''])

            sheet.append(['Category', 'Total Orders', 'Tax Collected', 'Revenue Generated'])

        else:
            sheet.append([language_localization['Start Date'], f'{formatted_start_date}'])
            sheet.append([language_localization['End Date'], f'{formatted_end_date}'])
            sheet.append([''])
        
            sheet.append([
                language_localization['Category'],
                language_localization['Total Orders'],
                language_localization['Tax Collected'],
                language_localization['Revenue Generated']
            ])

        if language == "English":
            for key, value in report_data.items():
                sheet.append([key, value["orders"], value["tax"], value["revenue"]])

        else:
            for key, value in report_data.items():
                sheet.append([language_localization[key], value["orders"], value["tax"], value["revenue"]])
        
        directory = os.path.join(settings.MEDIA_ROOT, 'Excel Downloads')
        os.makedirs(directory, exist_ok=True) # Create the directory if it doesn't exist inside MEDIA_ROOT

        file_name = f"Finance_Report_{vendor_id}.xlsx"

        relative_file_path = os.path.join('Excel Downloads', file_name)

        file_path = os.path.join(settings.MEDIA_ROOT, relative_file_path)

        workbook.save(file_path)

        print(f"Excel file '{file_name}' has been created.")
        
        response = "/media/" + relative_file_path
        
        return HttpResponse(response, status=200)
    
    else:
        return Response("Invalid 'download' key", status=status.HTTP_400_BAD_REQUEST)


@api_view(["GET"])
def footfall_revenue_report(request):
    vendor_id = request.GET.get("vendorId")
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    is_download = request.GET.get('download')
    filter_by = request.GET.get('filter')
    filter_type = request.GET.get('type')
    language = request.GET.get('language', 'English')

    if not vendor_id:
        return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id = int(vendor_id)

    except ValueError:
        return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
    
    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor_instance:
        return Response("Vendor does not exist", status=status.HTTP_400_BAD_REQUEST)
    
    if (((not start_date) or (not end_date)) or (start_date > end_date)):
        return Response("Invalid start date or end date", status=status.HTTP_400_BAD_REQUEST)
    
    if not is_download:
        is_download = "false"
    
    else:
        is_download = str(is_download)

    if filter_by not in ('date', 'month', 'day', 'hour'):
        return Response("Invalid filter parameter", status=status.HTTP_400_BAD_REQUEST)
    
    if filter_type not in ('footfall', 'revenue'):
        return Response("Invalid type parameter", status=status.HTTP_400_BAD_REQUEST)
    
    platform = Platform.objects.filter(Name__in=('Mobile App', 'Website'), isActive=True, VendorId=vendor_id).first()
    
    all_orders = OrderPayment.objects.filter(
        status = True,
        orderId__OrderDate__date__range = (start_date, end_date),
        orderId__vendorId = vendor_id,
        orderId__masterOrder = None
    ).exclude(orderId__Status = master_order_status_number["Canceled"])

    # ISO standard mapping
    weekday_names = {
        1: 'Monday',
        2: 'Tuesday',
        3: 'Wednesday',
        4: 'Thursday',
        5: 'Friday',
        6: 'Saturday',
        7: 'Sunday'
    }
    
    if filter_type == 'footfall':
        if filter_by == 'date':
            top_data = all_orders.values('orderId__OrderDate__date').distinct() \
                        .annotate(total_orders=Count('orderId__id')) \
                        .filter(total_orders__gt=0) \
                        .order_by('-total_orders')[:5]

        
        elif filter_by == 'month':
            top_data = all_orders.annotate(order_month=ExtractMonth('orderId__OrderDate')) \
                        .values('order_month').distinct() \
                        .annotate(total_orders=Count('orderId__id')) \
                        .filter(total_orders__gt=0) \
                        .order_by('-total_orders')[:6]
        
        elif filter_by == 'day':
            top_data = all_orders.annotate(weekday=ExtractWeekDay('orderId__OrderDate')) \
                .values('weekday').distinct() \
                .annotate(total_orders=Count('orderId__id')) \
                .filter(total_orders__gt=0) \
                .order_by('-total_orders')[:4]

            weekday_counts = defaultdict(int)

            for order in all_orders:
                weekday = order.orderId.OrderDate.weekday() + 1
                weekday_counts[weekday] += 1

            non_zero_weekdays = {}
            
            for weekday, count in weekday_counts.items():
                if count > 0:
                    non_zero_weekdays[weekday] = count

            top_data = sorted(non_zero_weekdays, key=non_zero_weekdays.get, reverse=True)[:4]
        
        elif filter_by == 'hour':
            top_data = all_orders.annotate(order_hour=ExtractHour('orderId__OrderDate')) \
                        .values('order_hour').distinct() \
                        .annotate(total_orders=Count('orderId__id')) \
                        .filter(total_orders__gt=0) \
                        .order_by('-total_orders')[:5]
    
    elif filter_type == 'revenue':
        if filter_by == 'date':
            top_data = all_orders.values('orderId__OrderDate__date').distinct() \
                        .annotate(subtotal_sum=Sum('orderId__subtotal')) \
                        .filter(subtotal_sum__gt=0) \
                        .order_by('-subtotal_sum')[:5]

        elif filter_by == 'month':
            top_data = all_orders.annotate(order_month=ExtractMonth('orderId__OrderDate')) \
                        .values('order_month').distinct() \
                        .annotate(subtotal_sum=Sum('orderId__subtotal')) \
                        .filter(subtotal_sum__gt=0) \
                        .order_by('-subtotal_sum')[:6]
        
        elif filter_by == 'day':
            top_data = all_orders.annotate(weekday=(ExtractWeekDay('orderId__OrderDate'))) \
                        .values('weekday').distinct() \
                        .annotate(subtotal_sum=Sum('orderId__subtotal')) \
                        .filter(subtotal_sum__gt=0) \
                        .order_by('-subtotal_sum')[:4]
            
            weekday_subtotal_sum = defaultdict(int)

            for order in all_orders:
                weekday = order.orderId.OrderDate.weekday() + 1
                weekday_subtotal_sum[weekday] += order.orderId.subtotal

            non_zero_weekdays = {}

            for weekday, subtotal_sum in weekday_subtotal_sum.items():
                if subtotal_sum > 0:
                    non_zero_weekdays[weekday] = subtotal_sum

            top_data = sorted(non_zero_weekdays, key=non_zero_weekdays.get, reverse=True)[:4]
    
        elif filter_by == 'hour':
            top_data = all_orders.annotate(order_hour=ExtractHour('orderId__OrderDate')) \
                        .values('order_hour').distinct() \
                        .annotate(subtotal_sum=Sum('orderId__subtotal')) \
                        .filter(subtotal_sum__gt=0) \
                        .order_by('-subtotal_sum')[:5]
    
    filtered_data = []
    
    for instance in top_data:
        if filter_by == 'date':
            orders = all_orders.filter(orderId__OrderDate__date=instance['orderId__OrderDate__date'].strftime('%Y-%m-%d'))

            actual_instance = instance['orderId__OrderDate__date'].strftime('%Y-%m-%d')

        elif filter_by == 'month':
            orders = all_orders.filter(orderId__OrderDate__month=instance["order_month"])

            actual_instance = calendar.month_name[instance["order_month"]]
        
        elif filter_by == 'day':
            orders = all_orders.filter(orderId__OrderDate__week_day=instance+1)

            for key,value in weekday_names.items():
                if key == instance:
                    actual_instance = value

        elif filter_by == 'hour':
            orders = all_orders.filter(orderId__OrderDate__hour=instance['order_hour'])

            actual_instance = datetime.strptime(str(instance['order_hour']), '%H').strftime('%I%p') + ' - ' + \
            datetime.strptime(str(instance['order_hour'] + 1), '%H').strftime('%I%p')

        delivery_orders = orders.filter(orderId__orderType = order_type_number["Delivery"])
        pickup_orders = orders.filter(orderId__orderType = order_type_number["Pickup"])
        dinein_orders = orders.filter(orderId__orderType = order_type_number["Dinein"])
        cash_payment_orders = orders.filter(type = payment_type_number["Cash"])
        online_payment_orders = orders.filter(type = payment_type_number["Online"])
        card_payment_orders = orders.filter(type = payment_type_number["Card"])

        total_orders = orders.count()
        delivery_orders_count = delivery_orders.count()
        pickup_orders_count = pickup_orders.count()
        dinein_orders_count = dinein_orders.count()
        cash_payment_orders_count = cash_payment_orders.count()
        online_payment_orders_count = online_payment_orders.count()
        card_payment_orders_count = card_payment_orders.count()
        
        total_tax_collected = orders.aggregate(total_tax=Sum('orderId__tax'))['total_tax'] or 0.0
        tax_collected_from_delivery = delivery_orders.aggregate(total_tax=Sum('orderId__tax'))['total_tax'] or 0.0
        tax_collected_from_pickup = pickup_orders.aggregate(total_tax=Sum('orderId__tax'))['total_tax'] or 0.0
        tax_collected_from_dinein = dinein_orders.aggregate(total_tax=Sum('orderId__tax'))['total_tax'] or 0.0
        tax_collected_from_cash_payment = cash_payment_orders.aggregate(total_tax=Sum('orderId__tax'))['total_tax'] or 0.0
        tax_collected_from_online_payment = online_payment_orders.aggregate(total_tax=Sum('orderId__tax'))['total_tax'] or 0.0
        tax_collected_from_card_payment = card_payment_orders.aggregate(total_tax=Sum('orderId__tax'))['total_tax'] or 0.0

        total_orders_aggregation = orders.aggregate(
            subtotal_sum=Coalesce(Sum('orderId__subtotal'), 0.0),
            discount_sum=Coalesce(Sum('orderId__discount'), 0.0)
        )

        delivery_orders_aggregation = delivery_orders.aggregate(
            subtotal_sum=Coalesce(Sum('orderId__subtotal'), 0.0),
            discount_sum=Coalesce(Sum('orderId__discount'), 0.0)
        )

        pickup_orders_aggregation = pickup_orders.aggregate(
            subtotal_sum=Coalesce(Sum('orderId__subtotal'), 0.0),
            discount_sum=Coalesce(Sum('orderId__discount'), 0.0)
        )

        dinein_orders_aggregation = dinein_orders.aggregate(
            subtotal_sum=Coalesce(Sum('orderId__subtotal'), 0.0),
            discount_sum=Coalesce(Sum('orderId__discount'), 0.0)
        )

        cash_payment_orders_aggregation = cash_payment_orders.aggregate(
            subtotal_sum=Coalesce(Sum('orderId__subtotal'), 0.0),
            discount_sum=Coalesce(Sum('orderId__discount'), 0.0)
        )

        online_payment_orders_aggregation = online_payment_orders.aggregate(
            subtotal_sum=Coalesce(Sum('orderId__subtotal'), 0.0),
            discount_sum=Coalesce(Sum('orderId__discount'), 0.0)
        )

        card_payment_orders_aggregation = card_payment_orders.aggregate(
            subtotal_sum=Coalesce(Sum('orderId__subtotal'), 0.0),
            discount_sum=Coalesce(Sum('orderId__discount'), 0.0)
        )
        
        total_revenue = total_orders_aggregation['subtotal_sum'] - total_orders_aggregation['discount_sum']
        revenue_from_delivery = delivery_orders_aggregation['subtotal_sum'] - delivery_orders_aggregation['discount_sum']
        revenue_from_pickup = pickup_orders_aggregation['subtotal_sum'] - pickup_orders_aggregation['discount_sum']
        revenue_from_dinein = dinein_orders_aggregation['subtotal_sum'] - dinein_orders_aggregation['discount_sum']
        revenue_from_cash_payment = cash_payment_orders_aggregation['subtotal_sum'] - cash_payment_orders_aggregation['discount_sum']
        revenue_from_online_payment = online_payment_orders_aggregation['subtotal_sum'] - online_payment_orders_aggregation['discount_sum']
        revenue_from_card_payment = card_payment_orders_aggregation['subtotal_sum'] - card_payment_orders_aggregation['discount_sum']
        
        online_orders_count = 0
        offline_orders_count = 0

        tax_collected_from_online = 0.0
        tax_collected_from_offline = 0.0

        revenue_from_online = 0.0
        revenue_from_offline = 0.0
        
        if platform:
            online_orders = orders.filter(orderId__platform=platform.pk)
            offline_orders = orders.exclude(orderId__platform=platform.pk)

            online_orders_count = online_orders.count()
            offline_orders_count = offline_orders.count()

            tax_collected_from_online = online_orders.aggregate(total_tax=Sum('orderId__tax'))['total_tax'] or 0.0
            tax_collected_from_offline = offline_orders.aggregate(total_tax=Sum('orderId__tax'))['total_tax'] or 0.0

            online_orders_aggregation = online_orders.aggregate(
                subtotal_sum=Coalesce(Sum('orderId__subtotal'), 0.0),
                discount_sum=Coalesce(Sum('orderId__discount'), 0.0)
            )

            offline_orders_aggregation = offline_orders.aggregate(
                subtotal_sum=Coalesce(Sum('orderId__subtotal'), 0.0),
                discount_sum=Coalesce(Sum('orderId__discount'), 0.0)
            )
            
            revenue_from_online = online_orders_aggregation['subtotal_sum'] - online_orders_aggregation['discount_sum']
            revenue_from_offline = offline_orders_aggregation['subtotal_sum'] - offline_orders_aggregation['discount_sum']

        filtered_data.append({
            "instance": actual_instance,
            "total_orders": total_orders,
            "delivery_orders_count": delivery_orders_count,
            "pickup_orders_count": pickup_orders_count,
            "dinein_orders_count": dinein_orders_count,
            "offline_orders_count": offline_orders_count,
            "online_orders_count": online_orders_count,
            "cash_payment_orders_count": cash_payment_orders_count,
            "online_payment_orders_count": online_payment_orders_count,
            "card_payment_orders_count": card_payment_orders_count,
            "total_tax_collected": round(total_tax_collected, 2),
            "tax_collected_from_delivery": round(tax_collected_from_delivery, 2),
            "tax_collected_from_pickup": round(tax_collected_from_pickup, 2),
            "tax_collected_from_dinein": round(tax_collected_from_dinein, 2),
            "tax_collected_from_offline": round(tax_collected_from_offline, 2),
            "tax_collected_from_online": round(tax_collected_from_online, 2),
            "tax_collected_from_cash_payment": round(tax_collected_from_cash_payment, 2),
            "tax_collected_from_online_payment": round(tax_collected_from_online_payment, 2),
            "tax_collected_from_card_payment": round(tax_collected_from_card_payment, 2),
            "total_revenue": round(total_revenue, 2),
            "revenue_from_delivery": round(revenue_from_delivery, 2),
            "revenue_from_pickup": round(revenue_from_pickup, 2),
            "revenue_from_dinein": round(revenue_from_dinein, 2),
            "revenue_from_offline": round(revenue_from_offline, 2),
            "revenue_from_online": round(revenue_from_online, 2),
            "revenue_from_cash_payment": round(revenue_from_cash_payment, 2),
            "revenue_from_online_payment": round(revenue_from_online_payment, 2),
            "revenue_from_card_payment": round(revenue_from_card_payment, 2)
        })
    
    if is_download.lower() == "false":
        return JsonResponse({"data": filtered_data})
    
    elif is_download.lower() == "true":
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')

        formatted_start_date = start_date.strftime('%d-%m-%Y')
        formatted_end_date = end_date.strftime('%d-%m-%Y')

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        if language == "English":
            sheet.append(['Start Date', f'{formatted_start_date}'])
            sheet.append(['End Date', f'{formatted_end_date}'])
            sheet.append(['Filtered by', f'{filter_by}'])

        else:
            sheet.append([language_localization['Start Date'], f'{formatted_start_date}'])
            sheet.append([language_localization['End Date'], f'{formatted_end_date}'])
            sheet.append([language_localization['Filtered by'], f'{filter_by}'])

        sheet.append([''])

        if platform:
            if language == "English":
                sheet.append(['Order Count Data'])

                sheet.append([
                    'Instance',
                    'Total Orders',
                    'Total Delivery Orders',
                    'Total Pickup Orders',
                    'Total DineIn Orders',
                    'Total Offline Orders',
                    'Total Online Orders',
                    'Cash Payment Orders',
                    'Online Payment Orders',
                    'Card Payment Orders',
                ])

            else:
                sheet.append([language_localization['Order Count Data']])

                sheet.append([
                    language_localization['Instance'],
                    language_localization['Total Orders'],
                    language_localization['Total Delivery Orders'],
                    language_localization['Total Pickup Orders'],
                    language_localization['Total DineIn Orders'],
                    language_localization['Total Offline Orders'],
                    language_localization['Total Online Orders'],
                    language_localization['Cash Payment Orders'],
                    language_localization['Online Payment Orders'],
                    language_localization['Card Payment Orders'],
                ])

            for data in filtered_data:
                sheet.append([
                    data['instance'],
                    data['total_orders'],
                    data['delivery_orders_count'],
                    data['pickup_orders_count'],
                    data['dinein_orders_count'],
                    data['offline_orders_count'],
                    data['online_orders_count'],
                    data['cash_payment_orders_count'],
                    data['online_payment_orders_count'],
                    data['card_payment_orders_count'],
                ])

            sheet.append([''])

            if language == "English":
                sheet.append(['Tax Collection Data'])

                sheet.append([
                    'Instance',
                    'Total Tax Collected',
                    'Tax Collection from Delivery',
                    'Tax Collection from Pickup',
                    'Tax Collection from DineIn',
                    'Tax Collection from Offline Orders',
                    'Tax Collection from Online Orders',
                    'Tax Collection from Cash Payment',
                    'Tax Collection from Online Payment',
                    'Tax Collection from Card Payment',
                ])

            else:
                sheet.append([language_localization['Tax Collection Data']])

                sheet.append([
                    language_localization['Instance'],
                    language_localization['Total Tax Collected'],
                    language_localization['Tax Collection from Delivery'],
                    language_localization['Tax Collection from Pickup'],
                    language_localization['Tax Collection from DineIn'],
                    language_localization['Tax Collection from Offline Orders'],
                    language_localization['Tax Collection from Online Orders'],
                    language_localization['Tax Collection from Cash Payment'],
                    language_localization['Tax Collection from Online Payment'],
                    language_localization['Tax Collection from Card Payment'],
                ])


            for data in filtered_data:
                sheet.append([
                    data['instance'],
                    data['total_tax_collected'],
                    data['tax_collected_from_delivery'],
                    data['tax_collected_from_pickup'],
                    data['tax_collected_from_dinein'],
                    data['tax_collected_from_offline'],
                    data['tax_collected_from_online'],
                    data['tax_collected_from_cash_payment'],
                    data['tax_collected_from_online_payment'],
                    data['tax_collected_from_card_payment'],
                ])
            
            sheet.append([''])

            if language == "English":
                sheet.append(['Revenue Generation Data'])

                sheet.append([
                    'Instance',
                    'Total Revenue Generated',
                    'Revenue from Delivery',
                    'Revenue from Pickup',
                    'Revenue from DineIn',
                    'Revenue from Offline Orders',
                    'Revenue from Online Orders',
                    'Revenue from Cash Payment',
                    'Revenue from Online Payment',
                    'Revenue from Card Payment',
                ])

            else:
                sheet.append([language_localization['Revenue Generation Data']])

                sheet.append([
                    language_localization['Instance'],
                    language_localization['Total Revenue Generated'],
                    language_localization['Revenue from Delivery'],
                    language_localization['Revenue from Pickup'],
                    language_localization['Revenue from DineIn'],
                    language_localization['Revenue from Offline Orders'],
                    language_localization['Revenue from Online Orders'],
                    language_localization['Revenue from Cash Payment'],
                    language_localization['Revenue from Online Payment'],
                    language_localization['Revenue from Card Payment'],
                ])

            for data in filtered_data:
                sheet.append([
                    data['instance'],
                    data['total_revenue'],
                    data['revenue_from_delivery'],
                    data['revenue_from_pickup'],
                    data['revenue_from_dinein'],
                    data['revenue_from_offline'],
                    data['revenue_from_online'],
                    data['revenue_from_cash_payment'],
                    data['revenue_from_online_payment'],
                    data['revenue_from_card_payment'],
                ])

        else:
            if language == "English":
                sheet.append(['Order Count Data'])

                sheet.append([
                    'Instance',
                    'Total Orders',
                    'Total Delivery Orders',
                    'Total Pickup Orders',
                    'Total DineIn Orders',
                    'Cash Payment Orders',
                    'Online Payment Orders',
                    'Card Payment Orders',
                ])

            else:
                sheet.append([language_localization['Order Count Data']])

                sheet.append([
                    language_localization['Instance'],
                    language_localization['Total Orders'],
                    language_localization['Total Delivery Orders'],
                    language_localization['Total Pickup Orders'],
                    language_localization['Total DineIn Orders'],
                    language_localization['Cash Payment Orders'],
                    language_localization['Online Payment Orders'],
                    language_localization['Card Payment Orders'],
                ])

            for data in filtered_data:
                sheet.append([
                    data['instance'],
                    data['total_orders'],
                    data['delivery_orders_count'],
                    data['pickup_orders_count'],
                    data['dinein_orders_count'],
                    data['cash_payment_orders_count'],
                    data['online_payment_orders_count'],
                    data['card_payment_orders_count'],
                ])

            sheet.append([''])

            if language == "English":
                sheet.append(['Tax Collection Data'])

                sheet.append([
                    'Instance',
                    'Total Tax Collected',
                    'Tax Collection from Delivery',
                    'Tax Collection from Pickup',
                    'Tax Collection from DineIn',
                    'Tax Collection from Cash Payment',
                    'Tax Collection from Online Payment',
                    'Tax Collection from Card Payment',
                ])

            else:
                sheet.append([language_localization['Tax Collection Data']])

                sheet.append([
                    language_localization['Instance'],
                    language_localization['Total Tax Collected'],
                    language_localization['Tax Collection from Delivery'],
                    language_localization['Tax Collection from Pickup'],
                    language_localization['Tax Collection from DineIn'],
                    language_localization['Tax Collection from Cash Payment'],
                    language_localization['Tax Collection from Online Payment'],
                    language_localization['Tax Collection from Card Payment'],
                ])

            for data in filtered_data:
                sheet.append([
                    data['instance'],
                    data['total_tax_collected'],
                    data['tax_collected_from_delivery'],
                    data['tax_collected_from_pickup'],
                    data['tax_collected_from_dinein'],
                    data['tax_collected_from_cash_payment'],
                    data['tax_collected_from_online_payment'],
                    data['tax_collected_from_card_payment'],
                ])
            
            sheet.append([''])

            if language == "English":
                sheet.append(['Revenue Generation Data'])

                sheet.append([
                    'Instance',
                    'Total Revenue Generated',
                    'Revenue from Delivery',
                    'Revenue from Pickup',
                    'Revenue from DineIn',
                    'Revenue from Cash Payment',
                    'Revenue from Online Payment',
                    'Revenue from Card Payment',
                ])

            else:
                sheet.append([language_localization['Revenue Generation Data']])

                sheet.append([
                    language_localization['Instance'],
                    language_localization['Total Revenue Generated'],
                    language_localization['Revenue from Delivery'],
                    language_localization['Revenue from Pickup'],
                    language_localization['Revenue from DineIn'],
                    language_localization['Revenue from Cash Payment'],
                    language_localization['Revenue from Online Payment'],
                    language_localization['Revenue from Card Payment'],
                ])

            for data in filtered_data:
                sheet.append([
                    data['instance'],
                    data['total_revenue'],
                    data['revenue_from_delivery'],
                    data['revenue_from_pickup'],
                    data['revenue_from_dinein'],
                    data['revenue_from_cash_payment'],
                    data['revenue_from_online_payment'],
                    data['revenue_from_card_payment'],
                ])
        
        directory = os.path.join(settings.MEDIA_ROOT, 'Excel Downloads')

        os.makedirs(directory, exist_ok=True) # Create the directory if it doesn't exist inside MEDIA_ROOT

        file_name = f"Footfall_Report_{vendor_id}.xlsx"

        relative_file_path = os.path.join('Excel Downloads', file_name)

        file_path = os.path.join(settings.MEDIA_ROOT, relative_file_path)

        workbook.save(file_path)

        print(f"Excel file '{file_name}' has been created.")
        
        response = "/media/" + relative_file_path
        
        return HttpResponse(response, status=200)
    
    else:
        return Response("Invalid 'download' key", status=status.HTTP_400_BAD_REQUEST)


@api_view(["GET"])
def order_report(request):
    vendor_id = request.GET.get("vendorId")
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    is_download = request.GET.get('download')
    language = request.GET.get('language', 'English')

    if not vendor_id:
        return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id = int(vendor_id)

    except ValueError:
        return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
    
    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor_instance:
        return Response("Vendor does not exist", status=status.HTTP_400_BAD_REQUEST)
    
    if (((not start_date) or (not end_date)) or (start_date > end_date)):
        return Response("Invalid start date or end date", status=status.HTTP_400_BAD_REQUEST)
    
    if is_download not in ("false", "true"):
        return Response("Invalid download parameter", status=status.HTTP_400_BAD_REQUEST)
    
    all_order_details = order_count(start_date, end_date, "all", vendor_id)
    
    delivery_order_details = order_count(start_date, end_date, "delivery", vendor_id)

    pickup_order_details = order_count(start_date, end_date, "pickup", vendor_id)

    dinein_order_details = order_count(start_date, end_date, "dinein", vendor_id)

    online_order_details = order_count(start_date, end_date, "online", vendor_id)

    offline_order_details = order_count(start_date, end_date, "offline", vendor_id)

    order_count_details = {
        "All": all_order_details,
        "Delivery": delivery_order_details,
        "Pickup": pickup_order_details,
        "DineIn": dinein_order_details,
        "Online": online_order_details,
        "Offline": offline_order_details
    }
    
    if is_download == "false":    
        return JsonResponse(order_count_details)
    
    elif is_download == "true":
        platform = Platform.objects.filter(Name__in=('Mobile App', 'Website'), isActive=True, VendorId=vendor_id).first()

        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')

        formatted_start_date = start_date.strftime('%d-%m-%Y')
        formatted_end_date = end_date.strftime('%d-%m-%Y')

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        if language == "English":
            sheet.append(['Start Date', f'{formatted_start_date}'])
            sheet.append(['End Date', f'{formatted_end_date}'])

        else:
            sheet.append([language_localization['Start Date'], f'{formatted_start_date}'])
            sheet.append([language_localization['End Date'], f'{formatted_end_date}'])

        sheet.append([''])

        if language == "English":
            sheet.append(["Category", "Total Orders", "Complete Orders", "Cancelled Orders", "Processing Orders"])
        
        else:
            sheet.append([
                language_localization["Category"],
                language_localization["Total Orders"],
                language_localization["Complete Orders"],
                language_localization["Cancelled Orders"],
                language_localization["Processing Orders"]
            ])

        for category, orders_info in order_count_details.items():
            if category == "Online" and not platform:
                pass
            
            else:
                sheet.append([
                    category,
                    orders_info.get("total_orders"),
                    orders_info.get("complete_orders"),
                    orders_info.get("cancelled_orders"),
                    orders_info.get("processing_orders")
                ])

        directory = os.path.join(settings.MEDIA_ROOT, 'Excel Downloads')
        os.makedirs(directory, exist_ok=True) # Create the directory if it doesn't exist inside MEDIA_ROOT

        file_name = f"Order_Count_Report_{vendor_id}.xlsx"

        relative_file_path = os.path.join('Excel Downloads', file_name)

        file_path = os.path.join(settings.MEDIA_ROOT, relative_file_path)

        workbook.save(file_path)

        print(f"Excel file '{file_name}' has been created.")
        
        response = "/media/" + relative_file_path
        
        return HttpResponse(response, status=200)


@api_view(["GET"])
def cancel_order_report(request):
    vendor_id = request.GET.get("vendorId")
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    order_type = request.GET.get('type')
    is_download = request.GET.get('download')
    top_number = request.GET.get('top')
    sort_by = request.GET.get('sort')
    language = request.GET.get('language', 'English')

    if not vendor_id:
        return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id = int(vendor_id)

    except ValueError:
        return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
    
    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor_instance:
        return Response("Vendor does not exist", status=status.HTTP_400_BAD_REQUEST)
    
    if (((not start_date) or (not end_date)) or (start_date > end_date)):
        return Response("Invalid start date or end date", status=status.HTTP_400_BAD_REQUEST)
    
    if is_download not in ("false", "true"):
        return Response("Invalid download parameter", status=status.HTTP_400_BAD_REQUEST)

    if order_type not in ("all", "delivery", "pickup", "dinein", "online", "offline"):
        return Response("Invalid type parameter", status=status.HTTP_400_BAD_REQUEST)

    try:
        top_number = int(top_number)

    except ValueError:
        return Response("Invalid top parameter", status=status.HTTP_400_BAD_REQUEST)
    
    if sort_by not in ("ascending", "descending"):
        return Response("Invalid sort parameter", status=status.HTTP_400_BAD_REQUEST)
    
    if order_type == "delivery":
        order_type_code = order_type_number["Delivery"]

    elif order_type == "pickup":
        order_type_code = order_type_number["Pickup"]

    elif order_type == "dinein":
        order_type_code = order_type_number["Dinein"]
    
    platform = Platform.objects.filter(Name__in = ('Mobile App', 'Website'), isActive = True, VendorId = vendor_id).first()
    
    cancelled_orders = KOMSOrder.objects.filter(
        master_order__Status = master_order_status_number["Canceled"],
        master_order__OrderDate__date__range = (start_date, end_date),
        master_order__vendorId = vendor_id,
        vendorId = vendor_id
    )

    most_cancelled_products = Order_content.objects.filter(
        status = 5,
        orderId__order_status = 5,
        orderId__master_order__Status = master_order_status_number["Canceled"],
        orderId__master_order__OrderDate__date__range = (start_date, end_date),
        orderId__master_order__vendorId = vendor_id,
        orderId__vendorId = vendor_id
    )
    
    if order_type == "all":
        pass
    
    elif order_type == "online":
        if platform:
            cancelled_orders = cancelled_orders.filter(master_order__platform=platform.pk)

            most_cancelled_products = most_cancelled_products.filter(orderId__master_order__platform=platform.pk)
        
        else:
            return Response("Contact you administrator to activate the platform", status=status.HTTP_400_BAD_REQUEST)
            
    elif order_type == "offline":
        if platform:
            cancelled_orders = cancelled_orders.exclude(master_order__platform=platform.pk)

            most_cancelled_products = most_cancelled_products.exclude(orderId__master_order__platform=platform.pk)

        else:
            return Response("Contact you administrator to activate the platform", status=status.HTTP_400_BAD_REQUEST)

    else:
        cancelled_orders = cancelled_orders.filter(master_order__orderType=order_type_code)

        most_cancelled_products = most_cancelled_products.filter(orderId__master_order__orderType=order_type_code)

    cancelled_orders_count = len(list(cancelled_orders.values_list('master_order', flat=True).distinct()))

    cancelled_products_count = most_cancelled_products.count()
    
    most_cancelled_products = most_cancelled_products.values('SKU').distinct() \
                            .annotate(cancel_count=Count('SKU')) \
                            .filter(cancel_count__gt=0) \
                            .order_by('-cancel_count')[:top_number]

    if sort_by == "descending":
        pass

    elif sort_by == "ascending":
        most_cancelled_products = sorted(most_cancelled_products, key=itemgetter('cancel_count'))

    cancelled_product_details = []
    
    for product in most_cancelled_products:
        product_info = Product.objects.filter(PLU=product["SKU"], vendorId=vendor_id).first()

        images = ProductImage.objects.filter(product=product_info.pk)

        image_list = []
        
        for image in images:
            if image.url:
                image_list.append(image.url)

        product_name = ""

        if language == "English":
            product_name = product_info.productName

        else:
            product_name = product_info.productName_locale

        cancelled_product_details.append({
            'product_id': product_info.pk,
            'product_name': product_name,
            'image': image_list[0] if len(image_list)!=0 else 'https://www.stockvault.net/data/2018/08/31/254135/preview16.jpg',
            'price': product_info.productPrice,
            'cancel_count': product['cancel_count'],
            'cancelled_estimated_sale': product['cancel_count'] * product_info.productPrice
        })

    estimated_revenue_from_cancelled_orders = 0.0

    for product_detail in cancelled_product_details:
        estimated_revenue_from_cancelled_orders = estimated_revenue_from_cancelled_orders + product_detail["cancelled_estimated_sale"]

    if is_download == "false":    
        return JsonResponse({
            "cancelled_orders_count": cancelled_orders_count,
            "cancelled_products_count": cancelled_products_count,
            "cancelled_estimated_revenue": estimated_revenue_from_cancelled_orders,
            "cancelled_products": cancelled_product_details
        })
    
    elif is_download == "true":
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')

        formatted_start_date = start_date.strftime('%d-%m-%Y')
        formatted_end_date = end_date.strftime('%d-%m-%Y')

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        if language == "English":
            sheet.append(['Start Date', f'{formatted_start_date}'])
            sheet.append(['End Date', f'{formatted_end_date}'])
            sheet.append(['Order Type', f'{order_type}'])
            sheet.append([''])

            sheet.append(["Cancelled Orders", "Cancelled Products", "Loss Made"])
            sheet.append([cancelled_orders_count, cancelled_products_count, estimated_revenue_from_cancelled_orders])

            sheet.append([''])

            sheet.append(['Cancelled Product Details'])
            sheet.append(['Top', top_number])
            sheet.append(['Sorted by', sort_by])

            sheet.append([''])

            sheet.append(['Product Name', 'Quantity Cancelled', 'Unit Price', 'Estimated Revenue'])

        else:
            sheet.append([language_localization['Start Date'], f'{formatted_start_date}'])
            sheet.append([language_localization['End Date'], f'{formatted_end_date}'])
            sheet.append([language_localization['Order Type'], f'{order_type}'])
            sheet.append([''])

            sheet.append([language_localization["Cancelled Orders"], language_localization["Cancelled Products"], language_localization["Loss Made"]])
            sheet.append([cancelled_orders_count, cancelled_products_count, estimated_revenue_from_cancelled_orders])

            sheet.append([''])

            sheet.append([language_localization['Cancelled Product Details']])
            sheet.append([language_localization['Top'], top_number])
            sheet.append([language_localization['Sorted by'], sort_by])

            sheet.append([''])

            sheet.append([
                language_localization['Product Name'],
                language_localization['Quantity Cancelled'],
                language_localization['Unit Price'],
                language_localization['Estimated Revenue']
            ])

        for product_detail in cancelled_product_details:
            sheet.append([
                product_detail["product_name"],
                product_detail["cancel_count"],
                product_detail["price"],
                product_detail["cancelled_estimated_sale"]
            ])

        directory = os.path.join(settings.MEDIA_ROOT, 'Excel Downloads')
        os.makedirs(directory, exist_ok=True) # Create the directory if it doesn't exist inside MEDIA_ROOT

        file_name = f"Cancel_Order_Report_{vendor_id}.xlsx"

        relative_file_path = os.path.join('Excel Downloads', file_name)

        file_path = os.path.join(settings.MEDIA_ROOT, relative_file_path)

        workbook.save(file_path)

        print(f"Excel file '{file_name}' has been created.")
        
        response = "/media/" + relative_file_path
        
        return HttpResponse(response, status=200)


@api_view(["GET"])
def pincode_report(request):
    vendor_id = request.GET.get("vendorId")
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    order_type = request.GET.get('type')
    is_download = request.GET.get('download')
    top_number = request.GET.get('top')
    sort_by = request.GET.get('sort')
    language = request.GET.get('language', 'English')

    if not vendor_id:
        return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id = int(vendor_id)
    except ValueError:
        return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
    
    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor_instance:
        return Response("Vendor does not exist", status=status.HTTP_400_BAD_REQUEST)
    
    if (((not start_date) or (not end_date)) or (start_date > end_date)):
        return Response("Invalid start date or end date", status=status.HTTP_400_BAD_REQUEST)
    
    if is_download not in ("false", "true"):
        return Response("Invalid download parameter", status=status.HTTP_400_BAD_REQUEST)

    if order_type not in ("all", "delivery", "pickup", "dinein", "online", "offline"):
        return Response("Invalid type parameter", status=status.HTTP_400_BAD_REQUEST)

    try:
        top_number = int(top_number)
    except ValueError:
        return Response("Invalid top parameter", status=status.HTTP_400_BAD_REQUEST)
    
    if sort_by not in ("ascending", "descending"):
        return Response("Invalid sort parameter", status=status.HTTP_400_BAD_REQUEST)
    
    if order_type == "delivery":
        order_type_code = order_type_number["Delivery"]

    elif order_type == "pickup":
        order_type_code = order_type_number["Pickup"]

    elif order_type == "dinein":
        order_type_code = order_type_number["Dinein"]
    
    platform = Platform.objects.filter(Name__in=('Mobile App', 'Website'), isActive=True, VendorId=vendor_id).first()

    order_items = Order_content.objects.filter(
        orderId__order_status = 10,
        orderId__master_order__Status = master_order_status_number["Completed"],
        orderId__master_order__orderpayment__status = True,
        orderId__master_order__OrderDate__date__range = (start_date, end_date),
        orderId__master_order__customerId__address__isnull = False,
        orderId__master_order__customerId__address__is_selected = True,
        orderId__vendorId = vendor_id
    ).exclude(status = 5)
    
    if order_type == "all":
        pass
    
    elif order_type == "online":
        if platform:
            order_items = order_items.filter(orderId__master_order__platform=platform.pk)
        
        else:
            return Response("Contact you administrator to activate the platform", status=status.HTTP_400_BAD_REQUEST)
            
    elif order_type == "offline":
        if platform:
            order_items = order_items.exclude(orderId__master_order__platform=platform.pk)
        
        else:
            return Response("Contact you administrator to activate the platform", status=status.HTTP_400_BAD_REQUEST)
            
    else:
        order_items = order_items.filter(orderId__master_order__orderType=order_type_code)

    top_pincodes = order_items.values('orderId__master_order__customerId__address__zipcode').distinct() \
                    .annotate(order_count=Count('orderId__master_order__pk', distinct=True)) \
                    .filter(order_count__gt=0) \
                    .order_by('-order_count')[:top_number]
    
    if sort_by == "ascending":
        top_pincodes = sorted(top_pincodes, key=itemgetter('order_count'))

    pincode_list = []

    # Initialize pgeocode
    nomi = pgeocode.Nominatim("in")
    
    for details in top_pincodes:
        order_items_by_area = order_items.filter(
            orderId__master_order__customerId__address__zipcode=details["orderId__master_order__customerId__address__zipcode"]
        )

        master_order_ids = set(order_items_by_area.values_list('orderId__master_order', flat=True))

        subtotal_sum = Order.objects.filter(pk__in=master_order_ids).aggregate(subtotal_sum=Sum('subtotal'))['subtotal_sum'] or 0.0

        most_ordered_items = order_items_by_area.values('SKU').distinct() \
                            .annotate(item_count=Sum('quantity')) \
                            .filter(item_count__gt=0) \
                            .order_by('-item_count')[:10]

        product_details = []
        
        for product in most_ordered_items:
            product_info = Product.objects.filter(PLU=product["SKU"], vendorId=vendor_id).first()

            images = ProductImage.objects.filter(product=product_info.pk)

            image_list = []
            
            for image in images:
                if image.url:
                    image_list.append(image.url)

            product_name = ""

            if language == "English":
                product_name = product_info.productName

            else:
                product_name = product_info.productName_locale

            product_details.append({
                'product_id': product_info.pk,
                'product_name': product_name,
                'image': image_list[0] if len(image_list)!=0 else 'https://www.stockvault.net/data/2018/08/31/254135/preview16.jpg',
                'price': product_info.productPrice,
                'quantity_sold': product['item_count'],
                'revenue': product['item_count'] * product_info.productPrice
            })

        # url = f'https://api.postalpincode.in/pincode/{details["customerId__address__zipcode"]}'

        # response = requests.get(url)
        # data = response.json()
        
        # locality_name = ""

        # if data and data[0]['Status'] == 'Success':
        #     locality_name = data[0]['PostOffice'][0]['Name']

        location = nomi.query_postal_code(details["orderId__master_order__customerId__address__zipcode"])

        locality = ""

        if (not location.empty) and (str(location["place_name"]) != 'nan'):
            locality = location["place_name"].split(", ")[0]
            
        pincode_list.append({
            "pincode": details["orderId__master_order__customerId__address__zipcode"],
            "locality_name": locality,
            "total_orders": details["order_count"],
            "revenue": round(subtotal_sum, 2),
            "products": product_details
        })

    if is_download == "false":    
        return JsonResponse({"pincodes": pincode_list})
    
    elif is_download == "true":
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')

        formatted_start_date = start_date.strftime('%d-%m-%Y')
        formatted_end_date = end_date.strftime('%d-%m-%Y')

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        if language == "English":
            sheet.append(['Start Date', f'{formatted_start_date}'])
            sheet.append(['End Date', f'{formatted_end_date}'])
            sheet.append(['Order Type', f'{order_type}'])
            sheet.append([''])

            sheet.append(["Pincode", "Locality", "Orders", "Revenue", "Most Ordered Products"])

        else:
            sheet.append([language_localization['Start Date'], f'{formatted_start_date}'])
            sheet.append([language_localization['End Date'], f'{formatted_end_date}'])
            sheet.append([language_localization['Order Type'], f'{order_type}'])
            sheet.append([''])

            sheet.append([
                language_localization["Pincode"],
                language_localization["Locality"],
                language_localization["Orders"],
                language_localization["Revenue"],
                language_localization["Most Ordered Products"]
            ])
        
        for info in pincode_list:
            list_of_items = ''
            
            for product in info.get('products', []):
                if list_of_items:
                    list_of_items = list_of_items + ", " + product["product_name"]

                else:
                    list_of_items = product["product_name"]

            sheet.append([info["pincode"], info["locality_name"], info["total_orders"], info["revenue"], list_of_items])

        directory = os.path.join(settings.MEDIA_ROOT, 'Excel Downloads')
        os.makedirs(directory, exist_ok=True) # Create the directory if it doesn't exist inside MEDIA_ROOT

        file_name = f"Pincode_Report_{vendor_id}.xlsx"

        relative_file_path = os.path.join('Excel Downloads', file_name)

        file_path = os.path.join(settings.MEDIA_ROOT, relative_file_path)

        workbook.save(file_path)

        print(f"Excel file '{file_name}' has been created.")
        
        response = "/media/" + relative_file_path
        
        return HttpResponse(response, status=200)


@api_view(["GET"])
def get_delivery_settings(request):
    vendor_id = request.GET.get('vendor')

    if not vendor_id:
        return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id = int(vendor_id)

    except ValueError:
        return JsonResponse({"message": "Invalid vendor ID"}, status=status.HTTP_400_BAD_REQUEST)
    
    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor_instance:
        return JsonResponse({"message": "Vendor does not exist"}, status=status.HTTP_400_BAD_REQUEST)

    delivery_setting = POSSetting.objects.filter(vendor=vendor_id).first()

    if not delivery_setting:
        return JsonResponse({"message": "POS setting not created for the Vendor"}, status=status.HTTP_400_BAD_REQUEST)
    
    return JsonResponse({
        "message": "",
        "kilometer_limit": delivery_setting.delivery_kilometer_limit,
        "delivery_charges": delivery_setting.delivery_charges_for_kilometer_limit
    })


@api_view(["POST"])
def update_delivery_settings(request):
    vendor_id = request.GET.get('vendor')

    if not vendor_id:
        return JsonResponse("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id = int(vendor_id)

    except ValueError:
        return JsonResponse({"message": "Invalid vendor ID"}, status=status.HTTP_400_BAD_REQUEST)
    
    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor_instance:
        return JsonResponse({"message": "Vendor does not exist"}, status=status.HTTP_400_BAD_REQUEST)
    
    kilometer_limit = request.data.get('kilometer_limit')
    delivery_charges = request.data.get('delivery_charges')

    if (kilometer_limit is None) or (delivery_charges is None):
        return JsonResponse({"message": "Invalid request data"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        kilometer_limit = int(kilometer_limit)
        delivery_charges = int(delivery_charges)

    except ValueError:
        return JsonResponse({"message": "Invalid request data"}, status=status.HTTP_400_BAD_REQUEST)
    
    delivery_settting_instance = POSSetting.objects.filter(vendor=vendor_id).first()

    if not delivery_settting_instance:
        return JsonResponse({"message": "POS setting not created for the Vendor"}, status=status.HTTP_400_BAD_REQUEST)

    delivery_settting_instance.delivery_kilometer_limit = kilometer_limit
    delivery_settting_instance.delivery_charges_for_kilometer_limit = delivery_charges

    delivery_settting_instance.save()
    
    return JsonResponse({
        "message": "",
        "kilometer_limit": delivery_settting_instance.delivery_kilometer_limit,
        "delivery_charges": delivery_settting_instance.delivery_charges_for_kilometer_limit
    })


import zipfile
@api_view(["POST"])
def product_excel_upload_for_pos(request):
    try:
        if 'excel_file' not in request.FILES:
            return Response("No file uploaded", status=status.HTTP_400_BAD_REQUEST)

        vendor_id = request.POST.get('vendor')

        if not vendor_id:
            return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
        
        try:
            vendor_id = int(vendor_id)
        except ValueError:
            return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
        
        vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

        if not vendor_instance:
            return Response("Vendor does not exist", status=status.HTTP_400_BAD_REQUEST)

        uploaded_file = request.FILES['excel_file']

        directory = os.path.join(settings.MEDIA_ROOT, 'Product Details Excel')
        os.makedirs(directory, exist_ok=True) # Create the directory if it doesn't exist inside MEDIA_ROOT

        file_name = f"{uploaded_file.name.split('.')[0]}_Vendor{vendor_id}.{uploaded_file.name.split('.')[-1]}"

        relative_file_path = os.path.join('Product Details Excel', file_name)

        file_path = os.path.join(settings.MEDIA_ROOT, relative_file_path)

        print(file_path)

        with default_storage.open(file_path, 'wb+') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)

        workbook = openpyxl.load_workbook(file_path)
        sheet_title = workbook.active.title
        
        file_status, response = process_product_excel(file_path, sheet_title, vendor_id)

        if (file_status == 1) and (response == None):
            return Response("Excel file uploaded successfully")

        elif (file_status == 1) and (response != None):
            return HttpResponse(response, status=status.HTTP_201_CREATED)
        
        elif file_status == 2:
            return Response(response, status=status.HTTP_202_ACCEPTED)

        elif file_status == 3:
            return Response(response, status=status.HTTP_203_NON_AUTHORITATIVE_INFORMATION)
        
        elif file_status == 4:
            return Response(response, status=status.HTTP_205_RESET_CONTENT)

        else:
            return Response(f"Excel file upload failed: {response}", status=status.HTTP_400_BAD_REQUEST)

    except zipfile.BadZipFile:
        return Response("Excel file corrupt", status=status.HTTP_206_PARTIAL_CONTENT)
    
    except Exception as e:
        return Response(f"{str(e)}", status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
def download_product_excel_upload_template(request):
    try:
        vendor_id = request.GET.get('vendor')

        if not vendor_id:
            return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
        
        try:
            vendor_id = int(vendor_id)
            
        except ValueError:
            return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
        
        vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

        if not vendor_instance:
            return Response("Vendor does not exist", status=status.HTTP_400_BAD_REQUEST)

        file_name = f"Product_Excel_Template.xlsx"

        relative_file_path = os.path.join('Excel Downloads', file_name)

        file_path = os.path.join(settings.MEDIA_ROOT, relative_file_path)

        response = "/media/" + relative_file_path

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        if not vendor_instance.secondary_language:
            sheet.append([
                "Category Station", "Category Name", "Category SKU", "Category Description", "Is Category Active (yes/no)", "Category Image",
                "Product Name", "Product SKU", "Product Description", "Tag", "Product Price", "Is Product Active (yes/no)", "Product Image",
                "Modifier Group Name", "Modifier Group SKU", "Modifier Group Description", "Modifier Group Min", "Modifier Group Max", "Is Modifier Group Active (yes/no)",
                "Modifier Name", "Modifier SKU", "Modifier Description", "Modifier Price", "Modifier Active (yes/no)", "Modifier Image"
            ])

        else:
            sheet.append([
                "Category Name", "Category Name (Locale)", "Category SKU", "Category Description", "Category Description (Locale)", "Is Category Active (yes/no)", "Category Image",
                "Product Name", "Product Name (Locale)", "Product SKU", "Product Description", "Product Description (Locale)", "Tag", "Product Price", "Is Product Active (yes/no)", "Product Image",
                "Modifier Group Name", "Modifier Group Name (Locale)", "Modifier Group SKU", "Modifier Group Description", "Modifier Group Description (Locale)", "Modifier Group Min", "Modifier Group Max", "Is Modifier Group Active (yes/no)",
                "Modifier Name", "Modifier Name (Locale)", "Modifier SKU", "Modifier Description", "Modifier Description (Locale)", "Modifier Price", "Modifier Active (yes/no)", "Modifier Image"
            ])
            
        directory = os.path.join(settings.MEDIA_ROOT, 'Excel Downloads')
        os.makedirs(directory, exist_ok=True) # Create the directory if it doesn't exist inside MEDIA_ROOT

        workbook.save(file_path)

        print(f"Excel file '{file_name}' has been created.")
        
        return HttpResponse(response, status=status.HTTP_200_OK)
    
    except Exception as e:
        print(e)
        return HttpResponse("Something went wrong", status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
def download_product_data_excel(request):
    try:
        vendor_id = request.GET.get('vendor')

        if not vendor_id:
            return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
        
        try:
            vendor_id = int(vendor_id)

        except ValueError:
            return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
        
        vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

        if not vendor_instance:
            return Response("Vendor does not exist", status=status.HTTP_400_BAD_REQUEST)
    
        categories = ProductCategory.objects.filter(vendorId=vendor_id).order_by("categoryName", "-is_active")

        if not categories.exists():
            return Response("No categories found", status=status.HTTP_404_NOT_FOUND)
        
        # Get the IP addresses associated with the hostname
        host_ip_list = socket.gethostbyname_ex(socket.gethostname())[2]

        # Filter out localhost addresses (those starting with "127.")
        filtered_ip_list = []
        
        for ip in host_ip_list:
            if not ip.startswith("127."):
                filtered_ip_list.append(ip)

        # Take the first non-localhost IP address if it exists
        if filtered_ip_list:
            server_ip = filtered_ip_list[0]

        else:
            server_ip = None

        # Create a socket to determine the external IP address by connecting to a public DNS server
        if not server_ip:
            try:
                soc = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)

                soc.connect(('8.8.8.8', 53))

                server_ip = soc.getsockname()[0]

                soc.close()

            except Exception as e:
                server_ip = None

        # Create a new Excel workbook and select the active sheet
        workbook = openpyxl.Workbook()
        
        sheet = workbook.active

        sheet.title = "Sheet1"
        
        if not vendor_instance.secondary_language:
            sheet.append([
                "Category Station", "Category Name", "Category SKU", "Category Description", "Is Category Active (yes/no)", "Category Image",
                "Product Name", "Product SKU", "Product Description", "Tag", "Product Price", "Is Product Active (yes/no)", "Product Image",
                "Modifier Group Name", "Modifier Group SKU", "Modifier Group Description", "Modifier Group Min", "Modifier Group Max", "Is Modifier Group Active (yes/no)",
                "Modifier Name", "Modifier SKU", "Modifier Description", "Modifier Price", "Modifier Active (yes/no)", "Modifier Image"
            ])

        else:
            sheet.append([
                "Category Station", "Category Name", "Category Name (Locale)", "Category SKU", "Category Description", "Category Description (Locale)", "Is Category Active (yes/no)", "Category Image",
                "Product Name", "Product Name (Locale)", "Product SKU", "Product Description", "Product Description (Locale)", "Tag", "Product Price", "Is Product Active (yes/no)", "Product Image",
                "Modifier Group Name", "Modifier Group Name (Locale)", "Modifier Group SKU", "Modifier Group Description", "Modifier Group Description (Locale)", "Modifier Group Min", "Modifier Group Max", "Is Modifier Group Active (yes/no)",
                "Modifier Name", "Modifier Name (Locale)", "Modifier SKU", "Modifier Description", "Modifier Description (Locale)", "Modifier Price", "Modifier Active (yes/no)", "Modifier Image"
            ])
        
        for category in categories:
            category_product_joint = ProductCategoryJoint.objects.filter(
                category=category,
                vendorId=vendor_id
            ).order_by("product__productName", "-product__active")

            if not category_product_joint.exists():
                return Response(f"No products found for category: {category.categoryName}", status=status.HTTP_404_NOT_FOUND)

            is_category_active = "no"
                        
            if category.is_active == True:
                is_category_active = "yes"
            
            category_image = category.categoryImageUrl if category.categoryImageUrl else ""
            
            for product_info in category_product_joint:
                is_product_active = "no"

                if product_info.product.active == True:
                    is_product_active = "yes"

                product_image = ""
                
                product_picture = ProductImage.objects.filter(product=product_info.product.pk, vendorId=vendor_id).first()

                if product_picture:
                    product_image = product_picture.url

                product_modifier_group_joint = ProductAndModifierGroupJoint.objects.filter(
                    product=product_info.product.pk,
                    vendorId=vendor_id
                ).order_by("modifierGroup__name", "-modifierGroup__active")

                category_description = category.categoryDescription if category.categoryDescription else ""
                product_description = product_info.product.productDesc if product_info.product.productDesc else ""
                category_description_locale = category.categoryDescription_locale if category.categoryDescription_locale else ""
                product_description_locale = product_info.product.productDesc_locale if product_info.product.productDesc_locale else ""
                
                if not product_modifier_group_joint.exists():
                    if not vendor_instance.secondary_language:
                        sheet.append([
                            f"{category.categoryStation.station_name}", f"{category.categoryName}", f"{category.categoryPLU}", f"{category_description}", f"{is_category_active}", f"{category_image}",
                            f"{product_info.product.productName}", f"{product_info.product.PLU}", f"{product_description}",
                            f"{product_info.product.tag}", f"{product_info.product.productPrice}", f"{is_product_active}", f"{product_image}",
                        ])

                    else:
                        sheet.append([
                            f"{category.categoryStation.station_name}", f"{category.categoryName}", f"{category.categoryName_locale}", f"{category.categoryPLU}", f"{category_description}", f"{category_description_locale}", f"{is_category_active}", f"{category_image}",
                            f"{product_info.product.productName}", f"{product_info.product.productName_locale}", f"{product_info.product.PLU}", f"{product_description}", f"{product_description_locale}",
                            f"{product_info.product.tag}", f"{product_info.product.productPrice}", f"{is_product_active}", f"{product_image}",
                        ])

                else:
                    for modifier_group_info in product_modifier_group_joint:
                        modifier_group_modifier_joint = ProductModifierAndModifierGroupJoint.objects.filter(
                            modifierGroup=modifier_group_info.modifierGroup.pk,
                            vendor=vendor_id
                        ).order_by("modifier__modifierName", "-modifier__active")

                        if not modifier_group_modifier_joint.exists():
                            return Response(f"Modifier group {modifier_group_info.modifierGroup.name} has no modifiers", status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                        
                        is_modifier_group_active = "no"
                        
                        if modifier_group_info.modifierGroup.active == True:
                            is_modifier_group_active = "yes"
                        
                        for modifier_info in modifier_group_modifier_joint:
                            modifier_group_description = modifier_group_info.modifierGroup.modifier_group_description if modifier_group_info.modifierGroup.modifier_group_description else ""
                            modifier_description = modifier_info.modifier.modifierDesc if modifier_info.modifier.modifierDesc else ""

                            is_modifier_active = "no"
                        
                            if modifier_info.modifier.active == True:
                                is_modifier_active = "yes"

                            modifier_image = ""
                
                            if modifier_info.modifier.modifierImg:
                                modifier_image = modifier_info.modifier.modifierImg

                            if not vendor_instance.secondary_language:
                                sheet.append([
                                    f"{category.categoryStation.station_name}", f"{category.categoryName}", f"{category.categoryPLU}", f"{category_description}", f"{is_category_active}", f"{category_image}",
                                    f"{product_info.product.productName}", f"{product_info.product.PLU}", f"{product_description}",
                                    f"{product_info.product.tag}", f"{product_info.product.productPrice}", f"{is_product_active}", f"{product_image}",
                                    f"{modifier_group_info.modifierGroup.name}", f"{modifier_group_info.modifierGroup.PLU}", f"{modifier_group_description}",
                                    f"{modifier_group_info.modifierGroup.min}", f"{modifier_group_info.modifierGroup.max}", f"{is_modifier_group_active}",
                                    f"{modifier_info.modifier.modifierName}", f"{modifier_info.modifier.modifierPLU}", f"{modifier_description}",
                                    f"{modifier_info.modifier.modifierPrice}", f"{is_modifier_active}", f"{modifier_image}",
                                ])

                            else:
                                modifier_group_description_locale = modifier_group_info.modifierGroup.modifier_group_description_locale if modifier_group_info.modifierGroup.modifier_group_description_locale else ""
                                modifier_description_locale = modifier_info.modifier.modifierDesc_locale if modifier_info.modifier.modifierDesc_locale else ""

                                sheet.append([
                                    f"{category.categoryStation.station_name}", f"{category.categoryName}", f"{category.categoryName_locale}", f"{category.categoryPLU}", f"{category_description}", f"{category_description_locale}", f"{is_category_active}", f"{category_image}",
                                    f"{product_info.product.productName}", f"{product_info.product.productName_locale}", f"{product_info.product.PLU}", f"{product_description}", f"{product_description_locale}",
                                    f"{product_info.product.tag}", f"{product_info.product.productPrice}", f"{is_product_active}", f"{product_image}",
                                    f"{modifier_group_info.modifierGroup.name}", f"{modifier_group_info.modifierGroup.name_locale}", f"{modifier_group_info.modifierGroup.PLU}", f"{modifier_group_description}", f"{modifier_group_description_locale}",
                                    f"{modifier_group_info.modifierGroup.min}", f"{modifier_group_info.modifierGroup.max}", f"{is_modifier_group_active}",
                                    f"{modifier_info.modifier.modifierName}", f"{modifier_info.modifier.modifierName_locale}", f"{modifier_info.modifier.modifierPLU}", f"{modifier_description}", f"{modifier_description_locale}",
                                    f"{modifier_info.modifier.modifierPrice}", f"{is_modifier_active}", f"{modifier_image}",
                                ])
        
        directory = os.path.join(settings.MEDIA_ROOT, 'Excel Downloads')
        os.makedirs(directory, exist_ok=True) # Create the directory if it doesn't exist inside MEDIA_ROOT

        file_name = f"Products_data_Vendor{vendor_id}.xlsx"

        relative_file_path = os.path.join('Excel Downloads', file_name)

        file_path = os.path.join(settings.MEDIA_ROOT, relative_file_path)

        workbook.save(file_path)

        print(f"Excel file '{file_name}' has been created.")
        
        response = "/media/" + relative_file_path
        
        return HttpResponse(response, status=status.HTTP_200_OK)
    
    except Exception as e:
        print(e)
        return HttpResponse("Something went wrong", status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST"])
def is_platform(request):
    try:
        request_data = request.data
        
        if not request_data:
            return Response("No data in the request body", status=status.HTTP_400_BAD_REQUEST)
        
        required_keys = {"platform", "vendor_id"}

        if not required_keys.issubset(request_data.keys()):
            return Response("Keys in request data should be: 'platform', 'vendor_id'", status=status.HTTP_400_BAD_REQUEST)

        platform_to_check = (request_data.get("platform")).lower()
        vendor_id = request_data.get("vendor_id")

        if not vendor_id:
            return Response("Vendor ID empty", status=status.HTTP_400_BAD_REQUEST)
        
        try:
            vendor_id = int(vendor_id)
        except ValueError:
            return Response("Invalid vendor ID", status=status.HTTP_400_BAD_REQUEST)
        
        vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

        if not vendor_instance:
            return Response("Vendor with given ID does not exist", status=status.HTTP_404_NOT_FOUND)    

        platform_names = ()

        vendor_platforms = Platform.objects.filter(VendorId=vendor_id)

        for platform in vendor_platforms:
            platform_names = platform_names + (platform.Name.lower(),)
        
        if (not platform_to_check) or (platform_to_check not in platform_names):
            return Response("Invalid Platform key", status=status.HTTP_400_BAD_REQUEST)
        
        specific_platform = Platform.objects.filter(Name__iexact=platform_to_check, VendorId=vendor_id).first()

        if not specific_platform:
            return Response("Platform does not exist", status=status.HTTP_404_NOT_FOUND)
        
        current_date = datetime.today().strftime("%Y-%m-%d")
        
        if (platform.isActive == True) and (current_date < platform.expiryDate.strftime("%Y-%m-%d")):
            return Response("True", status=status.HTTP_200_OK)
        
        return Response("False", status=status.HTTP_200_OK)
    
    except Exception as e:
        return Response(f"{str(e)}", status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST"])
def generate_language_translation_excel(request):
    try:
        uploaded_file = request.FILES.get('excel_file')
        language = request.GET.get("language")

        if not uploaded_file:
            return JsonResponse({"message": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        if not language:
            return JsonResponse({"message": "No language specified"}, status=status.HTTP_400_BAD_REQUEST)

        directory = os.path.join(settings.MEDIA_ROOT, 'Language Translation Excel')
        
        os.makedirs(directory, exist_ok=True)
        
        file_name = uploaded_file.name

        relative_file_path = os.path.join('Language Translation Excel', file_name)

        file_path = os.path.join(settings.MEDIA_ROOT, relative_file_path)

        with default_storage.open(file_path, 'wb+') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)

        workbook = openpyxl.load_workbook(file_path)

        sheet_name = workbook.active.title

        if not os.path.exists(file_path):
            return JsonResponse({"message": "File does not exist"}, status=status.HTTP_400_BAD_REQUEST)

        if not file_path.lower().endswith(".xlsx"):
            return JsonResponse({"message": "File format is not .xlsx"}, status=status.HTTP_400_BAD_REQUEST)

        if sheet_name != "Sheet1":
            return JsonResponse({"message": "Sheet name should be 'Sheet1'"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            data = pandas.read_excel(file_path, sheet_name=sheet_name)

        except ValueError as e:
            return JsonResponse({"message": "Wrong file format"}, status=status.HTTP_400_BAD_REQUEST)
        
        existing_column = data.columns.tolist()[0]

        if existing_column != "English":
            return JsonResponse({"message": "Column name should be 'English'"}, status=status.HTTP_400_BAD_REQUEST)

        translator = Translator()

        translations = []

        for string in data['English']:
            translation = translator.translate(string, dest=language).text

            translations.append(translation)
            
        data['Translation'] = translations

        translated_file_name = f"translated_{file_name}"

        translated_relative_file_path = os.path.join('Language Translation Excel', translated_file_name)

        translated_file_path = os.path.join(settings.MEDIA_ROOT, translated_relative_file_path)

        data.to_excel(translated_file_path, index=False)

        translations_dict = {}

        for _, row in data.iterrows():
            english_text = row['English']
            translated_text = row['Translation']

            translations_dict[english_text] = translated_text

        json_data = {'language': translations_dict}

        json_file_name = f"translations_{file_name.split('.')[0]}.json"

        json_relative_file_path = os.path.join('Language Translation Excel', json_file_name)

        json_file_path = os.path.join(settings.MEDIA_ROOT, json_relative_file_path)
        
        with open(json_file_path, 'w', encoding='utf-8') as json_file:
            json.dump(json_data, json_file, ensure_ascii=False, indent=4)

        response_excel_path = os.path.join('/media', translated_relative_file_path).replace('\\', '/')
        response_json_path = os.path.join('/media', json_relative_file_path).replace('\\', '/')

        return JsonResponse({
            "excel_file_path": response_excel_path,
            "json_file_path": response_json_path
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return JsonResponse({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
def get_cash_register_history(request):
    vendor_id = request.GET.get("vendor")
    date = request.GET.get("date")

    if not vendor_id:
        return JsonResponse({"message": "Invalid vendor ID"}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        vendor_id = int(vendor_id)

    except ValueError:
        return JsonResponse({"message": "Invalid vendor ID"}, status=status.HTTP_400_BAD_REQUEST)
    
    vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

    if not vendor_instance:
        return JsonResponse({"message": "Vendor does not exist"}, status=status.HTTP_400_BAD_REQUEST)

    if date:
        cash_register_history = CashRegister.objects.filter(created_at__date=date, vendor=vendor_id).first()

        if cash_register_history:
            return JsonResponse({
                "message": "",
                "balance_while_store_opening": cash_register_history.balance_while_store_opening,
                "balance_while_store_closing": cash_register_history.balance_while_store_closing,
                "created_by": cash_register_history.created_by.pk,
                "created_at": cash_register_history.created_at.astimezone(local_timezone).strftime("%Y-%m-%dT%H:%M:%S"),
                "edited_by": cash_register_history.edited_by.pk,
                "edited_at": cash_register_history.edited_at.astimezone(local_timezone).strftime("%Y-%m-%dT%H:%M:%S"),
            })
        
        return JsonResponse({
            "message": "",
            "balance_while_store_opening": 0.0,
            "balance_while_store_closing": 0.0,
            "created_by": 0,
            "created_at": "",
            "edited_by": 0,
            "edited_at": "",
        })

    cash_register_history_list = []
    
    cash_register_history = CashRegister.objects.filter(vendor=vendor_id)

    for instance in cash_register_history:
        cash_register_history_list.append({
            "balance_while_store_opening": instance.balance_while_store_opening,
            "balance_while_store_closing": instance.balance_while_store_closing,
            "created_by": instance.created_by.pk,
            "created_at": instance.created_at.astimezone(local_timezone).strftime("%Y-%m-%dT%H:%M:%S"),
            "edited_by": instance.edited_by.pk,
            "edited_at": instance.edited_at.astimezone(local_timezone).strftime("%Y-%m-%dT%H:%M:%S"),
        })

    return JsonResponse({"message": "", "history": cash_register_history_list})


@api_view(["POST", "PATCH"])
def register_cash(request):
    try:
        balance = request.data.get("balance")
        user_id = request.data.get("user_id")
        vendor_id = request.data.get("vendor_id")

        if not all((balance, user_id, vendor_id)):
            return JsonResponse({"message": "Invalid request data"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            balance = float(balance)
            user_id = int(user_id)
            vendor_id = int(vendor_id)

        except ValueError:
            return JsonResponse({"message": "Invalid request data"}, status=status.HTTP_400_BAD_REQUEST)
        
        vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

        if not vendor_instance:
            return JsonResponse({"message": "Vendor does not exist"}, status=status.HTTP_400_BAD_REQUEST)
        
        user_instance = CoreUser.objects.filter(pk=user_id, vendor=vendor_id).first()
        
        if not user_instance:
            return JsonResponse({"message": "User does not exist"}, status=status.HTTP_400_BAD_REQUEST)

        if request.method == "POST":
            cash_register_instance = CashRegister.objects.filter(
                created_at__date = datetime.now().date(),
                vendor = vendor_id
            ).first()

            if cash_register_instance:
                return JsonResponse({"message": "You have already registered today's store opening cash balance today"}, status=status.HTTP_400_BAD_REQUEST)

            cash_register_instance = CashRegister.objects.create(
                balance_while_store_opening = balance,
                created_by = user_instance,
                edited_by = user_instance,
                vendor = vendor_instance
            )

        elif request.method == "PATCH":
            cash_register_instance = CashRegister.objects.filter(
                created_at__date = datetime.now().date(),
                vendor = vendor_id
            ).first()

            cash_register_instance.balance_while_store_closing = balance
            cash_register_instance.edited_by = user_instance

            cash_register_instance.save()

        return JsonResponse({
            "message": "",
            "balance_while_store_opening": cash_register_instance.balance_while_store_opening,
            "balance_while_store_closing": cash_register_instance.balance_while_store_closing,
            "created_by": cash_register_instance.created_by.pk,
            "created_at": cash_register_instance.created_at.astimezone(local_timezone).strftime("%Y-%m-%dT%H:%M:%S"),
            "edited_by": cash_register_instance.edited_by.pk,
            "edited_at": cash_register_instance.edited_at.astimezone(local_timezone).strftime("%Y-%m-%dT%H:%M:%S"),
        })
    
    except Exception as e:
        return JsonResponse({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
def get_core_user_categories(request):
    try:
        vendor_id = request.GET.get("vendor")
        search_parameter = request.GET.get("search")

        if not vendor_id:
            return JsonResponse({"message": "Invalid Vendor ID"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            vendor_id = int(vendor_id)

        except ValueError:
            return JsonResponse({"message": "Invalid Vendor ID"}, status=status.HTTP_400_BAD_REQUEST)
        
        vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

        if not vendor_instance:
            return JsonResponse({"message": "Vendor does not exist"}, status=status.HTTP_400_BAD_REQUEST)

        department_wise_categories = get_department_wise_categories(
            vendor_instance = vendor_instance,
            search_parameter = search_parameter
        )
        
        return JsonResponse({
            "message": "",
            "departments": department_wise_categories
        })
    
    except Exception as e:
        return JsonResponse({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST"])
def create_core_user_category(request):
    with transaction.atomic():
        try:
            name = request.data.get("name")
            name_locale = request.data.get("name_locale")
            is_active = request.data.get("is_active")
            department_id = request.data.get("department")
            vendor_id = request.data.get("vendor")

            if (not name) or (not vendor_id):
                return JsonResponse({"message": "Invalid request data"}, status=status.HTTP_400_BAD_REQUEST)

            try:
                vendor_id = int(vendor_id)

                if department_id:
                    department_id = int(department_id)

            except ValueError:
                return JsonResponse({"message": "Invalid Vendor ID or Department ID"}, status=status.HTTP_400_BAD_REQUEST)
            
            vendor_instance = Vendor.objects.filter(pk=vendor_id).first()

            if not vendor_instance:
                return JsonResponse({"message": "Vendor does not exist"}, status=status.HTTP_400_BAD_REQUEST)
            
            if department_id:
                department_instance = Department.objects.filter(pk=department_id, vendor=vendor_id).first()

                if not department_instance:
                    return JsonResponse({"message": "Department does not exist"}, status=status.HTTP_400_BAD_REQUEST)
            
            category_name = f"{name}_{vendor_id}"

            existing_core_user_category = CoreUserCategory.objects.filter(name=category_name, vendor=vendor_id).first()

            if existing_core_user_category:
                return JsonResponse({"message": "Category with this name already exist"}, status=status.HTTP_400_BAD_REQUEST)

            if not name_locale:
                name_locale = name
            
            core_user_category = CoreUserCategory.objects.create(
                name = category_name,
                name_locale = name_locale,
                is_active = is_active,
                department = department_instance if department_id else None,
                vendor = vendor_instance
            )

            department_wise_categories = get_department_wise_categories(vendor_instance = vendor_instance)
        
            return JsonResponse({
                "message": "",
                "departments": department_wise_categories
            }, status = status.HTTP_201_CREATED)
        
        except Exception as e:
            transaction.set_rollback(True)
            
            return JsonResponse({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["PATCH"])
def update_core_user_category(request):
    with transaction.atomic():
        try:
            core_user_category_id = request.data.get("id")
            name = request.data.get("name")
            name_locale = request.data.get("name_locale")
            is_active = request.data.get("is_active")
            department_id = request.data.get("department")
            vendor_id = request.data.get("vendor")

            if not all((name, core_user_category_id, vendor_id)):
                return JsonResponse({"message": "Invalid request data"}, status=status.HTTP_400_BAD_REQUEST)

            try:
                vendor_id = int(vendor_id)
                core_user_category_id = int(core_user_category_id)

                if department_id:
                    department_id = int(department_id)

            except ValueError:
                return JsonResponse({"message": "Invalid Vendor ID, User Category ID or Department ID"}, status=status.HTTP_400_BAD_REQUEST)
            
            vendor_instance = Vendor.objects.filter(pk=vendor_id).first()
            core_user_category_instance = CoreUserCategory.objects.filter(pk=core_user_category_id, vendor=vendor_id).first()

            if (not vendor_instance) or (not core_user_category_instance):
                return JsonResponse({"message": "Vendor or User Category does not exist"}, status=status.HTTP_400_BAD_REQUEST)
            
            if department_id:
                department_instance = Department.objects.filter(pk=department_id, vendor=vendor_id).first()

                if not department_instance:
                    return JsonResponse({"message": "Department does not exist"}, status=status.HTTP_400_BAD_REQUEST)
            
            category_name = f"{name}_{vendor_id}"

            if not name_locale:
                name_locale = name

            core_user_category_instance.name = category_name
            core_user_category_instance.name_locale = name_locale
            core_user_category_instance.is_active = is_active
            core_user_category_instance.department = department_instance if department_id else None

            core_user_category_instance.save()

            department_wise_categories = get_department_wise_categories(vendor_instance = vendor_instance)
        
            return JsonResponse({
                "message": "",
                "departments": department_wise_categories
            })
        
        except Exception as e:
            transaction.set_rollback(True)

            return JsonResponse({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["DELETE"])
def delete_core_user_category(request):
    with transaction.atomic():
        try:
            core_user_category_id = request.data.get("id")
            vendor_id = request.data.get("vendor")

            if not all((core_user_category_id, vendor_id)):
                return JsonResponse({"message": "Invalid request data"}, status=status.HTTP_400_BAD_REQUEST)

            try:
                vendor_id = int(vendor_id)
                core_user_category_id = int(core_user_category_id)

            except ValueError:
                return JsonResponse({"message": "Invalid Vendor ID or User Category ID"}, status=status.HTTP_400_BAD_REQUEST)
            
            vendor_instance = Vendor.objects.filter(pk=vendor_id).first()
            core_user_category_instance = CoreUserCategory.objects.filter(pk=core_user_category_id, vendor=vendor_id).first()

            if (not vendor_instance) or (not core_user_category_instance):
                return JsonResponse({"message": "Vendor or User Category does not exist"}, status=status.HTTP_400_BAD_REQUEST)
            
            core_user_category_instance.delete()
            
            department_wise_categories = get_department_wise_categories(vendor_instance = vendor_instance)
        
            return JsonResponse({
                "message": "",
                "departments": department_wise_categories
            }, status = status.HTTP_200_OK)
        
        except Exception as e:
            transaction.set_rollback(True)

            return JsonResponse({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
